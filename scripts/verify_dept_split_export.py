#!/usr/bin/env python3
"""离线验收：分科室导出布局 / D8 解析逻辑（镜像 Java BillExportLayoutResolver + D8DisplayNameResolver）。"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
THRESHOLD = 1000
DEFAULT_RULE_NAMES = {"标准灭菌计费规则", "标准灭菌计费规则 v2.0"}


def normalize_bill_layout(raw: str | None) -> str:
    if not raw:
        return "auto"
    v = raw.strip().lower()
    if v in ("dept_split", "dept-split", "deptsplit"):
        return "dept_split"
    if v == "combined":
        return "combined"
    return "auto"


def use_dept_split(bill_layout: str, distinct_sheets: int, export_rows: int) -> bool:
    layout = normalize_bill_layout(bill_layout)
    if layout == "dept_split":
        return True
    if layout == "combined":
        return False
    return distinct_sheets > 1 and export_rows <= THRESHOLD


def resolve_d8_auto(hospital: str, rule_name: str | None, plan_name: str | None) -> str:
    if hospital and hospital.strip():
        return hospital.strip()
    if plan_name and plan_name.strip():
        return plan_name.strip()
    if rule_name and rule_name.strip() and rule_name.strip() not in DEFAULT_RULE_NAMES:
        if not rule_name.strip().startswith("标准灭菌计费规则"):
            return rule_name.strip()
    return hospital.strip() if hospital else ""


def run_unit_tests() -> list[tuple[str, bool, str]]:
    results: list[tuple[str, bool, str]] = []

    results.append(
        (
            "dept_split 覆盖 OOM 阈值",
            use_dept_split("dept_split", 30, THRESHOLD + 500),
            "1200 行仍分科室",
        )
    )
    results.append(
        (
            "combined 强制单表",
            not use_dept_split("combined", 30, 2000),
            "多 sheet 仍合并",
        )
    )
    results.append(
        (
            "auto 大行数合并",
            not use_dept_split("auto", 5, THRESHOLD + 1),
            ">1000 行走 combined",
        )
    )
    results.append(
        (
            "auto 小行数分科室",
            use_dept_split("auto", 5, THRESHOLD),
            "<=1000 行多 sheet 分科室",
        )
    )
    results.append(
        (
            "D8 auto 医院名优先",
            resolve_d8_auto("黑龙江省医院（南岗院区）", "标准灭菌计费规则", None)
            == "黑龙江省医院（南岗院区）",
            "",
        )
    )
    results.append(
        (
            "D8 auto 非默认规则名",
            resolve_d8_auto("", "2024年Q1计费规则", None) == "2024年Q1计费规则",
            "医院名为空时用规则名",
        )
    )
    return results


def check_seed() -> tuple[bool, str]:
    seed_path = ROOT / "backend/src/main/resources/billing-seeds/phase-export-dept-split-20260728.json"
    if not seed_path.is_file():
        return False, f"缺少种子 {seed_path.name}"
    data = json.loads(seed_path.read_text(encoding="utf-8"))
    codes = {
        o["code"]
        for o in data.get("exportTemplateOverrides", [])
        for t in o.get("templates", [])
        if t.get("type") == "bill"
    }
    required = {"SHENG-YY-NG", "ZYY-D1", "HRB-2ND", "SHENG-YY-XF"}
    missing = required - codes
    if missing:
        return False, f"种子缺少客户: {missing}"
    for o in data.get("exportTemplateOverrides", []):
        for t in o.get("templates", []):
            cm = t.get("columnMapping") or {}
            if cm.get("billLayout") == "dept_split" and cm.get("d8DisplaySource") != "hospitalName":
                return False, f"{o['code']} dept_split 未配 d8DisplaySource=hospitalName"
    return True, f"种子 OK · {len(codes)} 院 bill 模板"


def analyze_processed_bills() -> list[dict]:
    """统计处理后表格 Sheet 数（ground truth）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [{"hospital": "—", "status": "skip", "detail": "pip install openpyxl"}]

    from batch_june_price_reconciliation import pick_june_pair, HOSPITAL_PAIR_OVERRIDE

    targets = [
        "黑龙江省医院（南岗院区）",
        "黑龙江中医药大学附属第一医院",
        "哈尔滨市第二医院",
    ]
    rows: list[dict] = []
    test_case = ROOT / "测试用例"
    for folder in targets:
        base = test_case / folder
        if not base.is_dir():
            rows.append({"hospital": folder, "status": "skip", "detail": "目录不存在"})
            continue
        _raw, proc, label = pick_june_pair(base)
        if not proc or not proc.is_file():
            rows.append({"hospital": folder, "status": "skip", "detail": f"无处理后表 ({label})"})
            continue
        wb = load_workbook(proc, read_only=True, data_only=True)
        try:
            n = len(wb.sheetnames)
            rows.append(
                {
                    "hospital": folder,
                    "status": "pass" if n > 1 else "warn",
                    "detail": f"{label} · {n} sheets · 预期 dept_split",
                    "sheet_count": n,
                    "file": str(proc.relative_to(ROOT)),
                }
            )
        finally:
            wb.close()
    return rows


def main() -> int:
    print("=== 特色账单分科室导出 · 离线验收 ===\n")

    print("## 1. 单元逻辑（Python 镜像 Java）")
    failed = 0
    for name, ok, note in run_unit_tests():
        mark = "✅" if ok else "❌"
        print(f"  {mark} {name}" + (f" — {note}" if note else ""))
        if not ok:
            failed += 1

    print("\n## 2. 种子文件")
    ok, msg = check_seed()
    print(f"  {'✅' if ok else '❌'} {msg}")

    print("\n## 3. 处理后表格 Sheet 数（铂康 ground truth）")
    sys.path.insert(0, str(ROOT / "scripts"))
    for row in analyze_processed_bills():
        mark = {"pass": "✅", "warn": "🔄", "skip": "⏭"}.get(row["status"], "❌")
        print(f"  {mark} {row['hospital']}: {row['detail']}")

    print("\n## 4. 集成环境")
    import shutil
    import subprocess

    docker_bin = shutil.which("docker")
    if not docker_bin:
        print("  ⏭ docker 不可用，跳过容器检查")
        has_mysql = has_backend = False
    else:
        docker_ps = subprocess.run(
            [docker_bin, "ps", "--format", "{{.Names}}"],
            capture_output=True,
            text=True,
        )
        names = docker_ps.stdout.split()
        has_mysql = "hospital-mysql" in names
        has_backend = "hospital-backend" in names
        print(f"  {'✅' if has_mysql else '❌'} hospital-mysql")
        print(f"  {'✅' if has_backend else '❌'} hospital-backend")
        if has_backend:
            hc = subprocess.run(
                [
                    docker_bin,
                    "exec",
                    "hospital-backend",
                    "curl",
                    "-sS",
                    "http://127.0.0.1:8000/api/v1/base/health",
                ],
                capture_output=True,
                text=True,
            )
            healthy = hc.returncode == 0 and "healthy" in hc.stdout
            print(f"  {'✅' if healthy else '❌'} backend health (容器内 8000)")
        if has_mysql and (ROOT / ".env").is_file():
            pwd = ""
            for line in (ROOT / ".env").read_text(encoding="utf-8").splitlines():
                if line.startswith("MYSQL_ROOT_PASSWORD="):
                    pwd = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
            if pwd:
                mk = subprocess.run(
                    [
                        docker_bin,
                        "exec",
                        "hospital-mysql",
                        "mysql",
                        "-uroot",
                        f"-p{pwd}",
                        "-N",
                        "-B",
                        "hospital",
                        "-e",
                        "SELECT COUNT(*) FROM sys_setting WHERE setting_key='billing_seed_export_dept_split_20260728_v1'",
                    ],
                    capture_output=True,
                    text=True,
                )
                cnt = mk.stdout.strip() if mk.returncode == 0 else "?"
                print(
                    f"  {'✅' if cnt == '1' else '❌' if cnt == '0' else '⏭'} "
                    f"dept_split marker 落库 (count={cnt})"
                )
            else:
                print("  ⏭ dept_split marker：.env 无 MYSQL_ROOT_PASSWORD")

    print("\n## 5. Maven 编译")
    mvn = subprocess.run(
        ["mvn", "-q", "-DskipTests", "compile"],
        cwd=ROOT / "backend",
        capture_output=True,
        text=True,
    )
    compile_ok = mvn.returncode == 0
    if compile_ok:
        print("  ✅ mvn compile（宿主机）")
    else:
        print("  ⚠️ 宿主机 mvn compile 失败（多为 JDK 版本）；尝试 Docker JDK17…")
        docker_mvn = subprocess.run(
            [
                docker_bin or "docker",
                "run",
                "--rm",
                "-v",
                f"{ROOT / 'backend'}:/app",
                "-w",
                "/app",
                "maven:3.9-eclipse-temurin-17",
                "mvn",
                "-q",
                "-DskipTests",
                "compile",
            ],
            capture_output=True,
            text=True,
        )
        compile_ok = docker_mvn.returncode == 0
        print(f"  {'✅' if compile_ok else '❌'} Docker maven:3.9-eclipse-temurin-17 compile")
        if not compile_ok:
            for line in (docker_mvn.stderr or docker_mvn.stdout).splitlines()[-3:]:
                print(f"      {line}")

    out_path = ROOT / "测试用例" / "特色账单分科室导出-验收结果.md"
    # 报告由人工/CI 维护于 测试用例/特色账单分科室导出-验收结果.md
    print(f"\n详细报告: {out_path.relative_to(ROOT)}")
    return 1 if failed or not ok or not compile_ok else 0


def write_report(path: Path, unit_failed: int, seed_ok: bool, compile_ok: bool, docker_ok: bool) -> None:
    path.write_text(
        f"""# 特色账单分科室导出 · 验收结果

> 生成时间：2026-07-27 · 脚本：`scripts/verify_dept_split_export.py`

## 能力状态总览

| 能力项 | 状态 | 说明 |
|--------|:----:|------|
| 代码：billLayout 布局决策 | {'✅' if unit_failed == 0 else '❌'} | 离线逻辑测试 {6 - unit_failed}/6 |
| 代码：D8 医院名优先 | {'✅' if unit_failed == 0 else '❌'} | 同上 |
| 种子 JSON | {'✅' if seed_ok else '❌'} | phase-export-dept-split-20260728.json |
| Maven 编译 | {'✅' if compile_ok else '🚫'} | {'通过' if compile_ok else 'ProductMatchServiceImpl 等阻塞 rebuild'} |
| DB marker 落库 | ⏭ | 需 `docker compose build backend && up -d` 成功后查 sys_setting |
| export-v2 分科室实测 | ⏭ | 依赖 backend 新镜像 + S8 |
| 前端 ExportProfileBanner | ⏭ | 需 rebuild frontend + UI 人工确认 |

## P0 三家 · 处理后表 Sheet 数（预期 export 应对齐）

| 医院 | 预期 layout | 验收方式 |
|------|-------------|----------|
| 黑龙江省医院（南岗院区） | dept_split · 25+ sheets | S8 / 人工导出 |
| 黑龙江中医药大学附属第一医院 | dept_split · 35+ sheets | S8 / 人工导出 |
| 哈尔滨市第二医院 | dept_split · 25+ sheets | S8 / 人工导出 |

## 待执行（环境就绪后）

```bash
docker compose build backend && docker compose up -d backend
./scripts/verify-billing-seed.sh   # 确认 billing_seed_export_dept_split_20260728_v1
python3 scripts/batch_s8_export_compare.py --hospital "黑龙江省医院（南岗院区）"
python3 scripts/batch_s8_export_compare.py --hospital "黑龙江中医药大学附属第一医院"
python3 scripts/batch_s8_export_compare.py --hospital "哈尔滨市第二医院"
docker compose build frontend && docker compose up -d frontend  # 导出向导横幅
```

## 已知阻塞

- **Maven compile 失败**：`ProductMatchServiceImpl` 找不到 Lombok 生成 getter（本地与 Docker build 均失败），导致**无法部署含分科室改动的新 backend 镜像**。
- 当前运行的 `hospital-backend` 为**旧镜像**，API healthy 但**不含** dept_split 逻辑与种子 marker。
""",
        encoding="utf-8",
    )


if __name__ == "__main__":
    raise SystemExit(main())
