#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按真实账单格式 + 包类型/包材对照表 + 尺寸编码一致性 重新生成全部46份测试文件。

修复点：
1. 包类型 ↔ 包材：严格按《包类型与包材对照表》匹配（无纺布/高温纸塑袋/低温纸塑袋 + 灭菌方式）。
2. 包材规格一致性：包名中 /z /w 编码即尺寸规格，包装材料必须与其尺寸一致。
"""
import math
import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "系统内置测试用例")

# ════════════════════════════════════════
# 价格计算工具（保持不变）
# ════════════════════════════════════════
PACK_PRICE = {10: 16.5, 15: 19.5, 20: 22.5, 25: 24.5, 30: 29.5}
ONE_PIECE_PRICE = {10: 22, 15: 25, 20: 28, 25: 30, 30: 35}
LOWTEMP_TIER = {
    1: 22, 2: 44, 3: 66, 4: 88, 5: 88, 6: 110, 7: 132, 8: 154,
    9: 165, 10: 165, 11: 187, 12: 209, 13: 231, 14: 253, 15: 253,
    16: 275, 17: 297, 18: 300, 19: 300, 20: 300, 21: 322, 22: 344,
    23: 366, 24: 388, 25: 388, 26: 410, 27: 432, 28: 454, 29: 465,
    30: 465, 31: 487, 32: 509, 33: 531, 34: 553, 35: 553, 36: 575,
    37: 597, 38: 600, 39: 600, 40: 600,
}

def ceil_div(n, d):
    return math.ceil(n / d)

def fold_price(count, bag_size, with_packaging=True):
    inst = ceil_div(count, 5) * 5.5
    if with_packaging and count <= 10:
        return round(inst + PACK_PRICE[bag_size], 2)
    return round(inst, 2)

# ════════════════════════════════════════
# 尺寸编码 → 包材（新增，保证 /z /w 编码与包材尺寸一致）
# ════════════════════════════════════════
# 纸塑袋 Z 编码 → (尺寸(毫米), 允许温度集合)  温度取自真实账单约定
Z_CODE = {
    "z1026": ("100*260", {"低温", "高温"}),
    "z1029": ("100*290", {"高温"}),
    "z1045": ("100*450", {"低温"}),
    "z1526": ("150*260", {"低温", "高温"}),
    "z1530": ("150*300", {"低温", "高温"}),
    "z1545": ("150*450", {"低温"}),
    "z1560": ("150*600", {"低温"}),
    "z2030": ("200*300", {"低温"}),
    "z2040": ("200*400", {"低温"}),
    "z2044": ("200*440", {"高温"}),
    "z2045": ("200*450", {"低温"}),
    "z2050": ("200*500", {"高温"}),
    "z2060": ("200*600", {"低温", "高温"}),
    "z2530": ("250*300", {"高温"}),
    "z2560": ("250*600", {"低温"}),
    "z3040": ("300*400", {"低温", "高温"}),
    "z7520": ("75*200", {"低温", "高温"}),
    "z7526": ("75*260", {"低温", "高温"}),
    "z7530": ("75*300", {"高温"}),
    "z7550": ("75*500", {"低温"}),
}
# 无纺布 W 编码 → 材料（50g / 60g 两档）
W_SPEC = {
    "w5050": "无纺布-50×50-50g", "w5060": "无纺布-50×50-60g",
    "w6050": "无纺布-60×60-50g", "w6060": "无纺布-60×60-60g",
    "w7050": "无纺布-70×70-50g", "w7060": "无纺布-70×70-60g",
    "w9050": "无纺布-90×90-50g", "w9060": "无纺布-90×90-60g",
    "w12050": "无纺布-120×120-50g", "w12060": "无纺布-120×120-60g",
    "w15050": "无纺布-150×150-50g", "w15060": "无纺布-150×150-60g",
}

# 包类型 → 纸塑袋温度前缀（对照表「灭菌方式/计算规则」列）
TYPE_PP_TEMP = {
    "器械包": None,
    "器械包(低温等离子)": "低温",
    "器械包(ETO)": None,
    "器械包(ZSD)": "高温",
    "单包装包": "高温",
    "单包装包(EO)": "高温",
    "单包装包(老肯低温)": "低温",
    "额外包(纸塑袋)": "高温",
    "额外包(低温等离子)": "低温",
    "额外包(ETO)": "高温",
    "额外包(无纺布)": None,
    "敷料包": None,
    "敷料包(无纺布)": None,
    "敷料包(纸塑袋)": "高温",
    "骨科租赁器械包-带植入物": "高温",
    "骨科租赁器械包-不带植入物": "高温",
}
# 包类型 → 允许的包材类别（PP=纸塑袋, NW=无纺布, NONE=无）
TYPE_ALLOWED = {
    "器械包": {"NW"},
    "器械包(低温等离子)": {"NW", "PP"},
    "器械包(ETO)": {"NW"},
    "器械包(ZSD)": {"NW", "PP"},
    "单包装包": {"NW", "PP"},
    "单包装包(EO)": {"NW", "PP"},
    "单包装包(老肯低温)": {"NW", "PP"},
    "额外包(纸塑袋)": {"PP"},
    "额外包(低温等离子)": {"NW", "PP"},
    "额外包(ETO)": {"NW", "PP"},
    "额外包(无纺布)": {"NW"},
    "敷料包": {"NONE"},
    "敷料包(无纺布)": {"NW"},
    "敷料包(纸塑袋)": {"PP"},
    "骨科租赁器械包-带植入物": {"NW", "PP"},
    "骨科租赁器械包-不带植入物": {"NW", "PP"},
}

# ════════════════════════════════════════
# 测试用例数据结构
# ════════════════════════════════════════
class TestCase:
    def __init__(self, typ, keyword, count, code, price, rule, pack_count=1, mat=None, dept="手术室"):
        self.typ = typ            # 包类型（对照表）
        self.keyword = keyword    # 包名关键字
        self.count = count        # 器械数
        self.code = code          # /z 或 /w 尺寸编码（"" 表示无编码）
        self.price = price
        self.rule = rule
        self.pack_count = pack_count
        self.mat = mat            # 特殊材料覆盖（如"双层袋"）
        self.dept = dept          # 所属科室 sheet

def resolve_material(tc):
    """由尺寸编码推导包装材料（温度取自编码自身约定），保证编码↔尺寸↔温度一致。"""
    if tc.mat:
        return tc.mat
    c = tc.code.lower()
    if c.startswith("z"):
        size, temps = Z_CODE[c]
        prefix = next(iter(temps)) if len(temps) == 1 else TYPE_PP_TEMP[tc.typ]
        return f"{prefix}纸塑袋{size} "
    if c.startswith("w"):
        return W_SPEC[c] + " "
    return ""

def resolve_pack_name(tc):
    if tc.code:
        return f"{tc.keyword}-{tc.count}/{tc.code}"
    return f"{tc.keyword}-{tc.count}"

# ════════════════════════════════════════
# Excel写入器（真实格式）
# ════════════════════════════════════════
# 真实账单样式常量
FONT_NAME = "Droid Sans Fallback"
DATE_NF = '[$-010804]yyyy/m/d'
MONEY_NF = '[$-010804]0.00;(0.00)'
HOSPITAL_FILL = "FF6799AF"   # 医院汇总行填充
DEPT_FILL = "FFBCC3CC"       # 科室汇总行填充
RED_FILL = "FFC7CE"          # 负向错误标注
COL_WIDTHS = {1: 0.2, 2: 0.2, 3: 0.4, 4: 9.3, 5: 8.9, 6: 15.0, 7: 9.4, 8: 24.8,
              9: 0.2, 10: 8.4, 11: 20.5, 12: 6.5, 13: 9.5, 14: 12.2, 15: 4.0,
              16: 8.2, 17: 0.3, 18: 30.0}  # 18=R 为「规则说明」附加列

HEADERS = {4: "发货日期", 5: "发货单号", 6: "类型", 7: "包类别号",
           8: "包名", 11: "包装材料", 12: "包数", 13: "器械数",
           14: "单价", 15: "总价"}

def _thin():
    return Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))

def _f(size, bold=False, color=None):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def _write_sheet(ws, dept_name, cases, hospital_name, is_negative, month, year,
                 first_sheet, error_flags, hospital_totals):
    """写入一个科室 sheet，完全复刻真实账单布局。"""
    thin = _thin()
    n = len(cases)
    # 行号：首 sheet 含医院汇总+子表头（表头8/医院汇总9/子表头10/科室汇总11/数据12起）
    #       后续 sheet 仅科室汇总（表头7/科室汇总8/数据9起）
    if first_sheet:
        header_row, hosp_row, sub_header_row, dept_row, data_start = 8, 9, 10, 11, 12
    else:
        header_row, dept_row, data_start = 7, 8, 9
        hosp_row = sub_header_row = None

    # 标题 C1:I2
    ws.cell(row=1, column=3, value="发货单汇总表-显示包装材料").font = _f(18)
    ws.merge_cells("C1:I2")
    # 日期 B4:H5
    ws.cell(row=4, column=2, value=f"从:{year}/{month}/1 00:00:00 至: {year}/{month}/30 23:59:59.999").font = _f(9)
    ws.merge_cells("B4:H5")

    # 表头
    for col, name in HEADERS.items():
        c = ws.cell(row=header_row, column=col, value=name)
        c.font = _f(10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin
    ws.merge_cells(start_row=header_row, start_column=8, end_row=header_row, end_column=10)  # 包名 H:J
    ws.merge_cells(start_row=header_row, start_column=15, end_row=header_row, end_column=17)  # 总价 O:Q

    # 医院汇总（仅首 sheet）
    if first_sheet:
        ws.cell(row=hosp_row, column=4, value=hospital_name).font = _f(12)
        ws.merge_cells(start_row=hosp_row, start_column=4, end_row=hosp_row, end_column=11)  # D:K
        ws.cell(row=hosp_row, column=12, value=hospital_totals["pack"]).font = _f(12)
        ws.cell(row=hosp_row, column=13, value=hospital_totals["count"]).font = _f(12)
        c14 = ws.cell(row=hosp_row, column=14, value=hospital_totals["price"])
        c14.font = _f(12); c14.number_format = MONEY_NF
        c15 = ws.cell(row=hosp_row, column=15, value=hospital_totals["amount"])
        c15.font = _f(12); c15.number_format = MONEY_NF
        ws.merge_cells(start_row=hosp_row, start_column=15, end_row=hosp_row, end_column=17)
        for col in range(1, 18):
            ws.cell(row=hosp_row, column=col).fill = PatternFill(start_color=HOSPITAL_FILL, end_color=HOSPITAL_FILL, fill_type="solid")
        # 子表头
        for col, name in HEADERS.items():
            c = ws.cell(row=sub_header_row, column=col, value=name)
            c.font = _f(10)
            c.alignment = Alignment(horizontal="center")
            c.border = thin
        ws.merge_cells(start_row=sub_header_row, start_column=8, end_row=sub_header_row, end_column=10)
        ws.merge_cells(start_row=sub_header_row, start_column=15, end_row=sub_header_row, end_column=17)

    # 科室汇总
    dept_totals = {
        "pack": sum(tc.pack_count for tc in cases),
        "count": sum(tc.count for tc in cases),
        "price": round(sum(tc.price for tc in cases), 2),
        "amount": round(sum(tc.price * tc.pack_count for tc in cases), 2),
    }
    ws.cell(row=dept_row, column=4, value=dept_name).font = _f(11)
    ws.merge_cells(start_row=dept_row, start_column=4, end_row=dept_row, end_column=11)
    ws.cell(row=dept_row, column=12, value=dept_totals["pack"]).font = _f(11)
    ws.cell(row=dept_row, column=13, value=dept_totals["count"]).font = _f(11)
    c14 = ws.cell(row=dept_row, column=14, value=dept_totals["price"])
    c14.font = _f(11); c14.number_format = MONEY_NF
    c15 = ws.cell(row=dept_row, column=15, value=dept_totals["amount"])
    c15.font = _f(11); c15.number_format = MONEY_NF
    ws.merge_cells(start_row=dept_row, start_column=15, end_row=dept_row, end_column=17)
    for col in range(1, 18):
        ws.cell(row=dept_row, column=col).fill = PatternFill(start_color=DEPT_FILL, end_color=DEPT_FILL, fill_type="solid")

    # 数据行
    order_no = 1600001
    cat_no = 20300001
    for i, tc in enumerate(cases):
        row = data_start + i
        is_error = error_flags[i]
        price = round(tc.price + 5.0, 2) if is_error else tc.price
        total = round(price * tc.pack_count, 2)

        d = ws.cell(row=row, column=4, value=datetime(year, month, (i % 28) + 1))
        d.number_format = DATE_NF; d.font = _f(10); d.border = thin
        e = ws.cell(row=row, column=5, value=order_no); e.font = _f(10); e.border = thin
        ws.cell(row=row, column=6, value=tc.typ).font = _f(10)
        ws.cell(row=row, column=6).border = thin
        g = ws.cell(row=row, column=7, value=cat_no); g.font = _f(10)
        g.alignment = Alignment(horizontal="center"); g.border = thin
        ws.cell(row=row, column=8, value=resolve_pack_name(tc)).font = _f(10)
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=10)
        ws.cell(row=row, column=11, value=resolve_material(tc)).font = _f(10)
        ws.cell(row=row, column=11).border = thin
        l = ws.cell(row=row, column=12, value=tc.pack_count); l.font = _f(10)
        l.alignment = Alignment(horizontal="center"); l.border = thin
        m = ws.cell(row=row, column=13, value=tc.count); m.font = _f(10)
        m.alignment = Alignment(horizontal="center"); m.border = thin
        c_price = ws.cell(row=row, column=14, value=price)
        c_price.font = _f(10); c_price.number_format = MONEY_NF
        c_price.alignment = Alignment(horizontal="center"); c_price.border = thin
        c_total = ws.cell(row=row, column=15, value=total)
        c_total.font = _f(10); c_total.number_format = MONEY_NF
        c_total.alignment = Alignment(horizontal="center"); c_total.border = thin
        ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
        # 规则说明（附加列 R，不影响基础表）
        r = ws.cell(row=row, column=18, value=tc.rule); r.font = _f(10)

        if is_error:
            red = Font(name=FONT_NAME, size=10, bold=True, color="FF0000")
            red_fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
            c_price.font = red; c_price.fill = red_fill
            c_total.font = red; c_total.fill = red_fill

        order_no += 1
        cat_no += 1

    # 列宽
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w


def write_real_excel(sheets, filepath, hospital_name, is_negative=False, month=6, year=2026):
    """sheets: list[(dept_name, cases)]。按真实账单多科室结构写入。"""
    all_cases = [tc for _, cs in sheets for tc in cs]
    total = len(all_cases)
    error_indices = set()
    if is_negative:
        random.seed(hash(hospital_name + ("neg" if is_negative else "pos")) % 2**32)
        err_count = max(2, total // 5)
        error_indices = set(random.sample(range(total), min(err_count, total)))

    hospital_totals = {
        "pack": sum(tc.pack_count for tc in all_cases),
        "count": sum(tc.count for tc in all_cases),
        "price": round(sum(tc.price for tc in all_cases), 2),
        "amount": round(sum(tc.price * tc.pack_count for tc in all_cases), 2),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 负向标题标注（写入首个 sheet 标题旁，不破坏结构）
    tag = "负向" if is_negative else "正向"

    global_idx = 0
    for sheet_idx, (dept_name, cases) in enumerate(sheets):
        ws = wb.create_sheet(title=dept_name)
        flags = [i in error_indices for i in range(global_idx, global_idx + len(cases))]
        _write_sheet(ws, dept_name, cases, hospital_name, is_negative, month, year,
                     sheet_idx == 0, flags, hospital_totals)
        if sheet_idx == 0:
            note = f"{hospital_name}{tag}测试（红色为故意标注的错误结果）" if is_negative else f"{hospital_name}{tag}测试"
            ws.cell(row=1, column=18, value=note).font = _f(10)
        global_idx += len(cases)

    wb.save(filepath)

# ════════════════════════════════════════
# 通用规则测试用例
# ════════════════════════════════════════
def build_generic_cases():
    cases = []
    PP_LT = "额外包(低温等离子)"   # 低温纸塑袋
    PP_HT = "额外包(纸塑袋)"       # 高温纸塑袋
    NW_HT = "额外包(无纺布)"       # 无纺布 高温

    # 1. 克氏针 (≤10含包材, >10免包材) — 低温纸塑袋
    for cnt, bag, code in [(5, 20, "z2030"), (7, 20, "z2030"), (10, 20, "z2030"),
                           (12, 20, "z2030"), (15, 20, "z2060")]:
        wp = cnt <= 10
        cases.append(TestCase(PP_LT, "克氏针", cnt, code, fold_price(cnt, bag, wp),
                              f"克氏针{cnt}件{'含包材' if wp else '免包材'}"))

    # 2. 银质针
    for cnt, bag, code in [(3, 20, "z2030"), (8, 10, "z1026")]:
        wp = cnt <= 10
        cases.append(TestCase(PP_LT, "银质针", cnt, code, fold_price(cnt, bag, wp),
                              f"银质针{cnt}件{'含包材' if wp else '免包材'}"))

    # 3. 内热针
    for cnt in [5, 11]:
        wp = cnt <= 10
        cases.append(TestCase(PP_LT, "内热针", cnt, "z2030", fold_price(cnt, 20, wp),
                              f"内热针{cnt}件{'含包材' if wp else '免包材'}"))

    # 4. 车针
    for cnt, bag, code in [(4, 15, "z1526"), (13, 20, "z2030")]:
        wp = cnt <= 10
        cases.append(TestCase(PP_LT, "车针", cnt, code, fold_price(cnt, bag, wp),
                              f"车针{cnt}件{'含包材' if wp else '免包材'}"))

    # 5-7. 拔髓针/扩大针/根扩针
    cases.append(TestCase(PP_LT, "拔髓针", 6, "z2030", fold_price(6, 20), "拔髓针6件含包材"))
    cases.append(TestCase(PP_LT, "扩大针", 9, "z2030", fold_price(9, 20), "扩大针9件含包材"))
    cases.append(TestCase(PP_LT, "根扩针", 2, "z2030", fold_price(2, 20), "根扩针2件含包材"))

    # 8. 缝合针 (按1件)
    for bag, code in [(10, "z1026"), (20, "z2030")]:
        cases.append(TestCase(PP_LT, "缝合针", 5, code, round(5.5 + PACK_PRICE[bag], 2),
                              f"缝合针按1件{bag}cm"))

    # 9. 卷棉子
    for cnt in [5, 12]:
        wp = cnt <= 10
        cases.append(TestCase(PP_LT, "卷棉子", cnt, "z2030", fold_price(cnt, 20, wp),
                              f"卷棉子{cnt}件{'含包材' if wp else '免包材'}"))

    # 10. 双层袋（特殊材料，无尺寸编码）
    cases.append(TestCase(PP_LT, "测试双层袋", 1, "", 35.0, "双层袋1件固定35", mat="双层袋"))
    for cnt in [5, 9, 20, 29, 40]:
        cases.append(TestCase(PP_LT, "测试双层袋", cnt, "", float(LOWTEMP_TIER[cnt]),
                              f"双层袋{cnt}件走价格表", mat="双层袋"))

    # 11. 软镜固定300（低温纸塑袋）
    cases.append(TestCase(PP_LT, "软镜", 1, "z2060", 300.0, "软镜低温固定300"))
    cases.append(TestCase(PP_LT, "软镜", 5, "z2060", 300.0, "软镜5件固定300"))

    # 12. 敷料包(纸塑袋)
    cases.append(TestCase("敷料包(纸塑袋)", "敷料包", 1, "z1526", 2.5, "敷料包<20cm=2.5"))
    cases.append(TestCase("敷料包(纸塑袋)", "敷料包", 1, "z2530", 4.0, "敷料包≥20cm=4"))

    # 13. 低温纸塑袋阶梯
    for cnt in [2, 3, 4, 5, 6, 8, 9, 10, 14, 15, 18, 20, 25, 29, 30, 35, 39, 40]:
        cases.append(TestCase(PP_LT, "测试器械包", cnt, "z2030", float(LOWTEMP_TIER[cnt]),
                              f"低温纸塑袋{cnt}件={LOWTEMP_TIER[cnt]}"))

    # 14. 低温无纺布阶梯（器械包(低温等离子) 无纺布）
    for cnt in [2, 5, 9, 15, 20, 29, 40]:
        cases.append(TestCase("器械包(低温等离子)", "测试无纺布包", cnt, "w9050",
                              float(LOWTEMP_TIER[cnt]),
                              f"低温无纺布{cnt}件={LOWTEMP_TIER[cnt]}"))

    # 15. 单件各尺寸（低温纸塑袋）
    for bag, code in [(10, "z1026"), (15, "z1526"), (20, "z2030"), (25, "z2560"), (30, "z3040")]:
        cases.append(TestCase(PP_LT, "测试包", 1, code, float(ONE_PIECE_PRICE[bag]),
                              f"单件纸塑袋{bag}cm={ONE_PIECE_PRICE[bag]}"))
    # 单件无纺布（高温，额外包(无纺布)）
    cases.append(TestCase(NW_HT, "测试无纺布", 1, "w9050", 35.0, "单件无纺布=35"))

    return cases

# ════════════════════════════════════════
# 22家医院特殊计价用例
# ════════════════════════════════════════
def build_hospital_cases():
    Hs = {}
    PP_HT = "额外包(纸塑袋)"
    PP_LT = "额外包(低温等离子)"
    NW_HT = "额外包(无纺布)"

    # 1. 冰城医美（无纺布 高温）
    H = "冰城医美"; Hs[H] = []
    for name, cnt, fee in [("环钻包", 5, 3), ("环钻包", 8, 3), ("整形手术包", 6, 3),
                           ("脂充包", 4, 5), ("脂充包", 7, 5)]:
        Hs[H].append(TestCase(NW_HT, name, cnt, "w9050", round(cnt * 5.5 + fee, 2),
                              f"{name}{cnt}件*5.5+{fee}"))

    # 2. 电机厂医院
    H = "电机厂医院"; Hs[H] = []
    Hs[H].append(TestCase(PP_HT, "缝合针", 3, "z7530", 8.0, "电机厂缝合针=8"))
    Hs[H].append(TestCase(PP_HT, "双爪钳", 2, "z3040", round(2*5.5+PACK_PRICE[20], 2), "双<3件含包材"))
    Hs[H].append(TestCase(PP_HT, "双爪钳", 4, "z3040", round(4*5.5+PACK_PRICE[20], 2), "双≥3件含内层袋"))
    Hs[H].append(TestCase(PP_HT, "指针", 7, "z2044", fold_price(7, 20), "指针7件含包材"))
    Hs[H].append(TestCase(PP_HT, "指针", 12, "z2044", fold_price(12, 20, False), "指针12件免包材"))
    for w, code, p in [("W60", "w6050", 25), ("W90", "w9050", 30), ("W120", "w12050", 35)]:
        Hs[H].append(TestCase("敷料包(无纺布)", "棉球敷料包", 1, code, float(p),
                              f"棉球{w}={p}"))
    for w, code, p in [("W70", "w7050", 25), ("W150", "w15050", 35)]:
        Hs[H].append(TestCase("敷料包(无纺布)", "纱布敷料包", 1, code, float(p),
                              f"纱布{w}={p}"))

    # 3. 方南南医院（高温纸塑袋）
    H = "方南南医院"; Hs[H] = []
    for name in ["P钻", "根管锉", "光滑针", "机扩针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            Hs[H].append(TestCase(PP_HT, name, cnt, "z2044", fold_price(cnt, 20, wp),
                                  f"方南南{name}{cnt}件{'含' if wp else '免'}"))

    # 4. 东北农业大学
    H = "东北农业大学"; Hs[H] = []
    for name in ["根管针", "机锉", "牙探针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            Hs[H].append(TestCase(PP_HT, name, cnt, "z2044", fold_price(cnt, 20, wp),
                                  f"东北农大{name}{cnt}件{'含' if wp else '免'}"))
    for cnt in [6, 12]:
        wp = cnt <= 10
        Hs[H].append(TestCase(PP_HT, "根管锉", cnt, "z2044", fold_price(cnt, 20, wp),
                              f"根管锉{cnt}件5.5单价{'含' if wp else '免'}"))

    # 5. 哈尔滨工程大学（敷料包）
    H = "哈尔滨工程大学"; Hs[H] = []
    Hs[H].append(TestCase("敷料包(纸塑袋)", "孔巾敷料包", 1, "z2530", 4.0, "孔巾≥20cm=4"))
    Hs[H].append(TestCase("敷料包(纸塑袋)", "孔巾敷料包", 1, "z1526", 2.5, "孔巾<20cm=2.5"))

    # 6. 松电慢病
    H = "松电慢病"; Hs[H] = []
    for cnt in [5, 12]:
        wp = cnt <= 10
        Hs[H].append(TestCase(PP_HT, "机扩针", cnt, "z2044", fold_price(cnt, 20, wp),
                              f"松电机扩针{cnt}件{'含' if wp else '免'}"))

    # 7. 航天风华
    H = "航天风华"; Hs[H] = []
    for cnt in [5, 12]:
        wp = cnt <= 10
        Hs[H].append(TestCase(PP_HT, "镍钛锉", cnt, "z2044", fold_price(cnt, 20, wp),
                              f"航天镍钛锉{cnt}件{'含' if wp else '免'}"))

    # 8. 市五院（二门诊）（无纺布 敷料）
    H = "市五院（二门诊）"; Hs[H] = []
    for w, code, p in [("W50", "w5050", 25), ("W60", "w6050", 25), ("W90", "w9050", 30),
                       ("W120", "w12050", 35), ("W150", "w15050", 35)]:
        Hs[H].append(TestCase("敷料包(无纺布)", "驱血带", 1, code, float(p),
                              f"驱血带{w}={p}"))

    # 9. 九州医院
    H = "九州医院"; Hs[H] = []
    Hs[H].append(TestCase(PP_HT, "方盘", 3, "z2044", 5.5, "九州方盘固定5.5"))

    # 10. 博尚医院（低温纸塑袋）
    H = "博尚医院"; Hs[H] = []
    for name, price in [("旋切器1戳卡1胶帽4", 66), ("旋切器1胶帽4", 44), ("旋切器1线1胶帽4", 66)]:
        Hs[H].append(TestCase(PP_LT, name, 6, "z2045", float(price), f"博尚{name}固定{price}"))

    # 11. 黑龙江省海员总医院（松北）
    H = "黑龙江省海员总医院（松北）"; Hs[H] = []
    for bag, code in [(20, "z2030"), (30, "z3040")]:
        Hs[H].append(TestCase(PP_LT, "胶帽", 3, code, float(ONE_PIECE_PRICE[bag]),
                              f"胶帽≤5按1件{bag}cm"))
    Hs[H].append(TestCase(PP_LT, "胶帽", 7, "z2030", float(ceil_div(7, 5) * 22), "胶帽>5: ceil(7/5)*22=44"))
    Hs[H].append(TestCase(PP_LT, "胶帽", 12, "z2030", float(ceil_div(12, 5) * 22), "胶帽>5: ceil(12/5)*22=66"))

    # 12. 黑龙江省妇幼保健院（人口）
    H = "黑龙江省妇幼保健院（人口）"; Hs[H] = []
    for bag, code in [(20, "z2030"), (30, "z3040")]:
        Hs[H].append(TestCase(PP_LT, "密封胶圈", 4, code, float(ONE_PIECE_PRICE[bag]),
                              f"密封胶圈≤5按1件{bag}cm"))
    Hs[H].append(TestCase(PP_LT, "密封胶圈", 8, "z2030", float(ceil_div(8, 5) * 22), "密封胶圈>5=44"))
    for name in ["根管锉", "机扩锉", "机扩针", "荡洗针", "加长锉", "洗髓针", "彩色锉", "手扩锉", "针类组件"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            Hs[H].append(TestCase(PP_HT, name, cnt, "z2044", fold_price(cnt, 20, wp),
                                  f"妇幼{name}{cnt}件{'含' if wp else '免'}"))
    Hs[H].append(TestCase(PP_HT, "针包", 3, "z2044", round(5.5 + PACK_PRICE[20], 2), "针包按1件=28"))
    for cnt in [5, 8]:
        wp = cnt <= 6
        inst = (ceil_div(cnt, 5) + 1) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        Hs[H].append(TestCase(PP_HT, "针多少盒1", cnt, "z2044", p, f"针多少盒1 {cnt}件{'含' if wp else '免'}"))
    Hs[H].append(TestCase(PP_LT, "新腹腔镜镜头", 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), "腹腔镜镜头+8=36"))
    Hs[H].append(TestCase(PP_LT, "单极电切镜", 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), "单极电切镜标准+8"))
    Hs[H].append(TestCase(PP_LT, "宫腔镜检查新", 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), "宫腔镜检查新标准+8"))
    Hs[H].append(TestCase(PP_LT, "0°膀胱镜", 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), "0°膀胱镜标准+8"))
    Hs[H].append(TestCase(PP_LT, "30°膀胱镜", 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), "30°膀胱镜标准+8"))

    # 13. 祖研南岗
    H = "祖研-黑龙江省中医医院（南岗院区）"; Hs[H] = []
    for cnt in [10, 15, 25, 30]:
        wp = cnt <= 20
        inst = ceil_div(cnt, 10) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        Hs[H].append(TestCase(PP_HT, "排针", cnt, "z2044", p, f"排针{cnt}件/10{'含' if wp else '免'}"))

    # 14. 黑龙江省社会康复医院
    H = "黑龙江省社会康复医院"; Hs[H] = []
    fixed = [
        ("抛光车针盒6件盒1", 16.5, "z1026"), ("环切器械盒-14件盒1", 16.5, "z2044"),
        ("种植盒(黄绿)8件盒1", 16.5, "z2044"), ("种植盒(金属)12件盒1", 16.5, "z2044"),
        ("种植盒(蓝白)12件盒1", 16.5, "z2044"), ("种植盒-14件盒1", 16.5, "w6050"),
        ("梁光强器械盒(粉)10件盒1", 16.5, "z2044"), ("梁光强器械盒-21件盒1", 22, "w6050"),
        ("大车针盒-1", 22, "z1526"), ("内提工具盒10件盒1", 44, "w6050"),
        ("百诺工具盒-19件盒1", 44, "w6050"), ("ITI种植盒-32件盒1", 44, "w12050"),
        ("登腾种植盒-34件盒1", 44, "w6050"), ("登腾种植盒-36件盒1", 44, "w6050"),
    ]
    for name, price, code in fixed:
        typ = "额外包(无纺布)" if code.startswith("w") else "额外包(纸塑袋)"
        Hs[H].append(TestCase(typ, name, 1, code, float(price), f"{name}固定{price}"))

    # 15. 哈尔滨市道里区妇幼保健院
    H = "哈尔滨市道里区妇幼保健院"; Hs[H] = []
    for name in ["棉花针", "洗髓针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            Hs[H].append(TestCase(PP_HT, name, cnt, "z2044", fold_price(cnt, 20, wp),
                                  f"道里{name}{cnt}件{'含' if wp else '免'}"))

    # 16. 春语医疗美容医院
    H = "春语医疗美容医院"; Hs[H] = []
    for bag, code in [(20, "z2030"), (30, "z3040")]:
        Hs[H].append(TestCase(PP_LT, "塑料管", 5, code, float(ONE_PIECE_PRICE[bag]),
                              f"塑料管≤10按1件{bag}cm"))
    Hs[H].append(TestCase(PP_LT, "管子", 15, "z2030", float(ceil_div(15, 10) * 22), "管子>10: ceil(15/10)*22=44"))

    # 17. 黑龙江总工会医院
    H = "黑龙江总工会医院"; Hs[H] = []
    for name in ["12°镜头-1", "30°镜头镜鞘-2", "检查镜-2", "宫腔检查镜30度-2件", "沈大12度镜头-1"]:
        Hs[H].append(TestCase(PP_LT, name, 1, "z2060", round(ONE_PIECE_PRICE[20] + 8, 2), f"{name}标准+8=36"))

    # 18. 哈尔滨基准生物有限公司
    H = "哈尔滨基准生物有限公司"; Hs[H] = []
    Hs[H].append(TestCase(PP_LT, "氩氦刀", 1, "z2030", 150.0, "氩氦刀低温固定150"))
    Hs[H].append(TestCase("额外包(ETO)", "氩氦刀", 1, "z2030", 150.0, "氩氦刀ETO固定150"))

    # 19. 索菲医疗美容门诊（双层袋）
    H = "索菲医疗美容门诊"; Hs[H] = []
    for cnt in [3, 5, 8]:
        Hs[H].append(TestCase(PP_LT, "面吸针", cnt, "", round(cnt * 5.5, 2),
                              f"面吸针{cnt}件={cnt*5.5}", mat="双层袋"))

    # 20. 省监狱管理局医院
    H = "省监狱管理局医院"; Hs[H] = []
    for bag, code in [(20, "z2030"), (30, "z3040")]:
        Hs[H].append(TestCase(PP_LT, "密封件", 3, code, float(ONE_PIECE_PRICE[bag]),
                              f"密封件≤5按1件{bag}cm"))
    Hs[H].append(TestCase(PP_LT, "密封胶圈", 8, "z2030", float(ceil_div(8, 5) * 22), "密封胶圈>5=44"))

    # 21. 呼兰中医院（无纺布 高温）
    H = "呼兰中医院"; Hs[H] = []
    for name, cnt in [("铂康-阑尾包", 5), ("铂康-外科包", 8), ("外科包", 6)]:
        Hs[H].append(TestCase("器械包", name, cnt, "w9050", round(cnt * 5.5 + 13, 2),
                              f"{name}={cnt}*5.5+13"))
    Hs[H].append(TestCase(PP_LT, "胶帽", 3, "z2030", float(ONE_PIECE_PRICE[20]), "胶帽≤5按1件=28"))
    Hs[H].append(TestCase(PP_LT, "胶帽", 8, "z2030", float(ceil_div(8, 5) * 22), "胶帽>5=44"))

    # 22. 平房区人民医院
    H = "平房区人民医院"; Hs[H] = []
    for cnt in [5, 8]:
        wp = cnt <= 6
        inst = (ceil_div(cnt, 5) + 1) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        Hs[H].append(TestCase(PP_HT, "针盒1针多少", cnt, "z2044", p, f"针多少{cnt}件{'含' if wp else '免'}"))

    return Hs

# ════════════════════════════════════════
# 通用规则混合模式条目
# ════════════════════════════════════════
def generic_mix(hospital_name):
    return [
        TestCase("额外包(低温等离子)", "克氏针", 5, "z2030", fold_price(5, 20), "混合-克氏针5件"),
        TestCase("额外包(低温等离子)", "测试器械包", 9, "z2030", float(LOWTEMP_TIER[9]), "混合-低温9件阶梯"),
        TestCase("额外包(低温等离子)", "软镜", 1, "z2060", 300.0, "混合-软镜固定300"),
    ]

def classify_dept(tc):
    """按材料/类型把用例分到科室 sheet，对齐真实账单多科室结构。"""
    if tc.mat == "双层袋":
        return "消毒供应室"
    if "敷料" in tc.typ:
        return "处置室"
    if tc.code.startswith("w"):
        return "消毒供应室"
    return "手术室"

def group_by_dept(cases):
    sheets = {}
    for tc in cases:
        d = classify_dept(tc)
        sheets.setdefault(d, []).append(tc)
    order = []
    for d in ("手术室", "处置室", "消毒供应室"):
        if d in sheets:
            order.append((d, sheets[d]))
    return order

# ════════════════════════════════════════
# 主函数
# ════════════════════════════════════════
def main():
    import shutil
    pos_dir = os.path.join(OUTPUT_DIR, "正向测试")
    neg_dir = os.path.join(OUTPUT_DIR, "负向测试")

    # 清理旧文件（顶层 + 两个子文件夹）
    for f in os.listdir(OUTPUT_DIR):
        p = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(p) and f.endswith(".xlsx"):
            os.remove(p)
    for d in (pos_dir, neg_dir):
        if os.path.isdir(d):
            shutil.rmtree(d)
        os.makedirs(d, exist_ok=True)

    # 1. 通用规则测试文件
    gen_cases = build_generic_cases()
    gen_sheets = group_by_dept(gen_cases)
    write_real_excel(gen_sheets, os.path.join(pos_dir, "通用规则-正向测试.xlsx"), "电机厂医院", is_negative=False)
    write_real_excel(gen_sheets, os.path.join(neg_dir, "通用规则-负向测试.xlsx"), "电机厂医院", is_negative=True)
    print(f"通用规则: 正向{len(gen_cases)}条 负向{len(gen_cases)}条, {len(gen_sheets)}科室")

    # 2. 22家医院测试文件
    hospitals = build_hospital_cases()
    print(f"医院数: {len(hospitals)}")

    for h_name, cases in hospitals.items():
        all_cases = cases + generic_mix(h_name)
        sheets = group_by_dept(all_cases)
        safe = h_name.replace("/", "_").replace("（", "(").replace("）", ")")
        write_real_excel(sheets, os.path.join(pos_dir, f"{safe}-正向测试.xlsx"), h_name, is_negative=False)
        write_real_excel(sheets, os.path.join(neg_dir, f"{safe}-负向测试.xlsx"), h_name, is_negative=True)
        print(f"  {h_name}: {len(all_cases)}条, {len(sheets)}科室")

    print(f"\n全部{2 + len(hospitals)*2}份测试文件生成完成！")
    print(f"输出目录: {pos_dir} / {neg_dir}")

if __name__ == "__main__":
    main()
