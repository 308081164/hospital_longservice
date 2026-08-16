#!/usr/bin/env python3
"""Parse docs/source/特殊收费(11).xlsx into rule-type-registry.json for SC11 unit tests."""

from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs/source/特殊收费(11).xlsx"
DEFAULT_OUT = ROOT / "backend/src/test/resources/pricing-engine/rule-type-registry.json"

HOSPITAL_TO_CODE: dict[str, str] = {
    "东北农业大学": "NEAU-YY",
    "九州医院": "JIUZHOU-FK",
    "冰城医美": "BINGCHENG-YM",
    "博尚医院": "BOSHANG-YY",
    "哈尔滨基准生物有限公司": "JZSW-BIO",
    "哈尔滨工程大学": "HRB-HEU",
    "哈尔滨市道里区妇幼保健院": "DL-FUCHAN",
    "市五院（二门诊）": "HRB-WY-EM",
    "方南南医院": "FNN-YY",
    "春语医疗美容医院": "CHUNYU-YL",
    "松电慢病": "SONGDIAN-MB",
    "电机厂医院": "GUOYAO-2",
    "省监狱管理局医院": "HLJ-JYGLJ-YY",
    "祖研-黑龙江省中医医院（南岗院区）": "ZUYAN-NG",
    "索菲医疗美容门诊": "SUOFEI-YL",
    "航天风华": "HRB-HTFH",
    "黑龙江总工会医院": "HL-ZGH",
    "黑龙江省妇幼保健院（人口）": "HLJ-FY-RK",
    "黑龙江省海员总医院（松北）": "HAIYUAN-SB",
    "黑龙江省社会康复医院": "SHKF-YY",
}


def classify_rule(rule_text: str, pack_name: str, count_hint: str | None, pack_type: str | None) -> str:
    t = (rule_text or "").strip()
    pack = (pack_name or "").strip()
    count = (count_hint or "").strip()
    ptype = (pack_type or "").strip()

    if not t and ("镜头" in pack or "镜" in pack):
        return "SC11-T13"
    if "件数*5.5" in t and "＋" in t:
        return "SC11-T01"
    if re.search(r"1\s*\*\s*5\.5", t):
        return "SC11-T02"
    if "双" in (pack or "") and "固定收费35" in t:
        return "SC11-T03b"
    if "双" in (pack or "") and ("5.5*件数" in t or "5.5×件数" in t):
        return "SC11-T03"
    if "/10" in t:
        return "SC11-T06"
    if "/5" in t and "5.6" in t:
        return "SC11-T04b"
    if "/5" in t and "＞10" in count:
        return "SC11-T05"
    if "/5" in t:
        return "SC11-T04"
    if any(w in t for w in ("W50", "W60", "W70", "W90", "W120", "W150")):
        return "SC11-T07"
    if "纸塑袋宽度≥20" in t or "≥20CM" in t:
        return "SC11-T09"
    if "纸塑袋宽＜20" in t or "＜20CM" in t:
        return "SC11-T10"
    if "按一件算" in t or ("≤5" in count and "密封" in pack):
        return "SC11-T11"
    if "按一件一包" in t:
        return "SC11-T12"
    if "标准价上加" in t or "标准价上" in t:
        return "SC11-T13"
    if "双层袋" in t and "不额外收费" in t:
        return "SC11-T14"
    if "固定收费300" in t or "固定300" in t:
        return "SC11-T15"
    if re.fullmatch(r"(16\.5|22|44)", t):
        return "SC11-T08"
    if "固定" in t or re.search(r"固定收?\d", t):
        return "SC11-T08"
    if ptype and "低温" in ptype and not t:
        return "SC11-T15"
    return "SC11-UNCLASSIFIED"


def parse_hospital_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict] = []
    cur_hospital = None
    cur_seq = None
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        if row[1]:
            cur_hospital = str(row[1]).strip()
        if row[2]:
            cur_seq = str(row[2]).strip()
        pack_name = row[3]
        if not pack_name:
            continue
        pack_name = str(pack_name).strip()
        pack_type = str(row[4]).strip() if row[4] else None
        count_hint = str(row[5]).strip() if row[5] is not None and str(row[5]).strip() else None
        rule_text = str(row[6]).strip() if row[6] else ""
        note = str(row[7]).strip() if row[7] else None
        sc11_type = classify_rule(rule_text, pack_name, count_hint, pack_type)
        entries.append(
            {
                "id": f"hospital_{len(entries)+1:03d}",
                "sheet": "各医院特殊收费",
                "row": row_idx,
                "hospitalName": cur_hospital,
                "customerCode": HOSPITAL_TO_CODE.get(cur_hospital or ""),
                "seq": cur_seq,
                "packName": pack_name,
                "packType": pack_type,
                "instrumentCountHint": count_hint,
                "ruleText": rule_text,
                "note": note,
                "sc11Type": sc11_type,
            }
        )
    return entries


def parse_generic_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict] = []
    cur_pack = None
    cur_pack_type = None
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        if row[1]:
            cur_pack = str(row[1]).strip()
        if row[2]:
            cur_pack_type = str(row[2]).strip()
        rule_text = str(row[4]).strip() if row[4] else ""
        if not rule_text:
            continue
        count_hint = str(row[3]).strip() if row[3] is not None and str(row[3]).strip() else None
        sc11_type = classify_rule(rule_text, cur_pack or "", count_hint, cur_pack_type)
        entries.append(
            {
                "id": f"generic_{len(entries)+1:03d}",
                "sheet": "通用特殊收费",
                "row": row_idx,
                "hospitalName": None,
                "customerCode": None,
                "seq": str(row[0]).strip() if row[0] else None,
                "packName": cur_pack,
                "packType": cur_pack_type,
                "instrumentCountHint": count_hint,
                "ruleText": rule_text,
                "note": str(row[5]).strip() if row[5] else None,
                "sc11Type": sc11_type,
            }
        )
    return entries


def parse_tier_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict] = []
    for row_idx, row in enumerate(rows[2:14], start=3):
        if row[0] and row[1] is not None:
            entries.append(
                {
                    "id": f"tier_lt_left_{len(entries)+1:03d}",
                    "sheet": "环氧与低温通用收费",
                    "row": row_idx,
                    "section": "lt_eo_piece_tier_left",
                    "pieceCountLabel": str(row[0]).strip(),
                    "price": row[1],
                    "sc11Type": "SC11-T16",
                }
            )
        if row[3] and row[4] is not None:
            entries.append(
                {
                    "id": f"tier_lt_right_{len(entries)+1:03d}",
                    "sheet": "环氧与低温通用收费",
                    "row": row_idx,
                    "section": "lt_eo_piece_tier_right",
                    "pieceCountLabel": str(row[3]).strip(),
                    "price": row[4],
                    "sc11Type": "SC11-T16",
                }
            )
    for row_idx, row in enumerate(rows[2:10], start=3):
        if row[6] and row[7] is not None:
            entries.append(
                {
                    "id": f"tier_width_{len(entries)+1:03d}",
                    "sheet": "环氧与低温通用收费",
                    "row": row_idx,
                    "section": "paper_plastic_width_addon",
                    "bagWidthLabel": str(row[6]).strip(),
                    "price": row[7],
                    "note": str(row[8]).strip() if len(row) > 8 and row[8] else None,
                    "sc11Type": "SC11-T16",
                }
            )
    seen: set[str] = set()
    deduped: list[dict] = []
    for e in entries:
        key = json.dumps(e, sort_keys=True, ensure_ascii=False)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(e)
    return deduped


def build_registry(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    hospital = parse_hospital_sheet(wb["各医院特殊收费"])
    generic = parse_generic_sheet(wb["通用特殊收费"])
    tier = parse_tier_sheet(wb["环氧与低温通用收费"])

    all_entries = hospital + generic + tier
    type_counts: dict[str, int] = {}
    for e in all_entries:
        t = e["sc11Type"]
        type_counts[t] = type_counts.get(t, 0) + 1

    return {
        "version": "2026-08-14",
        "sourceFile": str(xlsx_path.relative_to(ROOT)),
        "generatedAt": date.today().isoformat(),
        "counts": {
            "hospitalRules": len(hospital),
            "genericRules": len(generic),
            "tierRules": len(tier),
            "total": len(all_entries),
        },
        "sc11TypeCounts": dict(sorted(type_counts.items())),
        "hospitalToCustomerCode": HOSPITAL_TO_CODE,
        "entries": {
            "hospital": hospital,
            "generic": generic,
            "tier": tier,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    registry = build_registry(args.xlsx)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(registry, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {args.out} ({registry['counts']['total']} entries)")
    print("SC11 types:", registry["sc11TypeCounts"])


if __name__ == "__main__":
    main()
