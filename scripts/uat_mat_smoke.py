#!/usr/bin/env python3
"""
UAT MAT smoke helper — row-count / total reconciliation for export xlsx pairs.

Usage:
  python scripts/uat_mat_smoke.py --mat01 raw.xlsx --mat02 expected.xlsx --mat02-actual actual.xlsx
  python scripts/uat_mat_smoke.py --settlement expected_settlement.xlsx actual_settlement.xlsx --tolerance 0.01
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def sheet_totals(path: Path, sheet_index: int = 0) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    numeric = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if isinstance(cell, (int, float)):
                numeric.append(float(cell))
    return {
        "file": str(path),
        "sheet": wb.sheetnames[sheet_index],
        "rows": max(ws.max_row - 1, 0),
        "numeric_cells": len(numeric),
        "numeric_sum": round(sum(numeric), 2) if numeric else 0.0,
    }


def compare_totals(a: dict, b: dict, tolerance: float) -> list[str]:
    issues = []
    if a["rows"] != b["rows"]:
        issues.append(f"row count {a['rows']} != {b['rows']}")
    delta = abs(a["numeric_sum"] - b["numeric_sum"])
    if delta > tolerance:
        issues.append(f"numeric sum delta {delta:.4f} > tolerance {tolerance}")
    return issues


def main() -> int:
    parser = argparse.ArgumentParser(description="UAT MAT smoke checks")
    parser.add_argument("--mat01", type=Path, help="MAT-01 raw bill xlsx")
    parser.add_argument("--mat02", type=Path, help="MAT-02 expected xlsx")
    parser.add_argument("--mat02-actual", type=Path, help="MAT-02 system export xlsx")
    parser.add_argument("--settlement", nargs=2, type=Path, metavar=("EXPECTED", "ACTUAL"))
    parser.add_argument("--tolerance", type=float, default=0.01)
    args = parser.parse_args()

    ok = True
    if args.mat02 and args.mat02_actual:
        exp = sheet_totals(args.mat02)
        act = sheet_totals(args.mat02_actual)
        issues = compare_totals(exp, act, args.tolerance)
        print(f"MAT-02 compare: expected={exp['rows']} rows sum={exp['numeric_sum']}, "
              f"actual={act['rows']} rows sum={act['numeric_sum']}")
        if issues:
            ok = False
            for issue in issues:
                print("  FAIL:", issue)
        else:
            print("  OK — within tolerance")

    if args.settlement:
        exp = sheet_totals(args.settlement[0])
        act = sheet_totals(args.settlement[1])
        issues = compare_totals(exp, act, args.tolerance)
        print(f"MAT-03 settlement: expected sum={exp['numeric_sum']}, actual sum={act['numeric_sum']}")
        if issues:
            ok = False
            for issue in issues:
                print("  FAIL:", issue)
        else:
            print("  OK — within tolerance")

    if args.mat01:
        raw = sheet_totals(args.mat01)
        print(f"MAT-01 baseline: {raw['rows']} rows, numeric sum={raw['numeric_sum']}")

    if not args.mat02 and not args.settlement and not args.mat01:
        parser.print_help()
        return 2

    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
