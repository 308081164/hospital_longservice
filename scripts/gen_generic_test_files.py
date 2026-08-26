#!/usr/bin/env python3
"""生成通用规则测试文件（正向+负向）"""
import math
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "系统内置测试用例")

# ── 包装材料价格表 ──
PACK_PRICE = {10: 16.5, 15: 19.5, 20: 22.5, 25: 24.5, 30: 29.5}
ONE_PIECE_PRICE = {10: 22, 15: 25, 20: 28, 25: 30, 30: 35}

# ── 低温阶梯价格表（1-40件）──
LOWTEMP_TIER = {
    1: 22, 2: 44, 3: 66, 4: 88, 5: 88, 6: 110, 7: 132, 8: 154,
    9: 165, 10: 165, 11: 187, 12: 209, 13: 231, 14: 253, 15: 253,
    16: 275, 17: 297, 18: 300, 19: 300, 20: 300, 21: 322, 22: 344,
    23: 366, 24: 388, 25: 388, 26: 410, 27: 432, 28: 454, 29: 465,
    30: 465, 31: 487, 32: 509, 33: 531, 34: 553, 35: 553, 36: 575,
    37: 597, 38: 600, 39: 600, 40: 600,
}

def ceil_div(n, d):
    return math.ceil(n / d)

def fold_price(count, bag_size, with_packaging=True):
    """小件5合1规则价格"""
    instrument = ceil_div(count, 5) * 5.5
    if with_packaging and count <= 10:
        return round(instrument + PACK_PRICE[bag_size], 2)
    return round(instrument, 2)

# ── 通用规则测试用例 ──
GENERIC_POSITIVE_CASES = []

# 1. 克氏针
for cnt, bag, desc in [(5, 20, "克氏针-5件≤10含包材"), (7, 20, "克氏针-7件≤10含包材"),
                        (10, 20, "克氏针-10件≤10含包材"), (12, 20, "克氏针-12件>10免包材"),
                        (15, 25, "克氏针-15件>10免包材")]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"克氏针-{cnt}件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": cnt, "packCount": 1, "rule": desc,
        "price": fold_price(cnt, bag),
    })

# 2. 银质针
for cnt, bag in [(3, 20), (8, 10)]:
    wp = cnt <= 10
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"银质针-{cnt}件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": cnt, "packCount": 1,
        "rule": f"银质针-{cnt}件{'≤10含包材' if wp else '>10免包材'}",
        "price": fold_price(cnt, bag, wp),
    })

# 3. 内热针
for cnt, bag in [(5, 20), (11, 20)]:
    wp = cnt <= 10
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"内热针-{cnt}件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": cnt, "packCount": 1,
        "rule": f"内热针-{cnt}件{'≤10含包材' if wp else '>10免包材'}",
        "price": fold_price(cnt, bag, wp),
    })

# 4. 车针
for cnt, bag in [(4, 15), (13, 20)]:
    wp = cnt <= 10
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"车针-{cnt}件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": cnt, "packCount": 1,
        "rule": f"车针-{cnt}件{'≤10含包材' if wp else '>10免包材'}",
        "price": fold_price(cnt, bag, wp),
    })

# 5. 拔髓针
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "拔髓针-6件/Z7530", "material": "纸塑袋 20cm",
    "count": 6, "packCount": 1, "rule": "拔髓针-6件≤10含包材",
    "price": fold_price(6, 20),
})

# 6. 扩大针
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "扩大针-9件/Z7530", "material": "纸塑袋 20cm",
    "count": 9, "packCount": 1, "rule": "扩大针-9件≤10含包材",
    "price": fold_price(9, 20),
})

# 7. 根扩针
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "根扩针-2件/Z7530", "material": "纸塑袋 20cm",
    "count": 2, "packCount": 1, "rule": "根扩针-2件≤10含包材",
    "price": fold_price(2, 20),
})

# 8. 缝合针（按1件含包材）
for bag in [10, 20]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "缝合针-5件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": 5, "packCount": 1, "rule": f"缝合针按1件含包材{bag}cm",
        "price": round(5.5 + PACK_PRICE[bag], 2),
    })

# 9. 卷棉子
for cnt, bag in [(5, 20), (12, 20)]:
    wp = cnt <= 10
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"卷棉子-{cnt}件/Z7530", "material": f"纸塑袋 {bag}cm",
        "count": cnt, "packCount": 1,
        "rule": f"卷棉子-{cnt}件{'≤10含包材' if wp else '>10免包材'}",
        "price": fold_price(cnt, bag, wp),
    })

# 10. 双层袋低温1件固定35
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "测试双层袋-1件", "material": "双层袋",
    "count": 1, "packCount": 1, "rule": "双层袋低温1件固定35",
    "price": 35.0,
})
# 双层袋多件走价格表
for cnt in [5, 9, 20, 29, 40]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"测试双层袋-{cnt}件", "material": "双层袋",
        "count": cnt, "packCount": 1, "rule": f"双层袋{cnt}件走低温价格表",
        "price": float(LOWTEMP_TIER[cnt]),
    })

# 11. 软镜固定300
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "软镜-1件", "material": "纸塑袋 20cm",
    "count": 1, "packCount": 1, "rule": "软镜低温固定300",
    "price": 300.0,
})
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "软镜-5件", "material": "纸塑袋 20cm",
    "count": 5, "packCount": 1, "rule": "软镜低温5件固定300",
    "price": 300.0,
})

# 12. 敷料包(纸塑袋)
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "高温灭菌",
    "packName": "敷料包-纸塑袋15cm", "material": "纸塑袋 15cm",
    "count": 1, "packCount": 1, "rule": "敷料包纸塑袋<20cm=2.5",
    "price": 2.5,
})
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "高温灭菌",
    "packName": "敷料包-纸塑袋25cm", "material": "纸塑袋 25cm",
    "count": 1, "packCount": 1, "rule": "敷料包纸塑袋>=20cm=4",
    "price": 4.0,
})

# 13. 低温纸塑袋阶梯（2-40件关键件数）
for cnt in [2, 3, 4, 5, 6, 8, 9, 10, 14, 15, 18, 20, 25, 29, 30, 35, 39, 40]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"测试器械包-{cnt}件", "material": "纸塑袋 20cm",
        "count": cnt, "packCount": 1, "rule": f"低温纸塑袋{cnt}件={LOWTEMP_TIER[cnt]}",
        "price": float(LOWTEMP_TIER[cnt]),
    })

# 14. 低温无纺布阶梯
for cnt in [2, 5, 9, 15, 20, 29, 40]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": f"测试无纺布包-{cnt}件", "material": "无纺布",
        "count": cnt, "packCount": 1, "rule": f"低温无纺布{cnt}件={LOWTEMP_TIER[cnt]}",
        "price": float(LOWTEMP_TIER[cnt]),
    })

# 15. 单件纸塑袋各尺寸
for bag in [10, 15, 20, 25, 30]:
    GENERIC_POSITIVE_CASES.append({
        "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "测试包-1件", "material": f"纸塑袋 {bag}cm",
        "count": 1, "packCount": 1, "rule": f"单件纸塑袋{bag}cm={ONE_PIECE_PRICE[bag]}",
        "price": float(ONE_PIECE_PRICE[bag]),
    })

# 16. 单件无纺布
GENERIC_POSITIVE_CASES.append({
    "hospital": "电机厂医院", "dept": "测试科", "type": "低温等离子灭菌",
    "packName": "测试无纺布-1件", "material": "无纺布",
    "count": 1, "packCount": 1, "rule": "单件无纺布=35",
    "price": 35.0,
})

# ── 生成Excel文件 ──
HEADERS = ["序号", "医院名称", "科室", "灭菌类型", "包名称", "包装材料",
           "器械数", "包数", "单价", "总价", "规则说明"]

def make_excel(cases, filepath, error_indices=None):
    """生成Excel测试文件。error_indices中的行号(0-based)会被故意写错并标红"""
    error_indices = error_indices or set()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 标题行
    title = "通用规则正向测试" if not error_indices else "通用规则负向测试（红色为故意标注的错误结果）"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    # 表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    red_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, case in enumerate(cases):
        row = i + 3
        price = case["price"]
        is_error = i in error_indices
        if is_error:
            # 引入错误：+5或-5
            price = round(price + 5.0, 2)

        vals = [
            i + 1, case["hospital"], case["dept"], case["type"],
            case["packName"], case["material"], case["count"],
            case["packCount"], price, round(price * case["packCount"], 2),
            case["rule"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if col != 11 else "left")
            if is_error and col in (9, 10):  # 单价和总价标红
                c.font = red_font
                c.fill = red_fill

    # 列宽
    widths = [6, 16, 10, 16, 28, 16, 8, 8, 10, 10, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filepath)
    print(f"  生成: {os.path.basename(filepath)} ({len(cases)}条用例, {len(error_indices)}条错误)")

# 正向测试文件
positive_path = os.path.join(OUTPUT_DIR, "通用规则-正向测试.xlsx")
make_excel(GENERIC_POSITIVE_CASES, positive_path)

# 负向测试文件（在正向基础上引入10处错误）
import random
random.seed(42)
total = len(GENERIC_POSITIVE_CASES)
error_indices = set(random.sample(range(total), min(10, total)))
negative_path = os.path.join(OUTPUT_DIR, "通用规则-负向测试.xlsx")
make_excel(GENERIC_POSITIVE_CASES, negative_path, error_indices)

print(f"\n通用规则测试文件生成完成：")
print(f"  正向: {positive_path}")
print(f"  负向: {negative_path}")
print(f"  总用例数: {len(GENERIC_POSITIVE_CASES)}")
