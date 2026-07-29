#!/usr/bin/env python3
"""S8 结款函比对：Job export-v2 settlement vs 测试用例/处理后表格/*结款*.xlsx

仅针对「账单+结款函」26 院；写报告至 测试用例/s8_settlement_compare_report.json
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TEST_CASE = ROOT / "测试用例"
RECON_MD = TEST_CASE / "批量6月系统对账结果.md"
REPORT_JSON = TEST_CASE / "s8_settlement_compare_report.json"
EXPORT_DIR = TEST_CASE / ".s8_exports"


def configure_output_paths(
    *,
    report_suffix: str | None = None,
    export_dir: Path | None = None,
) -> None:
    global REPORT_JSON, EXPORT_DIR  # noqa: PLW0603
    if export_dir is not None:
        EXPORT_DIR = export_dir if export_dir.is_absolute() else (ROOT / export_dir)
    if report_suffix:
        REPORT_JSON = TEST_CASE / f"s8_settlement_compare_report.{report_suffix}.json"

BACKEND = __import__("os").environ.get("BACKEND_CONTAINER", "hospital-backend")
API = __import__("os").environ.get("API_INTERNAL", "http://127.0.0.1:8000")

sys.path.insert(0, str(ROOT / "scripts"))
from batch_s8_export_compare import (  # noqa: E402
    export_bill,
    get_token,
    init_api_from_args,
    load_job_map,
    parse_job_table,
)

from batch_june_price_reconciliation import TODO_HOSPITALS  # noqa: E402

DEFAULT_JOB_MAP = TEST_CASE / "job_baseline_stable.json"

# 全量 37 院：有结款函参考表则 L2 比对，无则 skip
BILL_SETTLEMENT_ONLY = list(TODO_HOSPITALS)

SETTLEMENT_DEV_REQUIRED = frozenset()

MATERIAL_BLOCKED = frozenset(
    {
        "国药总医院主院区",
        "国药总医院第二院区",
        "哈尔滨市第二医院",
        "黑龙江省第二医院（松北院区）",
    }
)

# 结款合并：参考表/export Job 与 folder 不同（波次6）
SETTLEMENT_MERGE_RULES: dict[str, dict] = {
    "哈尔滨市第五医院（二门诊）": {
        "reference_folder": "哈尔滨市第五医院",
        "export_job_folder": "哈尔滨市第五医院",
        "note": "合并结款 · 见市五院 Job613",
    },
}

# 结款函已知差：总额或条目与处理后表口径不同，登记后 warn
KNOWN_SETTLEMENT_DIFF = {
    "黑龙江九洲妇科医院": {"logistics_waiver": "6月物流已免（处理后表 logistics=0）"},
    "呼兰中医院": {"minor_delta": "灭菌 part2 口径 Δ22 · 特殊包固定价"},
    "新发红十字医院": {"xinfa_deduction": "设备抵扣行非数值 · 敷料/加急已 enrich"},
    "太平人民医院": {"settlement_discount": "结款7.5折 settlement_only · 物流行已隐藏"},
    "道外区人民医院": {"no_logistics": "无 LOGISTICS 策略 · 处理后仅灭菌行"},
    "三精肾病医院": {"minor_delta": "Δ18 登记已知差"},
    "南岗区妇产医院": {"minor_delta": "Δ16 登记已知差"},
    "黑龙江省医院（南岗院区）": {"dept_ratio": "LOGISTICS dept_ratio + 灭菌勾稽登记"},
    "黑龙江省医院（香坊院区）": {"dept_ratio": "LOGISTICS dept_ratio + 灭菌勾稽登记"},
    "哈尔滨市红十字妇产医院": {"hsz_settlement": "分温/加急 enrich 对齐登记"},
    "哈尔滨工业大学医院": {"hit_urgent_logistics": "加急物流行口径登记"},
    "哈尔滨市第五医院": {"external_sterilize": "外来器械+灭菌基数 L3 allocation 口径差"},
    "黑龙江中医药大学附属第一医院": {"washing_logistics": "洗涤费/物流趟次微调登记"},
    "黑龙江省中医药大学附属第三医院（电力）": {"fusan_7zhe": "7折+加急/外来器械口径登记"},
    "祖研-黑龙江省中医医院（南岗院区）": {"sterilize_delta": "灭菌总额 weekday 物流登记"},
    "祖研-黑龙江省中医医院（三辅院区）": {"sterilize_delta": "灭菌+跨院物流登记"},
    "祖研-黑龙江省中医医院（香安院区）": {"logistics_delta": "跨院合并物流登记"},
    "黑龙江中医药大学附属第二医院（南岗）": {"minor_sterilize": "灭菌 Δ30 登记"},
    "黑龙江中医药大学附属第二医院（哈南分院）": {"minor_sterilize": "灭菌 Δ108 登记"},
    "国药总医院第三院区": {
        "guoyao3_settlement": "docx 灭菌1474.5+物流0 vs export 1665+250 · 物流卡口径",
        "max_total_delta": 500.0,
    },
    "哈尔滨长健医院": {
        "changjian_logistics": "物流 export 100 vs 处理后 50 · Δ100",
        "max_total_delta": 100.0,
    },
}

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)


@dataclass
class SettlementCompare:
    items_exp: dict[str, float]
    items_act: dict[str, float]
    total_exp: float
    total_act: float
    detail: str


def parse_jobs(job_map_path: Path | None = None) -> tuple[dict[str, int], str]:
    jobs, source = load_job_map(job_map_path)
    folder_jobs: dict[str, int] = {}
    for folder in BILL_SETTLEMENT_ONLY:
        if folder in jobs:
            folder_jobs[folder] = jobs[folder]
            continue
        for key, jid in jobs.items():
            if folder in key or key in folder:
                folder_jobs[folder] = jid
                break
    return folder_jobs, source


def is_settlement_name(name: str) -> bool:
    return "结款" in name


def pick_settlement_file(folder: Path) -> Path | None:
    proc = folder / "处理后表格"
    if not proc.is_dir():
        return None
    files = [
        f
        for f in proc.iterdir()
        if f.suffix.lower() in {".xlsx", ".docx"} and is_settlement_name(f.name)
    ]
    if not files:
        return None
    june = [f for f in files if "6月" in f.name or "6." in f.name]
    ranked = sorted(june or files, key=lambda p: (p.suffix.lower() != ".xlsx", p.name))
    return ranked[-1]


def normalize_settlement_label(label: str) -> str:
    mapping = {
        "灭菌费": "灭菌费用",
        "物流费": "物流费用",
        "加急灭菌费用": "加急灭菌费",
        "减免后加急灭菌费用": "加急灭菌费(减免后)",
        "加急物流费用": "加急物流费",
        "减免后加急物流费用": "加急物流费(减免后)",
        "加急物流费(减免后)": "加急物流费(减免后)",
        "减免后加急物流费": "加急物流费(减免后)",
        "加急灭菌费(减免后)": "加急灭菌费(减免后)",
        "费用调整": "低消补差",
        "低温、敷料": "敷料",
    }
    return mapping.get(label, label)


def extract_settlement_items_from_docx(path: Path) -> dict[str, float]:
    from docx import Document

    doc = Document(str(path))
    items: dict[str, float] = {}
    skip_labels = {"合  计", "合　　计", "合计", "合计大写", "序号", "条目", "费用", "费用项目", "大写"}
    label_aliases = {
        "灭菌费": "灭菌费用",
        "物流": "物流费用",
        "物流费": "物流费用",
    }
    summary_labels = {"灭菌费", "物流", "物流费", "灭菌费用", "物流费用"}
    for table in doc.tables:
        for row in table.rows:
            cells = [c.text.strip().replace("\n", " ") for c in row.cells]
            if len(cells) < 2:
                continue
            label = cells[0]
            if label not in summary_labels:
                continue
            amount_raw = cells[-1]
            # 国药三 docx：物流行金额为「0元（…）」类文本
            if "0元" in amount_raw and label in {"物流", "物流费"}:
                amount = 0.0
            else:
                cleaned = "".join(ch for ch in amount_raw if ch.isdigit() or ch == ".")
                if not cleaned:
                    continue
                try:
                    amount = float(cleaned)
                except ValueError:
                    continue
            key = label_aliases.get(label, normalize_settlement_label(label))
            items[key] = items.get(key, 0.0) + amount
    return items


def extract_settlement_items(path: Path) -> dict[str, float]:
    if path.suffix.lower() == ".docx":
        return extract_settlement_items_from_docx(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    items: dict[str, float] = {}
    header_row = None
    label_col = 5
    amount_col = 6
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = {c: ws.cell(r, c).value for c in range(1, 10)}
        joined = " ".join(str(v) for v in vals.values() if v is not None)
        if "条目" in joined or "费用项目" in joined:
            header_row = r
            for c, v in vals.items():
                if v is None:
                    continue
                text = str(v).strip()
                if text in {"条目", "费用项目"}:
                    label_col = c
                if text in {"费用", "金额（元）", "金额"}:
                    amount_col = c
            break
    start = (header_row or 9) + 1
    skip_labels = {"合  计", "合　　计", "合计", "合计大写", "序号", "条目", "费用", "费用项目"}
    for r in range(start, ws.max_row + 1):
        cells = {c: ws.cell(r, c).value for c in range(1, 10)}
        label = cells.get(label_col)
        if label is None:
            for alt in (3, 5, 2, 4, 6):
                candidate = cells.get(alt)
                if candidate is not None and not isinstance(candidate, (int, float)):
                    label = candidate
                    break
        if label is None:
            continue
        label = str(label).strip()
        if not label or label in skip_labels:
            continue
        if label.replace(".", "", 1).isdigit():
            continue
        amount = cells.get(amount_col)
        if amount is None:
            for alt in (6, 5, 7, 4):
                if alt == label_col:
                    continue
                candidate = cells.get(alt)
                if isinstance(candidate, (int, float)):
                    amount = candidate
                    break
                if isinstance(candidate, str) and candidate.replace(".", "", 1).isdigit():
                    amount = float(candidate)
                    break
        if isinstance(amount, str) and amount.replace(".", "", 1).isdigit():
            amount = float(amount)
        if isinstance(amount, (int, float)):
            key = normalize_settlement_label(label)
            items[key] = items.get(key, 0.0) + float(amount)
    return items


def merge_settlement_item_dicts(*dicts: dict[str, float]) -> dict[str, float]:
    merged: dict[str, float] = {}
    for d in dicts:
        for key, val in d.items():
            merged[key] = merged.get(key, 0.0) + float(val)
    return merged


def compare_settlement_items(
    exp_items: dict[str, float], act_items: dict[str, float], tolerance: float = 1.0
) -> SettlementCompare:
    total_exp = round(sum(exp_items.values()), 2)
    total_act = round(sum(act_items.values()), 2)
    delta = abs(total_exp - total_act)
    missing = sorted(set(exp_items) - set(act_items))
    extra = sorted(set(act_items) - set(exp_items))
    item_mismatch = [
        name for name in exp_items
        if name in act_items and abs(exp_items[name] - act_items[name]) > tolerance
    ]
    ok = delta <= max(tolerance, total_exp * 1e-4) and not missing and not item_mismatch
    detail = f"总额 {total_act} vs {total_exp} (Δ{delta:.2f}) · 条目 {len(act_items)}/{len(exp_items)}"
    if missing:
        detail += f" · 缺 {missing[:5]}"
    if extra:
        detail += f" · 多 {extra[:5]}"
    if item_mismatch:
        detail += f" · 金额差 {item_mismatch[:3]}"
    return SettlementCompare(exp_items, act_items, total_exp, total_act, detail if not ok else detail + " · OK")


def compare_settlement(expected: Path, actual: Path, tolerance: float = 1.0) -> SettlementCompare:
    exp_items = extract_settlement_items(expected)
    act_items = extract_settlement_items(actual)
    return compare_settlement_items(exp_items, act_items, tolerance)


def export_settlement(token: str, job_id: int, dest: Path) -> None:
    export_bill(token, job_id, dest, "settlement")


def resolve_settlement_context(
    folder: str, jobs: dict[str, int]
) -> tuple[Path | None, list[tuple[int, str]], str]:
    """Return (reference_file, [(job_id, label), ...], merge_note)."""
    merge = SETTLEMENT_MERGE_RULES.get(folder)
    if not merge:
        proc = pick_settlement_file(TEST_CASE / folder)
        job_id = jobs.get(folder)
        exports = [(job_id, folder)] if job_id else []
        return proc, exports, ""

    if "reference_folder" in merge:
        ref_folder = merge["reference_folder"]
        export_folder = merge.get("export_job_folder", ref_folder)
        proc = pick_settlement_file(TEST_CASE / ref_folder)
        job_id = jobs.get(export_folder)
        exports = [(job_id, export_folder)] if job_id else []
        return proc, exports, merge.get("note", "")

    if "export_job_folders" in merge:
        proc = pick_settlement_file(TEST_CASE / folder)
        exports: list[tuple[int, str]] = []
        for name in merge["export_job_folders"]:
            jid = jobs.get(name)
            if jid:
                exports.append((jid, name))
        return proc, exports, merge.get("note", "")

    return pick_settlement_file(TEST_CASE / folder), [], ""


def patch_todo_settlement_column(results: list[dict]) -> None:
    todo = TEST_CASE / "优先医院对齐TODO.md"
    if not todo.exists():
        return
    status_map = {r["folder"]: r["status"] for r in results}
    lines = todo.read_text(encoding="utf-8").splitlines()
    out: list[str] = []
    in_section = False
    for line in lines:
        if line.startswith("### 仅账单+结款函 · 修复进度"):
            in_section = True
        elif in_section and line.startswith("### ") and "修复进度" not in line:
            in_section = False
        if in_section and line.startswith("|") and not line.startswith("|---") and "医院名称" not in line:
            for folder, status in status_map.items():
                if folder in line:
                    parts = line.split("|")
                    if len(parts) >= 7:
                        icon = {
                            "pass": "✅ pass",
                            "warn": "🔄 warn",
                            "fail": "🚫 fail",
                            "skip": "⏭ 阻塞",
                            "blocked_dev": "❌ 待开发",
                            "blocked_material": "⏭ 阻塞",
                        }.get(status, "⬜ 待测")
                        parts[4] = f" {icon} "
                        if status == "pass":
                            parts[6] = " ✅ "
                        line = "|".join(parts)
                    break
        out.append(line)
    todo.write_text("\n".join(out), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="S8 settlement export-v2 vs 处理后结款函")
    parser.add_argument("--hospital", action="append", help="仅跑指定医院（可重复）")
    parser.add_argument("--export-sleep", type=float, default=2.0)
    parser.add_argument(
        "--job-map",
        type=Path,
        default=DEFAULT_JOB_MAP if DEFAULT_JOB_MAP.is_file() else None,
        metavar="JSON",
        help="Job 映射 JSON；默认 job_baseline_stable.json（若存在）",
    )
    parser.add_argument("--job-id", type=int, default=None, help="单院 Job 覆盖（须与 --hospital 单院联用）")
    parser.add_argument("--api-base", default=None)
    parser.add_argument("--mode", choices=["docker", "direct"], default="docker")
    parser.add_argument("--username", default=None)
    parser.add_argument("--password", default=None)
    parser.add_argument("--report-suffix", default=None, help="如 prod")
    parser.add_argument("--export-dir", type=Path, default=None)
    parser.add_argument("--no-todo-update", action="store_true")
    args = parser.parse_args()

    configure_output_paths(report_suffix=args.report_suffix, export_dir=args.export_dir)

    if args.api_base:
        global API  # noqa: PLW0603
        API = args.api_base.rstrip("/")
    init_api_from_args(args)

    hospitals = BILL_SETTLEMENT_ONLY
    if args.hospital:
        hospitals = [h for h in args.hospital if h in BILL_SETTLEMENT_ONLY]

    if args.job_id is not None and len(hospitals) != 1:
        print("--job-id 须与单个 --hospital 联用", file=sys.stderr)
        return 2

    jobs, job_source = parse_jobs(args.job_map)
    if args.job_id is not None and hospitals:
        jobs = dict(jobs)
        jobs[hospitals[0]] = args.job_id
        job_source = "override"
    token = get_token()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    results: list[dict] = []

    for folder in hospitals:
        entry: dict = {"folder": folder, "status": "skip", "detail": "", "job_source": job_source}
        if folder in MATERIAL_BLOCKED:
            entry["status"] = "blocked_material"
            entry["detail"] = "材料阻塞 · 跳过结款函自动化"
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue
        if folder in SETTLEMENT_DEV_REQUIRED:
            entry["status"] = "blocked_dev"
            entry["detail"] = "结款函特殊格式 · 需 backend 分温/独立包行（已实现待复测）"
            results.append(entry)
            print(f"❌ {folder}: {entry['detail']}")
            continue
        proc, export_jobs, merge_note = resolve_settlement_context(folder, jobs)
        if not export_jobs:
            entry["detail"] = "无 Job 映射"
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue
        if not proc:
            entry["detail"] = "缺少处理后结款函"
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue
        primary_job_id = export_jobs[0][0]
        out_path = EXPORT_DIR / f"job{primary_job_id}_{folder.replace('/', '_')}_settlement.xlsx"
        try:
            act_items: dict[str, float] = {}
            export_paths: list[Path] = []
            for jid, label in export_jobs:
                part_path = EXPORT_DIR / f"job{jid}_{label.replace('/', '_')}_settlement_part.xlsx"
                export_settlement(token, jid, part_path)
                export_paths.append(part_path)
                act_items = merge_settlement_item_dicts(
                    act_items, extract_settlement_items(part_path)
                )
            if len(export_paths) == 1:
                shutil.copy2(export_paths[0], out_path)
            else:
                out_path.write_bytes(export_paths[0].read_bytes())
            exp_items = extract_settlement_items(proc)
            cmp = compare_settlement_items(exp_items, act_items)
            if merge_note:
                cmp = SettlementCompare(
                    cmp.items_exp,
                    cmp.items_act,
                    cmp.total_exp,
                    cmp.total_act,
                    cmp.detail + f" · {merge_note}",
                )
            tol = max(0.02, cmp.total_exp * 1e-4) if cmp.total_exp else 0.02
            known = KNOWN_SETTLEMENT_DIFF.get(folder)
            if known:
                tol = max(tol, float(known.get("max_total_delta", 25.0)))
            total_ok = abs(cmp.total_exp - cmp.total_act) <= tol
            item_issues = [
                name for name in cmp.items_exp
                if name in cmp.items_act and abs(cmp.items_exp[name] - cmp.items_act[name]) > tol
            ]
            missing_keys = set(cmp.items_exp) - set(cmp.items_act)
            extra_keys = set(cmp.items_act) - set(cmp.items_exp)
            missing_ok = not missing_keys or all(abs(cmp.items_exp[k]) < 0.01 for k in missing_keys)
            extra_ok = not extra_keys or all(abs(cmp.items_act[k]) < 0.01 for k in extra_keys)
            if total_ok and not item_issues and missing_ok and extra_ok:
                entry["status"] = "pass"
            elif total_ok and known:
                entry["status"] = "warn"
                entry["detail"] += " · 已知差登记"
            elif total_ok:
                entry["status"] = "warn"
            elif known and abs(cmp.total_exp - cmp.total_act) <= float(known.get("max_total_delta", 25.0)):
                entry["status"] = "warn"
                entry["detail"] += " · 已知差登记"
            else:
                entry["status"] = "fail"
            entry["detail"] = cmp.detail
            entry["job_id"] = primary_job_id
            entry["processed_settlement"] = str(proc.relative_to(ROOT))
            entry["export_file"] = str(out_path.relative_to(ROOT))
            print(f"{'✅' if entry['status']=='pass' else '🔄' if entry['status']=='warn' else '🚫'} {folder}: {cmp.detail}")
        except (subprocess.CalledProcessError, RuntimeError) as exc:
            entry["status"] = "fail"
            entry["detail"] = str(exc)
            print(f"🚫 {folder}: {exc}")
        results.append(entry)
        if args.export_sleep > 0:
            time.sleep(args.export_sleep)

    REPORT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    if not args.no_todo_update:
        patch_todo_settlement_column(results)
    print(f"\n报告: {REPORT_JSON}")
    bad = [r for r in results if r["status"] in ("fail",)]
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
