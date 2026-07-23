#!/usr/bin/env python3
"""Compare raw vs processed hospital bill Excel files under 测试用例/."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[1]
TEST_CASE_DIR = ROOT / "测试用例"
TOLERANCE = 0.011
DATA_EXTS = {".xlsx", ".xls"}

MONTH_PATTERN = re.compile(r"(\d{1,2})月")
DATE_RANGE_PATTERN = re.compile(r"(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})")
PROC_PREFIX_PATTERN = re.compile(r"^(\d{1,2})月__")

# 文件名含以下词则肯定不是主账单
BILL_HARD_EXCLUDE = (
    "结款",
    "结款函",
    "结款涵",
    "器械把数",
    "器械表",
    "灭菌",
    "物流",
    "洗涤",
    "把数表",
    "分科室",
    "总汇总",
)

# 处理后_workbook 中不参与逐行对比的 Sheet（汇总/调整类）
PROC_META_SHEETS = {"费用调整", "异常物流", "Sheet1"}

# 存在 consolidated「账单」Sheet 时，额外纳入对比的补充 Sheet
PROC_SUPPLEMENT_SHEETS = {"外来器械", "加急", "外来加急"}

DETAIL_HEADERS = ("发货日期", "发货单号", "包类别号", "包名", "包数", "单价", "总价")
OPTIONAL_HEADERS = ("类型", "包装材料", "器械数")


@dataclass
class DetailRow:
    sheet: str
    excel_row: int
    ship_date: str
    ship_no: str
    pack_code: str
    pack_name: str
    pack_type: str | None
    pack_count: float | None
    unit_price: float | None
    total_price: float | None
    material: str | None
    instrument_count: float | None

    def key(self) -> tuple[str, str, str, str]:
        return (self.ship_date, self.ship_no, self.pack_code, self.pack_name)


@dataclass
class SheetSummary:
    sheet: str
    excel_row: int
    dept_name: str
    pack_count: float | None
    total_price: float | None


@dataclass
class ParsedWorkbook:
    path: Path
    sheets: dict[str, list[DetailRow]] = field(default_factory=dict)
    summaries: dict[str, list[SheetSummary]] = field(default_factory=dict)
    header_row: int | None = None
    raw_col_count: int = 0
    proc_col_count: int = 0
    parse_errors: list[str] = field(default_factory=list)


@dataclass
class FieldDiff:
    month: str
    sheet: str
    issue_type: str
    pack_name: str
    ship_no: str
    field: str
    raw_value: Any
    proc_value: Any
    raw_row: int | None
    proc_row: int | None
    note: str = ""


@dataclass
class MonthComparison:
    month: str
    raw_file: Path | None
    proc_bill: Path | None
    proc_settlement: Path | None
    raw_rows: int = 0
    proc_rows: int = 0
    missing_rows: int = 0
    extra_rows: int = 0
    field_diffs: int = 0
    amount_delta: float | None = None
    issues: list[FieldDiff] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class HospitalReport:
    name: str
    months: list[MonthComparison] = field(default_factory=list)
    all_issues: list[FieldDiff] = field(default_factory=list)
    failed: bool = False
    error: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if " " in text:
        text = text.split(" ", 1)[0]
    return text


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:g}"


def extract_month_from_name(name: str) -> int | None:
    prefix = PROC_PREFIX_PATTERN.match(name)
    if prefix:
        return int(prefix.group(1))
    match = MONTH_PATTERN.search(name)
    if match:
        return int(match.group(1))
    dr = DATE_RANGE_PATTERN.search(name)
    if dr:
        # 账期区间如 4.15-5.14 对应 4 月账单
        return int(dr.group(1))
    return None


def extract_date_range_token(name: str) -> str | None:
    match = DATE_RANGE_PATTERN.search(name)
    if not match:
        return None
    return match.group(0)


def is_bill_file(name: str) -> bool:
    lower = name.lower()
    if not lower.endswith((".xlsx", ".xls")):
        return False
    if is_settlement_file(name):
        return False
    if any(k in name for k in BILL_HARD_EXCLUDE):
        return False
    if "账单" in name or "bill" in lower or DATE_RANGE_PATTERN.search(name) is not None:
        return True
    # 如 4月__省二院（南岗）.xlsx、6月__黑龙江维多利亚妇产医院6月.xlsx
    return PROC_PREFIX_PATTERN.match(name) is not None


def is_settlement_file(name: str) -> bool:
    return "结款" in name


def load_workbook_safe(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise RuntimeError(f"读取 .xls 需要 xlrd: {path.name}") from exc
        book = xlrd.open_workbook(path)
        # openpyxl-only path for now; convert via pandas if needed
        raise RuntimeError(f".xls 暂不支持直接解析: {path.name}")
    return load_workbook(path, data_only=True)


def find_header_row(ws, max_scan: int = 30) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, min(max_scan, (ws.max_row or 0) + 1)):
        headers: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            text = normalize_text(ws.cell(row=row_idx, column=col).value)
            if text in {
                "发货日期",
                "发货单号",
                "类型",
                "包类别号",
                "包名",
                "包装材料",
                "包数",
                "器械数",
                "单价",
                "总价",
            }:
                headers[text] = col
        if {"发货日期", "发货单号", "包名"}.issubset(headers):
            return row_idx, headers
    return None


def row_texts(ws, row_idx: int, max_col: int | None = None) -> list[str]:
    limit = max_col or (ws.max_column or 0)
    return [normalize_text(ws.cell(row=row_idx, column=c).value) for c in range(1, limit + 1)]


def is_detail_row(values: dict[str, Any]) -> bool:
    ship_no = values.get("发货单号")
    pack_name = normalize_text(values.get("包名"))
    if not pack_name:
        return False
    if ship_no is None or ship_no == "":
        return False
    if isinstance(ship_no, str) and not ship_no.isdigit():
        return False
    return True


def is_dept_summary_row(values: dict[str, Any], headers: dict[str, int]) -> bool:
    ship_date_col = headers.get("发货日期")
    ship_no = values.get("发货单号")
    pack_name = normalize_text(values.get("包名"))
    dept_name = normalize_text(values.get("发货日期"))
    total = to_float(values.get("总价"))
    pack_count = to_float(values.get("包数"))
    if ship_no not in (None, ""):
        return False
    if pack_name:
        return False
    if not dept_name:
        return False
    if total is None and pack_count is None:
        return False
    # Exclude header repeats
    if dept_name in {"发货日期", "类型", "包类别号", "包名"}:
        return False
    return True


def parse_workbook(path: Path) -> ParsedWorkbook:
    result = ParsedWorkbook(path=path)
    try:
        wb = load_workbook_safe(path)
    except Exception as exc:
        result.parse_errors.append(str(exc))
        return result

    result.raw_col_count = max((ws.max_column or 0) for ws in wb.worksheets) if wb.worksheets else 0

    for sheet_name in wb.sheetnames:
        if sheet_name in PROC_META_SHEETS:
            continue
        ws = wb[sheet_name]
        header_info = find_header_row(ws)
        if not header_info:
            result.parse_errors.append(f"[{sheet_name}] 未找到表头行")
            continue
        header_row, headers = header_info
        if result.header_row is None:
            result.header_row = header_row

        details: list[DetailRow] = []
        summaries: list[SheetSummary] = []

        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            values: dict[str, Any] = {}
            for name, col in headers.items():
                values[name] = ws.cell(row=row_idx, column=col).value

            if is_detail_row(values):
                details.append(
                    DetailRow(
                        sheet=sheet_name,
                        excel_row=row_idx,
                        ship_date=format_date(values.get("发货日期")),
                        ship_no=normalize_text(values.get("发货单号")),
                        pack_code=normalize_text(values.get("包类别号")),
                        pack_name=normalize_text(values.get("包名")),
                        pack_type=normalize_text(values.get("类型")) or None,
                        pack_count=to_float(values.get("包数")),
                        unit_price=to_float(values.get("单价")),
                        total_price=to_float(values.get("总价")),
                        material=normalize_text(values.get("包装材料")) or None,
                        instrument_count=to_float(values.get("器械数")),
                    )
                )
                continue

            if is_dept_summary_row(values, headers):
                summaries.append(
                    SheetSummary(
                        sheet=sheet_name,
                        excel_row=row_idx,
                        dept_name=normalize_text(values.get("发货日期")),
                        pack_count=to_float(values.get("包数")),
                        total_price=to_float(values.get("总价")),
                    )
                )

        result.sheets[sheet_name] = details
        result.summaries[sheet_name] = summaries

    result.proc_col_count = result.raw_col_count
    return result


def nums_close(a: float | None, b: float | None, tol: float = TOLERANCE) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def price_note(raw: DetailRow, proc: DetailRow, field: str) -> str:
    if field not in {"单价", "总价"}:
        return ""
    if raw.material and raw.instrument_count and raw.unit_price and proc.unit_price:
        if not nums_close(raw.unit_price, proc.unit_price):
            return (
                f"原始按器械数×基础单价({format_number(raw.unit_price)})计费；"
                f"处理后按标准价({format_number(proc.unit_price)})计费"
            )
    return ""


def compare_sheet_details(
    month: str,
    sheet: str,
    raw_rows: list[DetailRow],
    proc_rows: list[DetailRow],
) -> list[FieldDiff]:
    issues: list[FieldDiff] = []
    raw_map = {r.key(): r for r in raw_rows}
    proc_map = {r.key(): r for r in proc_rows}

    for key, raw in raw_map.items():
        proc = proc_map.get(key)
        if not proc:
            issues.append(
                FieldDiff(
                    month=month,
                    sheet=sheet,
                    issue_type="缺失行",
                    pack_name=raw.pack_name,
                    ship_no=raw.ship_no,
                    field="整行",
                    raw_value=f"{raw.pack_count}/{raw.unit_price}/{raw.total_price}",
                    proc_value="",
                    raw_row=raw.excel_row,
                    proc_row=None,
                )
            )
            continue
        for field_name, getter in (
            ("包数", lambda r: r.pack_count),
            ("单价", lambda r: r.unit_price),
            ("总价", lambda r: r.total_price),
        ):
            rv, pv = getter(raw), getter(proc)
            if not nums_close(rv, pv):
                issues.append(
                    FieldDiff(
                        month=month,
                        sheet=sheet,
                        issue_type=f"字段不一致-{field_name}",
                        pack_name=raw.pack_name,
                        ship_no=raw.ship_no,
                        field=field_name,
                        raw_value=rv,
                        proc_value=pv,
                        raw_row=raw.excel_row,
                        proc_row=proc.excel_row,
                        note=price_note(raw, proc, field_name),
                    )
                )

    for key, proc in proc_map.items():
        if key not in raw_map:
            issues.append(
                FieldDiff(
                    month=month,
                    sheet=sheet,
                    issue_type="多余行",
                    pack_name=proc.pack_name,
                    ship_no=proc.ship_no,
                    field="整行",
                    raw_value="",
                    proc_value=f"{proc.pack_count}/{proc.unit_price}/{proc.total_price}",
                    raw_row=None,
                    proc_row=proc.excel_row,
                )
            )

    return issues


def compare_summaries(
    month: str,
    sheet: str,
    raw_summaries: list[SheetSummary],
    proc_summaries: list[SheetSummary],
) -> list[FieldDiff]:
    issues: list[FieldDiff] = []
    raw_map = {s.dept_name: s for s in raw_summaries}
    proc_map = {s.dept_name: s for s in proc_summaries}
    for dept, raw in raw_map.items():
        proc = proc_map.get(dept)
        if not proc:
            continue
        if not nums_close(raw.total_price, proc.total_price):
            issues.append(
                FieldDiff(
                    month=month,
                    sheet=sheet,
                    issue_type="汇总不一致-总价",
                    pack_name=dept,
                    ship_no="",
                    field="总价",
                    raw_value=raw.total_price,
                    proc_value=proc.total_price,
                    raw_row=raw.excel_row,
                    proc_row=proc.excel_row,
                )
            )
    return issues


def flatten_raw_rows(wb: ParsedWorkbook) -> list[DetailRow]:
    rows: list[DetailRow] = []
    for sheet_rows in wb.sheets.values():
        rows.extend(sheet_rows)
    return rows


def flatten_raw_summaries(wb: ParsedWorkbook) -> list[SheetSummary]:
    rows: list[SheetSummary] = []
    for sheet_summaries in wb.summaries.values():
        rows.extend(sheet_summaries)
    return rows


def collect_proc_rows(wb: ParsedWorkbook) -> list[DetailRow]:
    if "账单" in wb.sheets and wb.sheets["账单"]:
        rows = list(wb.sheets["账单"])
        for name in PROC_SUPPLEMENT_SHEETS:
            rows.extend(wb.sheets.get(name, []))
        return rows
    rows: list[DetailRow] = []
    for sheet_name, sheet_rows in wb.sheets.items():
        if sheet_name in PROC_META_SHEETS:
            continue
        rows.extend(sheet_rows)
    return rows


def collect_proc_summaries(wb: ParsedWorkbook) -> list[SheetSummary]:
    if "账单" in wb.summaries:
        rows = list(wb.summaries["账单"])
        for name in PROC_SUPPLEMENT_SHEETS:
            rows.extend(wb.summaries.get(name, []))
        return rows
    rows: list[SheetSummary] = []
    for sheet_name, sheet_summaries in wb.summaries.items():
        if sheet_name in PROC_META_SHEETS:
            continue
        rows.extend(sheet_summaries)
    return rows


def should_flatten_compare(raw: ParsedWorkbook, proc: ParsedWorkbook) -> bool:
    if "账单" in proc.sheets and proc.sheets["账单"]:
        return True
    proc_only = set(proc.sheets) - set(raw.sheets) - PROC_META_SHEETS
    if proc_only & PROC_SUPPLEMENT_SHEETS:
        return True
    return sheet_overlap_ratio(raw, proc) < 0.35


def sheet_overlap_ratio(raw: ParsedWorkbook, proc: ParsedWorkbook) -> float:
    raw_names = {s for s, rows in raw.sheets.items() if rows}
    proc_names = {s for s, rows in proc.sheets.items() if rows} - {"账单"} - PROC_SUPPLEMENT_SHEETS
    if not raw_names or not proc_names:
        return 0.0
    return len(raw_names & proc_names) / len(raw_names)


def compare_workbooks(month: str, raw: ParsedWorkbook, proc: ParsedWorkbook) -> MonthComparison:
    comp = MonthComparison(
        month=month,
        raw_file=raw.path,
        proc_bill=proc.path,
        proc_settlement=None,
    )
    comp.warnings.extend(raw.parse_errors)
    comp.warnings.extend(proc.parse_errors)

    use_flat = should_flatten_compare(raw, proc)
    if use_flat:
        raw_rows = flatten_raw_rows(raw)
        proc_rows = collect_proc_rows(proc)
        comp.raw_rows = len(raw_rows)
        comp.proc_rows = len(proc_rows)
        compare_label = "全部科室(汇总对比)"
        comp.issues.extend(compare_sheet_details(month, compare_label, raw_rows, proc_rows))
        comp.issues.extend(
            compare_summaries(
                month,
                compare_label,
                flatten_raw_summaries(raw),
                collect_proc_summaries(proc),
            )
        )
        raw_total = sum(r.total_price or 0 for r in raw_rows)
        proc_total = sum(r.total_price or 0 for r in proc_rows)
    else:
        raw_total = 0.0
        proc_total = 0.0
        for sheet in sorted(set(raw.sheets) | set(proc.sheets)):
            raw_rows = raw.sheets.get(sheet, [])
            proc_rows = proc.sheets.get(sheet, [])
            comp.raw_rows += len(raw_rows)
            comp.proc_rows += len(proc_rows)
            comp.issues.extend(compare_sheet_details(month, sheet, raw_rows, proc_rows))
            comp.issues.extend(
                compare_summaries(
                    month,
                    sheet,
                    raw.summaries.get(sheet, []),
                    proc.summaries.get(sheet, []),
                )
            )
            raw_total += sum(r.total_price or 0 for r in raw_rows)
            proc_total += sum(r.total_price or 0 for r in proc_rows)

    comp.missing_rows = sum(1 for i in comp.issues if i.issue_type == "缺失行")
    comp.extra_rows = sum(1 for i in comp.issues if i.issue_type == "多余行")
    comp.field_diffs = sum(1 for i in comp.issues if i.issue_type.startswith("字段不一致"))
    if abs(raw_total - proc_total) > TOLERANCE:
        comp.amount_delta = round(proc_total - raw_total, 2)
    return comp


def pick_processed_bill(month: int, files: list[Path], raw_name: str) -> Path | None:
    month_prefix = f"{month}月__"
    candidates = [f for f in files if f.name.startswith(month_prefix) and is_bill_file(f.name)]
    if not candidates:
        return None

    raw_range = extract_date_range_token(raw_name)
    if raw_range:
        matched = [f for f in candidates if raw_range in f.name]
        if matched:
            candidates = matched

    def rank(p: Path) -> tuple[int, int, str]:
        name = p.name
        score = 0
        if "账单" in name:
            score += 10
        if "供应室" in name:
            score += 3
        if "财务" in name:
            score -= 4
        if re.search(r"\(\d+\)", name):
            score -= 2
        if DATE_RANGE_PATTERN.search(name):
            score += 2
        return (-score, len(name), name)

    candidates.sort(key=rank)
    return candidates[0]


def pick_settlement(month: int, files: list[Path]) -> Path | None:
    month_prefix = f"{month}月__"
    candidates = [f for f in files if f.name.startswith(month_prefix) and is_settlement_file(f.name)]
    if not candidates:
        return None
    candidates.sort(key=lambda p: (re.search(r"\(\d+\)", p.name) is not None, len(p.name), p.name))
    return candidates[0]


def match_raw_processed(raw_files: list[Path], proc_files: list[Path]) -> dict[int, tuple[Path, Path | None, Path | None]]:
    mapping: dict[int, tuple[Path, Path | None, Path | None]] = {}
    for raw in sorted(raw_files):
        month = extract_month_from_name(raw.name)
        if month is None:
            continue
        bill = pick_processed_bill(month, proc_files, raw.name)
        settlement = pick_settlement(month, proc_files)
        mapping[month] = (raw, bill, settlement)
    return mapping


def discover_hospitals(base: Path) -> list[Path]:
    hospitals = []
    for d in sorted(base.iterdir()):
        if not d.is_dir():
            continue
        if (d / "原始表格").is_dir() and (d / "处理后表格").is_dir():
            hospitals.append(d)
    return hospitals


def analyze_hospital(hospital_dir: Path) -> HospitalReport:
    name = hospital_dir.name
    report = HospitalReport(name=name)
    raw_dir = hospital_dir / "原始表格"
    proc_dir = hospital_dir / "处理后表格"

    raw_files = [p for p in raw_dir.iterdir() if p.suffix.lower() in DATA_EXTS]
    proc_files = [p for p in proc_dir.iterdir() if p.suffix.lower() in DATA_EXTS]

    if not raw_files:
        report.failed = True
        report.error = "原始表格为空"
        return report

    month_map = match_raw_processed(raw_files, proc_files)

    for month in sorted(month_map):
        raw_path, proc_bill, proc_settlement = month_map[month]
        month_label = f"{month}月"
        if proc_bill is None:
            mc = MonthComparison(
                month=month_label,
                raw_file=raw_path,
                proc_bill=None,
                proc_settlement=proc_settlement,
                warnings=[f"未找到 {month_label} 对应的处理后账单文件"],
            )
            report.months.append(mc)
            continue

        try:
            raw_wb = parse_workbook(raw_path)
            proc_wb = parse_workbook(proc_bill)
            mc = compare_workbooks(month_label, raw_wb, proc_wb)
            mc.proc_settlement = proc_settlement
        except Exception as exc:
            report.failed = True
            report.error = f"{month_label}: {exc}"
            mc = MonthComparison(
                month=month_label,
                raw_file=raw_path,
                proc_bill=proc_bill,
                proc_settlement=proc_settlement,
                warnings=[str(exc)],
            )
        report.months.append(mc)
        report.all_issues.extend(mc.issues)

    return report


def issue_count(report: HospitalReport) -> int:
    return len(report.all_issues)


def month_has_data_issues(mc: MonthComparison) -> bool:
    return bool(mc.issues) or mc.missing_rows or mc.extra_rows or mc.field_diffs


def render_markdown(report: HospitalReport, analysis_date: str) -> str:
    lines: list[str] = []
    lines.append(f"# {report.name} — 原始表格数据问题分析")
    lines.append("")
    lines.append(f"> 分析日期：{analysis_date}")
    lines.append("> 对比方法：以「处理后表格」为基准（ground truth），逐行匹配原始表格中的发货明细")
    lines.append("> 匹配键：发货日期 + 发货单号 + 包类别号 + 包名")
    lines.append("")
    lines.append("## 一、文件清单")
    lines.append("")
    lines.append("| 月份 | 原始表格 | 处理后表格（账单） | 处理后表格（结款函） |")
    lines.append("|------|----------|-------------------|-------------------|")
    for mc in report.months:
        raw_name = mc.raw_file.name if mc.raw_file else "—"
        bill_name = mc.proc_bill.name if mc.proc_bill else "—（缺失）"
        settle_name = mc.proc_settlement.name if mc.proc_settlement else "—"
        lines.append(f"| {mc.month} | {raw_name} | {bill_name} | {settle_name} |")
    lines.append("")
    lines.append("**说明：** 原始表格仅含账单明细，不含结款函；结款函仅存在于处理后表格中，为系统导出产物。")
    lines.append("")
    lines.append("## 二、总体对比摘要")
    lines.append("")
    lines.append("| 月份 | 原始行数 | 处理后行数 | 缺失行 | 多余行 | 字段不一致 | 金额差异 |")
    lines.append("|------|---------|-----------|--------|--------|-----------|---------|")
    total_issues = 0
    for mc in report.months:
        delta = "无" if mc.amount_delta is None else f"{mc.amount_delta:+.2f}元"
        if mc.proc_bill is None:
            lines.append(f"| {mc.month} | — | — | — | — | — | 缺少处理后账单 |")
            continue
        lines.append(
            f"| {mc.month} | {mc.raw_rows} | {mc.proc_rows} | {mc.missing_rows} | "
            f"{mc.extra_rows} | {mc.field_diffs} | {delta} |"
        )
        total_issues += len(mc.issues)
    lines.append("")
    if total_issues == 0 and not report.failed:
        lines.append("**结论：** 所有可对比月份的原始数据与处理后表格完全一致。")
    elif total_issues > 0:
        issue_months = [mc.month for mc in report.months if month_has_data_issues(mc)]
        if issue_months:
            lines.append(
                f"**结论：** {', '.join(issue_months)} 存在数据差异，详见下文明细。"
            )
    if report.failed and report.error:
        lines.append("")
        lines.append(f"**警告：** 分析过程中出现错误：{report.error}")
    lines.append("")

    # Structure section from first comparable month
    comparable = [mc for mc in report.months if mc.raw_file and mc.proc_bill]
    if comparable:
        mc0 = comparable[0]
        raw_wb = parse_workbook(mc0.raw_file)
        proc_wb = parse_workbook(mc0.proc_bill)
        lines.append("## 三、表格结构差异（原始 vs 处理后）")
        lines.append("")
        lines.append("原始表格与处理后表格在列结构上可能存在系统性差异，属正常导出转换，非数据错误：")
        lines.append("")
        lines.append("| 差异项 | 原始表格 | 处理后表格 |")
        lines.append("|--------|---------|-----------|")
        raw_has_mat = any(r.material for rows in raw_wb.sheets.values() for r in rows)
        raw_has_inst = any(r.instrument_count for rows in raw_wb.sheets.values() for r in rows)
        lines.append(f"| 表头行位置 | 约第 {raw_wb.header_row or '?'} 行 | 约第 {proc_wb.header_row or '?'} 行 |")
        lines.append(f"| 总列数（约） | {raw_wb.raw_col_count} 列 | {proc_wb.raw_col_count} 列 |")
        lines.append(f"| 包装材料列 | {'有' if raw_has_mat else '无/未识别'} | {'有' if any(r.material for rows in proc_wb.sheets.values() for r in rows) else '**已移除或未识别**'} |")
        lines.append(f"| 器械数列 | {'有' if raw_has_inst else '无/未识别'} | {'有' if any(r.instrument_count for rows in proc_wb.sheets.values() for r in rows) else '**已移除或未识别**'} |")
        lines.append("")
        raw_sheets = sorted({s for mc in comparable for s in []})
        all_raw_sheets = sorted({sn for mc in comparable for sn in (parse_workbook(mc.raw_file).sheets.keys() if mc.raw_file else [])})
        all_proc_sheets = sorted({sn for mc in comparable for sn in (parse_workbook(mc.proc_bill).sheets.keys() if mc.proc_bill else [])})
        lines.append("### Sheet 差异")
        lines.append("")
        lines.append(f"- 原始表格 Sheet：{', '.join(all_raw_sheets) or '—'}")
        lines.append(f"- 处理后表格 Sheet：{', '.join(all_proc_sheets) or '—'}")
        lines.append("")

    lines.append("## 四、原始表格数据问题明细")
    lines.append("")
    for mc in report.months:
        lines.append(f"### {mc.month}")
        lines.append("")
        if mc.warnings:
            for w in mc.warnings:
                lines.append(f"- ⚠️ {w}")
            lines.append("")
        if mc.proc_bill is None:
            lines.append("⏭️ **跳过：** 缺少对应的处理后账单文件，无法对比。")
            lines.append("")
            continue
        month_issues = [i for i in mc.issues if not i.issue_type.startswith("汇总不一致")]
        summary_issues = [i for i in mc.issues if i.issue_type.startswith("汇总不一致")]
        if not month_issues and not summary_issues:
            lines.append("✅ **无数据问题。** 所有明细行与处理后表格完全匹配。")
            lines.append("")
            continue

        by_sheet: dict[str, list[FieldDiff]] = defaultdict(list)
        for issue in month_issues:
            by_sheet[issue.sheet].append(issue)

        for sheet, issues in sorted(by_sheet.items()):
            field_issues = [i for i in issues if i.issue_type.startswith("字段不一致")]
            missing = [i for i in issues if i.issue_type == "缺失行"]
            extra = [i for i in issues if i.issue_type == "多余行"]
            if field_issues:
                lines.append(f"#### [{sheet}] 字段值不一致（{len(field_issues)} 条）")
                lines.append("")
                lines.append("| 行号(原始) | 包名 | 发货单号 | 字段 | 原始值 | 处理后值 | 说明 |")
                lines.append("|-----------|------|---------|------|--------|---------|------|")
                for i in field_issues:
                    lines.append(
                        f"| {i.raw_row or '—'} | {i.pack_name} | {i.ship_no} | {i.field} | "
                        f"{format_number(to_float(i.raw_value)) if i.raw_value != '' else ''} | "
                        f"{format_number(to_float(i.proc_value)) if i.proc_value != '' else ''} | {i.note or ''} |"
                    )
                lines.append("")
            if missing:
                lines.append(f"#### [{sheet}] 缺失行（{len(missing)} 条）")
                lines.append("")
                for i in missing:
                    lines.append(f"- 发货单号 {i.ship_no}，包名 `{i.pack_name}`（原始行 {i.raw_row}）")
                lines.append("")
            if extra:
                lines.append(f"#### [{sheet}] 多余行（{len(extra)} 条）")
                lines.append("")
                for i in extra:
                    lines.append(f"- 发货单号 {i.ship_no}，包名 `{i.pack_name}`（处理后行 {i.proc_row}）")
                lines.append("")

        if summary_issues:
            lines.append("#### 科室汇总不一致")
            lines.append("")
            for i in summary_issues:
                lines.append(
                    f"- **[{i.sheet}] {i.pack_name}** 总价：原始 {i.raw_value} → 处理后 {i.proc_value}"
                )
            lines.append("")

    lines.append("## 五、问题模式归纳")
    lines.append("")
    confirmed = [i for i in report.all_issues if i.issue_type.startswith("字段不一致")]
    structural_months = [mc.month for mc in report.months if mc.proc_bill and mc.raw_file]
    lines.append("### 5.1 已确认问题（需系统处理）")
    lines.append("")
    if confirmed:
        price_fixes = [i for i in confirmed if i.field in {"单价", "总价"}]
        lines.append(f"1. **字段不一致** 共 {len(confirmed)} 条（其中单价/总价 {len(price_fixes)} 条）")
        shown = 0
        for i in price_fixes[:5]:
            lines.append(
                f"   - {i.month}/{i.sheet} `{i.pack_name}` 发货单号 {i.ship_no}："
                f"{i.field} {i.raw_value} → {i.proc_value}"
            )
            shown += 1
        if len(price_fixes) > 5:
            lines.append(f"   - … 另有 {len(price_fixes) - 5} 条，见 CSV 清单")
    else:
        lines.append("无已确认的明细字段差异。")
    lines.append("")
    lines.append("### 5.2 结构性差异（非错误，导入时需适配）")
    lines.append("")
    lines.append("1. **列裁剪**：处理后可能移除「包装材料」「器械数」列")
    lines.append("2. **表头偏移**：原始与处理后表头行位置可能不同")
    lines.append("3. **合并单元格**：原始表格可能存在合并单元格，影响按列解析")
    lines.append("")
    lines.append("### 5.3 原始数据质量")
    lines.append("")
    if structural_months:
        clean = [mc for mc in report.months if mc.proc_bill and not month_has_data_issues(mc)]
        if clean:
            lines.append(f"- {', '.join(mc.month for mc in clean)} 原始与处理后 **完全一致**")
        problem = [mc for mc in report.months if mc.proc_bill and month_has_data_issues(mc)]
        if problem:
            lines.append(f"- {', '.join(mc.month for mc in problem)} 存在需关注的差异")
    lines.append("")
    lines.append("## 六、系统导入验证建议")
    lines.append("")
    lines.append("1. 将 `原始表格/` 中各月账单导入系统")
    lines.append("2. 导出账单明细，与 `处理后表格/` 中对应月份账单逐行对比")
    lines.append("3. 重点验证存在字段差异的明细行及科室汇总金额")
    lines.append("4. 结款函导出验证：与处理后结款函文件核对合计金额")
    lines.append("")
    return "\n".join(lines)


def write_csv(report: HospitalReport, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["月份", "Sheet", "问题类型", "包名", "发货单号", "字段", "原始值", "处理后值", "原始行号", "处理后行号", "说明"]
        )
        for i in report.all_issues:
            writer.writerow(
                [
                    i.month,
                    i.sheet,
                    i.issue_type,
                    i.pack_name,
                    i.ship_no,
                    i.field,
                    i.raw_value,
                    i.proc_value,
                    i.raw_row or "",
                    i.proc_row or "",
                    i.note,
                ]
            )


def render_index(reports: list[HospitalReport], analysis_date: str, failed: list[str]) -> str:
    lines = [
        "# 测试用例数据分析索引",
        "",
        f"> 生成日期：{analysis_date}",
        f"> 分析脚本：`scripts/analyze_test_case_excel.py`",
        "",
        "## 概览",
        "",
        "| 医院 | 可对比月份 | 问题条数 | 缺失处理后账单 | 状态 |",
        "|------|-----------|---------|---------------|------|",
    ]
    total_issues = 0
    for r in reports:
        months = ", ".join(mc.month for mc in r.months if mc.proc_bill) or "—"
        missing_proc = ", ".join(mc.month for mc in r.months if mc.proc_bill is None) or "—"
        count = issue_count(r)
        total_issues += count
        if r.failed:
            status = f"部分失败：{r.error}"
        elif count == 0:
            status = "✅ 一致"
        else:
            status = f"⚠️ {count} 条差异"
        lines.append(f"| {r.name} | {months} | {count} | {missing_proc} | {status} |")
    lines.extend(
        [
            "",
            f"**合计：** {len(reports)} 家医院，{total_issues} 条数据差异。",
            "",
        ]
    )
    if failed:
        lines.append("## 分析失败")
        lines.append("")
        for item in failed:
            lines.append(f"- {item}")
        lines.append("")
    lines.append("## 各医院报告")
    lines.append("")
    for r in reports:
        lines.append(f"- [{r.name}]({r.name}/数据问题分析.md) — {issue_count(r)} 条问题")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Analyze raw vs processed test case Excel files")
    parser.add_argument("--hospital", help="Only analyze one hospital folder name")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="Analysis date in report")
    args = parser.parse_args()

    hospitals = discover_hospitals(TEST_CASE_DIR)
    if args.hospital:
        hospitals = [p for p in hospitals if p.name == args.hospital]
        if not hospitals:
            print(f"未找到医院: {args.hospital}", file=sys.stderr)
            return 1

    reports: list[HospitalReport] = []
    hard_failed: list[str] = []

    for hospital_dir in hospitals:
        print(f"Analyzing {hospital_dir.name}...")
        try:
            report = analyze_hospital(hospital_dir)
        except Exception as exc:
            hard_failed.append(f"{hospital_dir.name}: {exc}")
            report = HospitalReport(name=hospital_dir.name, failed=True, error=str(exc))
        reports.append(report)

        md_path = hospital_dir / "数据问题分析.md"
        csv_path = hospital_dir / "数据问题清单.csv"
        md_path.write_text(render_markdown(report, args.date), encoding="utf-8")
        write_csv(report, csv_path)
        print(f"  -> {md_path.name}, {csv_path.name}, issues={issue_count(report)}")

    index_path = TEST_CASE_DIR / "数据分析索引.md"
    index_path.write_text(render_index(reports, args.date, hard_failed), encoding="utf-8")
    print(f"\nIndex written: {index_path}")
    print(f"Hospitals: {len(reports)}, Total issues: {sum(issue_count(r) for r in reports)}")
    if hard_failed:
        print("Failed:", "; ".join(hard_failed))
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
