#!/usr/bin/env python3
"""附一 11 列 export-v2 验收：42 sheet 表头 + golden 价格行（L/M/N）。"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[1]
EXPECTED_HEADERS = {
    9: "包数",
    10: "包装材料",
    11: "单包内器械数量/把",
    12: "单价（把）",
    13: "单价",
    14: "总价",
}

# reprice 后以引擎修正价为准（容差 ±0.05）
GOLDEN_ROWS = [
    {
        "sheet": "宫腔镜",
        "pack_substr": "镜头-3件",
        "m": 52.74,
        "check_l": False,
        "tol": 0.05,
    },
    {
        "sheet": "外二",
        "pack_substr": "换药包(120布)",
        "m": 21.99,
        "check_l": False,
        "tol": 0.05,
    },
    {
        "sheet": "手术室(一区)",
        "pack_substr": "30°腹腔镜",
        "m": 30.4,
        "n": 30.4,
        "check_l": False,
        "tol": 0.05,
    },
    {
        "sheet": "眼科门诊手术室（一）",
        "pack_substr": "球内注药",
        "m": 13.2,
        "check_l": False,
        "tol": 0.05,
    },
    {
        "sheet": "清洁区（手术室）",
        "pack_substr": "W15050",
        "l": 28.0,
        "m": 28.0,
        "n": 28.0,
        "tol": 0.05,
    },
    {
        "sheet": "手术室(一区)",
        "pack_substr": "王树人特器-26",
        "m": 328.0,
        "n": 328.0,
        "check_l": False,
        "tol": 0.05,
    },
]


def _cell_num(ws, row: int, col: int) -> float | None:
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def verify_export(path: Path, *, min_sheets: int = 42) -> list[str]:
    errors: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) < min_sheets:
            errors.append(f"sheet 数 {len(wb.sheetnames)} < {min_sheets}")
        for name in wb.sheetnames:
            ws = wb[name]
            for col_idx, expected in EXPECTED_HEADERS.items():
                actual = ws.cell(9, col_idx).value
                if str(actual or "").strip() != expected:
                    errors.append(f"[{name}] row9 col{col_idx} 期望「{expected}」实际「{actual}」")
        for golden in GOLDEN_ROWS:
            sheet_name = golden["sheet"]
            if sheet_name not in wb.sheetnames:
                errors.append(f"缺 sheet「{sheet_name}」")
                continue
            ws = wb[sheet_name]
            found = False
            for r in range(10, min(ws.max_row or 10, 5000) + 1):
                pack = ws.cell(r, 8).value
                if pack is None:
                    continue
                if golden["pack_substr"] not in str(pack):
                    continue
                found = True
                l = _cell_num(ws, r, 12)
                m = _cell_num(ws, r, 13)
                n = _cell_num(ws, r, 14)
                tol = golden.get("tol", 0.02)
                checks: list[tuple[str, float | None, float]] = []
                if golden.get("check_l", True) and "l" in golden:
                    checks.append(("L", l, golden["l"]))
                if "m" in golden:
                    checks.append(("M", m, golden["m"]))
                if "n" in golden:
                    checks.append(("N", n, golden["n"]))
                for label, actual, expected in checks:
                    if actual is None or abs(actual - expected) > tol:
                        errors.append(
                            f"[{sheet_name}] {golden['pack_substr']} row{r} {label}={actual} 期望≈{expected}"
                        )
                break
            if not found:
                errors.append(f"[{sheet_name}] 未找到包名含「{golden['pack_substr']}」的数据行")
    finally:
        wb.close()
    return errors


def main() -> int:
    p = argparse.ArgumentParser(description="附一 11 列 export 结构验收")
    p.add_argument("xlsx", type=Path, help="export-v2 输出的 xlsx")
    p.add_argument("--min-sheets", type=int, default=42)
    args = p.parse_args()
    if not args.xlsx.is_file():
        print(f"文件不存在: {args.xlsx}", file=sys.stderr)
        return 1
    errors = verify_export(args.xlsx, min_sheets=args.min_sheets)
    if errors:
        print("FAIL", args.xlsx)
        for e in errors:
            print(" -", e)
        return 1
    print(f"OK {args.xlsx} · {args.min_sheets}+ sheets · 11col headers · {len(GOLDEN_ROWS)} golden rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
