#!/usr/bin/env python3
"""Ingest 铂康/8.14新增 batch into 测试用例/ folder structure."""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "铂康" / "8.14新增"
TEST_CASE_DIR = ROOT / "测试用例"
MANIFEST_PATH = TEST_CASE_DIR / "814新增入库清单.json"

FOUR_HOSPITAL_SRC = SOURCE_DIR / "特殊收费" / "4家医院(1).xlsx"
FOUR_HOSPITAL_SPLIT: list[tuple[str, str, str]] = [
    ("冰城医美", "哈尔滨冰城医疗美容医院", "7月__冰城医美账单.xlsx"),
    ("电机厂", "国药总医院第二院区", "7月__电机厂账单.xlsx"),
    ("方南南", "方南南医院", "7月__方南南账单.xlsx"),
    ("东北农大", "东北农业大学", "7月__东北农大账单.xlsx"),
]


@dataclass
class IngestEntry:
    source: str
    dest: str
    hospital: str
    kind: str  # raw | proc_standard | proc_special | split
    month: int | None
    strict_eligible: bool
    note: str = ""


RAW_FILES: list[tuple[str, str, str, bool, str]] = [
    ("新建文件夹 (3)/冰成7月.xlsx", "哈尔滨冰城医疗美容医院", "7月__冰城原始.xlsx", True, "7月 strict 成对"),
    ("新建文件夹 (3)/电机厂.xlsx", "国药总医院第二院区", "7月__电机厂原始.xlsx", True, "7月 strict 成对"),
    ("新建文件夹 (3)/上德7月.xlsx", "黑龙江菁华上德生殖妇产医院", "7月__上德原始.xlsx", True, "7月 strict 成对"),
    ("新建文件夹 (3)/祖研7月.xlsx", "祖研-黑龙江省中医医院（南岗院区）", "7月__祖研南岗原始.xlsx", False, "缺 7 月处理后"),
    ("新建文件夹 (3)/五.xlsx", "哈尔滨市第五医院", "7月__市五院原始.xlsx", False, "主院区原始，无二门诊 proc"),
    ("新建文件夹 (3)/人口.xlsx", "黑龙江省妇幼保健院（人口）", "7月__人口原始.xlsx", False, "v8 省妇幼人口，缺 7 月处理后"),
]

STANDARD_PROC: list[tuple[str, str, str]] = [
    ("标准计费/三精肾病-1.xlsx", "三精肾病医院", "7月__三精肾病账单.xlsx"),
    ("标准计费/上德-1.xlsx", "黑龙江菁华上德生殖妇产医院", "7月__上德账单.xlsx"),
    ("标准计费/东大肛肠-1.xlsx", "黑龙江东大肛肠", "7月__东大肛肠账单.xlsx"),
    ("标准计费/中医附四-1.xlsx", "黑龙江中医药大学附属第四医院", "7月__中医附四账单.xlsx"),
    ("标准计费/仁胜-1.xlsx", "哈尔滨仁胜医院", "7月__仁胜账单.xlsx"),
    ("标准计费/先锋社区-1.xlsx", "南岗区先锋路社区卫生服务中心", "7月__先锋社区账单.xlsx"),
    ("标准计费/公安-1.xlsx", "哈尔滨市公安医院", "7月__公安账单.xlsx"),
    ("标准计费/南岗人民-1.xlsx", "哈尔滨市南岗区人民医院（九院）", "7月__南岗人民账单.xlsx"),
    ("标准计费/南岗妇产-1.xlsx", "南岗区妇产医院", "7月__南岗妇产账单.xlsx"),
    ("标准计费/和平社区-1.xlsx", "和平社区", "7月__和平社区账单.xlsx"),
    ("标准计费/国药总医院第三医院-1.xlsx", "国药总医院第三院区", "7月__国药三院账单.xlsx"),
    ("标准计费/奥美-1.xlsx", "奥美", "7月__奥美账单.xlsx"),
    ("标准计费/媛尚美-1.xlsx", "媛尚美", "7月__媛尚美账单.xlsx"),
    ("标准计费/悦美芳华-1.xlsx", "悦美芳华医疗门诊医院", "7月__悦美芳华账单.xlsx"),
    ("标准计费/维多利亚-1.xlsx", "黑龙江维多利亚妇产医院", "7月__维多利亚账单.xlsx"),
    ("标准计费/香坊中医院-1.xlsx", "香坊中医院", "7月__香坊中医院账单.xlsx"),
]

SPECIAL_PROC: list[tuple[str, str, str, str]] = [
    ("特殊收费/电机厂-2.xlsx", "国药总医院第二院区", "7月__电机厂特殊账单.xlsx", "6-7月跨期，备选 proc"),
    ("特殊收费/市五院二门诊-2.xlsx", "哈尔滨市第五医院（二门诊）", "6月__市五院二门诊特殊账单.xlsx", "6 月特殊 proc"),
    ("特殊收费/九州-2.xlsx", "黑龙江九洲妇科医院", "6月__九州特殊账单.xlsx", "3-4 月 proc"),
    ("特殊收费/工程大学-2.xlsx", "哈尔滨工程大学医院", "6月__工程大学特殊账单.xlsx", "6 月 proc"),
    ("特殊收费/松电慢病-2.xlsx", "松电慢病", "6月__松电慢病特殊账单.xlsx", "3 月 proc"),
    ("特殊收费/海员松北-2.xlsx", "黑龙江省海员总医院（松北）", "6月__海员松北特殊账单.xlsx", "3-4 月 proc"),
    ("特殊收费/航天风华-2.xlsx", "航天风华", "8月__航天风华特殊账单.xlsx", "8 月 proc"),
    ("特殊收费/东北农大-2.xlsx", "东北农业大学", "6月__东北农大特殊账单.xlsx", "6-7 月 proc"),
]

STRICT_JULY_FOLDERS = {
    "黑龙江菁华上德生殖妇产医院",
    "哈尔滨冰城医疗美容医院",
    "国药总医院第二院区",
}


def ensure_dirs(hospital: str, *, raw: bool = False, proc: bool = False) -> Path:
    if hospital == "待匹配":
        base = TEST_CASE_DIR / "待匹配"
        if raw:
            (base / "原始表格").mkdir(parents=True, exist_ok=True)
        if proc:
            (base / "处理后表格").mkdir(parents=True, exist_ok=True)
        return base
    base = TEST_CASE_DIR / hospital
    if raw:
        (base / "原始表格").mkdir(parents=True, exist_ok=True)
    if proc:
        (base / "处理后表格").mkdir(parents=True, exist_ok=True)
    return base


def split_four_hospital_workbook() -> list[IngestEntry]:
    if not FOUR_HOSPITAL_SRC.is_file():
        return []
    import openpyxl

    entries: list[IngestEntry] = []
    wb_src = openpyxl.load_workbook(FOUR_HOSPITAL_SRC)
    for sheet_name, hospital, dest_name in FOUR_HOSPITAL_SPLIT:
        if sheet_name not in wb_src.sheetnames:
            continue
        ensure_dirs(hospital, proc=True)
        dest = TEST_CASE_DIR / hospital / "处理后表格" / dest_name
        wb_dst = openpyxl.Workbook()
        ws_dst = wb_dst.active
        ws_dst.title = sheet_name
        ws_src = wb_src[sheet_name]
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row))
        wb_dst.save(dest)
        strict = hospital in STRICT_JULY_FOLDERS and sheet_name in {"冰城医美", "电机厂"}
        entries.append(
            IngestEntry(
                source=str(FOUR_HOSPITAL_SRC.relative_to(ROOT)) + f"#{sheet_name}",
                dest=str(dest.relative_to(ROOT)),
                hospital=hospital,
                kind="split",
                month=7,
                strict_eligible=strict,
                note="4家 workbook 拆分",
            )
        )
    wb_src.close()
    return entries


def copy_file(rel_src: str, hospital: str, subdir: str, dest_name: str) -> Path:
    src = SOURCE_DIR / rel_src
    if hospital == "待匹配":
        dest = TEST_CASE_DIR / "待匹配" / subdir / dest_name
    else:
        dest = TEST_CASE_DIR / hospital / subdir / dest_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    return dest


def ingest(*, dry_run: bool = False) -> list[IngestEntry]:
    entries: list[IngestEntry] = []

    if not dry_run:
        entries.extend(split_four_hospital_workbook())

    for rel, hospital, dest_name, strict, note in RAW_FILES:
        src = SOURCE_DIR / rel
        if not src.is_file():
            continue
        if hospital == "待匹配":
            ensure_dirs("待匹配", raw=True)
            dest_rel = f"测试用例/待匹配/原始表格/{dest_name}"
            if not dry_run:
                shutil.copy2(src, TEST_CASE_DIR / "待匹配" / "原始表格" / dest_name)
        else:
            ensure_dirs(hospital, raw=True)
            dest_rel = f"测试用例/{hospital}/原始表格/{dest_name}"
            if not dry_run:
                copy_file(rel, hospital, "原始表格", dest_name)
        entries.append(
            IngestEntry(
                source=str((SOURCE_DIR / rel).relative_to(ROOT)),
                dest=dest_rel,
                hospital=hospital,
                kind="raw",
                month=7,
                strict_eligible=strict,
                note=note,
            )
        )

    for rel, hospital, dest_name in STANDARD_PROC:
        src = SOURCE_DIR / rel
        if not src.is_file():
            continue
        ensure_dirs(hospital, proc=True)
        dest_rel = f"测试用例/{hospital}/处理后表格/{dest_name}"
        if not dry_run:
            copy_file(rel, hospital, "处理后表格", dest_name)
        strict = hospital in STRICT_JULY_FOLDERS and dest_name == "7月__上德账单.xlsx"
        if hospital == "黑龙江菁华上德生殖妇产医院":
            strict = True
        entries.append(
            IngestEntry(
                source=str(src.relative_to(ROOT)),
                dest=dest_rel,
                hospital=hospital,
                kind="proc_standard",
                month=7,
                strict_eligible=strict,
                note="标准计费 -1",
            )
        )

    for rel, hospital, dest_name, note in SPECIAL_PROC:
        src = SOURCE_DIR / rel
        if not src.is_file():
            continue
        ensure_dirs(hospital, proc=True)
        dest_rel = f"测试用例/{hospital}/处理后表格/{dest_name}"
        if not dry_run:
            copy_file(rel, hospital, "处理后表格", dest_name)
        month = 7 if dest_name.startswith("7月") else (8 if dest_name.startswith("8月") else 6)
        entries.append(
            IngestEntry(
                source=str(src.relative_to(ROOT)),
                dest=dest_rel,
                hospital=hospital,
                kind="proc_special",
                month=month,
                strict_eligible=False,
                note=note,
            )
        )

    if dry_run and FOUR_HOSPITAL_SRC.is_file():
        for sheet_name, hospital, dest_name in FOUR_HOSPITAL_SPLIT:
            strict = hospital in STRICT_JULY_FOLDERS and sheet_name in {"冰城医美", "电机厂"}
            entries.append(
                IngestEntry(
                    source=str(FOUR_HOSPITAL_SRC.relative_to(ROOT)) + f"#{sheet_name}",
                    dest=f"测试用例/{hospital}/处理后表格/{dest_name}",
                    hospital=hospital,
                    kind="split",
                    month=7,
                    strict_eligible=strict,
                    note="4家 workbook 拆分",
                )
            )

    return entries


def write_manifest(entries: list[IngestEntry]) -> None:
    payload = {
        "source_dir": str(SOURCE_DIR.relative_to(ROOT)),
        "strict_july_folders": sorted(STRICT_JULY_FOLDERS),
        "entries": [asdict(e) for e in entries],
    }
    MANIFEST_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_manifest() -> dict:
    if not MANIFEST_PATH.is_file():
        return {"strict_july_folders": sorted(STRICT_JULY_FOLDERS), "entries": []}
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


def strict_july_hospitals() -> list[str]:
    data = load_manifest()
    folders = data.get("strict_july_folders") or sorted(STRICT_JULY_FOLDERS)
    return list(folders)


def main() -> int:
    p = argparse.ArgumentParser(description="入库 铂康/8.14新增 至 测试用例/")
    p.add_argument("--write", action="store_true", help="执行复制/拆分并写入 manifest")
    p.add_argument("--dry-run", action="store_true", help="仅预览 manifest，不写文件")
    args = p.parse_args()
    if not SOURCE_DIR.is_dir():
        print(f"源目录不存在: {SOURCE_DIR}", file=__import__("sys").stderr)
        return 1
    entries = ingest(dry_run=args.dry_run and not args.write)
    if args.write:
        entries = ingest(dry_run=False)
        write_manifest(entries)
        print(f"已入库 {len(entries)} 项 → {MANIFEST_PATH}")
    else:
        print(json.dumps([asdict(e) for e in entries], ensure_ascii=False, indent=2))
    strict = [e for e in entries if e.strict_eligible]
    print(f"\n7 月 strict 可测: {len({e.hospital for e in strict})} 院")
    for h in sorted({e.hospital for e in strict}):
        print(f"  - {h}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
