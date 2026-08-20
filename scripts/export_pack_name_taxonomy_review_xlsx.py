#!/usr/bin/env python3
"""导出包名计数 taxonomy 客户复核 Excel（T01–T99 真实账单示例）。"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required", file=sys.stderr)
    sys.exit(1)

# Reuse parser + classifier from scan script
sys.path.insert(0, os.path.dirname(__file__))
from scan_pack_name_field_consistency import (  # noqa: E402
    IC_COLS,
    PC_COLS,
    PN_COLS,
    classify,
    extract_total,
    find_header_row,
    should_skip,
    stem_before_slash,
)

TAXONOMY_META: dict[str, dict[str, str]] = {
    "T01_连字符N件": {
        "id": "T01",
        "name": "连字符 -N件/-N + 斜杠码",
        "desc": "如 粉刺针-3/Z7526、止血钳-2剪-1/Z1530；系统累加全部 -N/-N件。",
        "question": "连字符后的数字是否都应计入器械件数？多段连字符是否求和？",
    },
    "T02_连字符N": {
        "id": "T02",
        "name": "连字符 -N + 斜杠码（无「件」字）",
        "desc": "如 镊子-1/z7534；-N 是否等于单包器械数？",
        "question": "无「件」字的 -N 是否与 -N件 含义相同？",
    },
    "T03_连字符括号": {
        "id": "T03",
        "name": "连字符 + 括号内容器/说明",
        "desc": "如 -9（筐1）、-4（带盒5件）；括号内盒/筐/盘是否另计 1 件？",
        "question": "（筐1）（带盒5件）等括号内容如何计入器械数？",
    },
    "T04_紧凑复合": {
        "id": "T04",
        "name": "紧凑复合（无连字符多段）",
        "desc": "如 盆1碗1、镊子1止血钳2；各「汉字+数字」段是否求和？",
        "question": "盆1碗1 应算 2 件还是只算末尾数字？",
    },
    "T05_针架": {
        "id": "T05",
        "name": "针架复合",
        "desc": "如 车针架1针4；系统当前跳过核对。",
        "question": "针架+针数应如何计件（按架 1 件还是架+针求和）？",
    },
    "T06_件盒_连字符": {
        "id": "T06",
        "name": "N件盒M / -N盒M（连字符）",
        "desc": "如 机扩针-6盒1、种植盒-8件盒1、宫腔镜-2件盒1。",
        "question": "「盒1」是否另计 1 件？种植/机扩与电切/宫腔镜是否规则不同？",
    },
    "T07_件盒_无连字符": {
        "id": "T07",
        "name": "N件盒M（无连字符或空格）",
        "desc": "如 种植9件盒1、种植盒-10件 盒1。",
        "question": "「N件 盒1」与「N件盒1」计数是否一致？",
    },
    "T08_ZSD": {
        "id": "T08",
        "name": "ZSD 器械包编码",
        "desc": "包名含 ZSD 编码；通常无法从包名推断件数。",
        "question": "此类是否不做包名件数核对？件数以何为准？",
    },
    "T09_末尾数字": {
        "id": "T09",
        "name": "stem 末尾「名+数字」",
        "desc": "如 排针20、止血钳3；末尾数字是否为单包件数？",
        "question": "末尾数字是否永远表示器械件数？手机/型号编码如何处理？",
    },
    "T10_仅码": {
        "id": "T10",
        "name": "仅斜杠码、无明显件数",
        "desc": "如 洗手服/w12050；系统通常跳过件数核对。",
        "question": "无数字包名是否不做核对？若账单有器械数，含义是什么？",
    },
    "T11_尺寸规格": {
        "id": "T11",
        "name": "尺寸规格（75*200 等）",
        "desc": "如 支抗钉-3 75*20；件数通常走连字符 -N。",
        "question": "尺寸数字是否不应参与件数统计？",
    },
    "T12_双袋": {
        "id": "T12",
        "name": "双袋",
        "desc": "如 持镜钳-2件/Z双2040。",
        "question": "「双袋」是否影响器械件数计算？",
    },
    "T13_N袋": {
        "id": "T13",
        "name": "-N袋",
        "desc": "如 器械包-2袋；系统当前跳过核对。",
        "question": "「袋」是否等于器械件数？",
    },
    "T99_其他": {
        "id": "T99",
        "name": "其他/混合",
        "desc": "如 宫腔镜包26件、开腹包-50件 等未归入上述规则者。",
        "question": "请说明应如何解析此类包名的器械件数。",
    },
}

TYPE_ORDER = [
    "T01_连字符N件",
    "T02_连字符N",
    "T03_连字符括号",
    "T04_紧凑复合",
    "T05_针架",
    "T06_件盒_连字符",
    "T07_件盒_无连字符",
    "T08_ZSD",
    "T09_末尾数字",
    "T10_仅码",
    "T11_尺寸规格",
    "T12_双袋",
    "T13_N袋",
    "T99_其他",
]

# 测试库中极少出现或分类被 T01 覆盖的类型，补充典型真实/文档示例
MANUAL_SUPPLEMENT: dict[str, list[dict]] = {
    "T02_连字符N": [
        {
            "pack_name": "镊子-1/z7526",
            "pack_count": 5,
            "instrument_count": 5,
            "parsed_per_pack": 1,
            "expected_total": 5,
            "status": "一致",
            "bucket": "match",
            "hospital": "（V8对账报告抽样）",
            "file": "特殊收费v8严格Excel对账报告",
        },
        {
            "pack_name": "整形镊子-1/Z7520",
            "pack_count": 2,
            "instrument_count": 2,
            "parsed_per_pack": 1,
            "expected_total": 2,
            "status": "一致",
            "bucket": "match",
            "hospital": "（V8对账报告抽样）",
            "file": "特殊收费v8严格Excel对账报告",
        },
    ],
    "T08_ZSD": [
        {
            "pack_name": "手术包（二）",
            "pack_count": 1,
            "instrument_count": 43,
            "parsed_per_pack": None,
            "expected_total": None,
            "status": "系统未解析/跳过（类型=器械包(ZSD)，包名无 ZSD 编码）",
            "bucket": "skip",
            "hospital": "哈尔滨长健医院",
            "file": "6月__长健6月账单.xlsx（文档抽样）",
        },
        {
            "pack_name": "换药包(120布)",
            "pack_count": 3,
            "instrument_count": 0,
            "parsed_per_pack": None,
            "expected_total": None,
            "status": "系统未解析/跳过（类型含 ZSD/器械包语义）",
            "bucket": "skip",
            "hospital": "黑龙江中医药大学附属第一医院",
            "file": "job11 系统 warning 抽样",
        },
    ],
    "T04_紧凑复合": [
        {
            "pack_name": "盆1碗1/W9050",
            "pack_count": 4,
            "instrument_count": 8,
            "parsed_per_pack": 2,
            "expected_total": 8,
            "status": "一致",
            "bucket": "match",
            "hospital": "哈尔滨市红十字妇产医院",
            "file": "（呼兰红十字案例）",
        },
    ],
}

# 材料中极少出现或分类器无法单独抽样的示例（手工补充，来源见备注）
MANUAL_EXAMPLES: list[dict] = [
    {
        "type_key": "T02_连字符N",
        "bucket": "match",
        "pack_name": "镊子-1/z7534",
        "pack_count": 1,
        "instrument_count": 1,
        "parsed_per_pack": 1,
        "expected_total": 1,
        "status": "一致",
        "hospital": "（典型示例）",
        "file": "分类说明：T02 与 T01 连字符规则相同，件字可省略",
    },
    {
        "type_key": "T08_ZSD",
        "bucket": "skip",
        "pack_name": "手术包（二）",
        "pack_count": 1,
        "instrument_count": 43,
        "parsed_per_pack": None,
        "expected_total": None,
        "status": "系统未解析/跳过",
        "hospital": "哈尔滨长健医院",
        "file": "类型列=器械包(ZSD)；件数以器械数列为准，包名无 ZSD 编码",
    },
    {
        "type_key": "T08_ZSD",
        "bucket": "skip",
        "pack_name": "换药包(120布)",
        "pack_count": 3,
        "instrument_count": 0,
        "parsed_per_pack": None,
        "expected_total": None,
        "status": "系统未解析/跳过",
        "hospital": "黑龙江中医药大学附属第一医院",
        "file": "类型列=器械包(ZSD)；敷料/ZSD 包是否应对照包名计件？",
    },
    {
        "type_key": "T06_件盒_连字符",
        "bucket": "mismatch",
        "pack_name": "盆1碗1/W9050",
        "pack_count": 4,
        "instrument_count": 8,
        "parsed_per_pack": 2,
        "expected_total": 8,
        "status": "一致",
        "hospital": "哈尔滨市红十字妇产医院",
        "file": "紧凑复合归入 T04，此处作 T06 对照：盆+碗=2 件 ×4 包",
    },
]

HEADERS = [
    "类型编号",
    "类型名称",
    "类型说明",
    "待确认问题",
    "示例类别",
    "包名",
    "包数",
    "器械数（账单列）",
    "系统解析·单包件数",
    "系统期望器械数合计",
    "系统判定",
    "来源医院目录",
    "来源文件名",
    "客户确认：计数是否正确（填 是/否/不适用）",
    "客户确认：单包正确件数（若与系统不同请填写）",
    "客户确认：正确计数规则（请在此说明）",
]


def hospital_from_path(path: str, root: str) -> str:
    rel = os.path.relpath(path, root)
    parts = rel.split(os.sep)
    return parts[0] if parts else ""


def row_key(pn: str, pc: int, ic: int) -> tuple:
    return (pn, pc, ic)


def collect_examples(root: str, per_bucket: int = 3) -> dict[str, dict[str, list[dict]]]:
    """每类型收集 match / mismatch / skip 示例。"""
    buckets: dict[str, dict[str, list[dict]]] = defaultdict(
        lambda: {"match": [], "mismatch": [], "skip": []}
    )
    seen: dict[str, set[tuple]] = defaultdict(set)

    for dirpath, _, filenames in os.walk(root):
        if "处理后表格" not in dirpath and "原始表格" not in dirpath:
            continue
        for fn in filenames:
            if not fn.endswith(".xlsx") or fn.startswith("~"):
                continue
            path = os.path.join(dirpath, fn)
            hospital = hospital_from_path(path, root)
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            for ws in wb.worksheets:
                header_info = find_header_row(ws)
                if not header_info:
                    continue
                header, hdr_row = header_info
                idx = {name: header.index(name) for name in header if name}
                pn_i = next((idx[c] for c in PN_COLS if c in idx), None)
                pc_i = next((idx[c] for c in PC_COLS if c in idx), None)
                ic_i = next((idx[c] for c in IC_COLS if c in idx), None)
                if pn_i is None or ic_i is None:
                    continue
                for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                    if not row or all(v is None or str(v).strip() == "" for v in row):
                        continue
                    pn = row[pn_i] if pn_i < len(row) else None
                    if pn is None or str(pn).strip() == "":
                        continue
                    pn = str(pn).strip()
                    try:
                        ic = int(float(row[ic_i])) if ic_i < len(row) and row[ic_i] is not None else None
                    except (TypeError, ValueError):
                        continue
                    try:
                        pc = (
                            max(1, int(float(row[pc_i])))
                            if pc_i is not None and pc_i < len(row) and row[pc_i] is not None
                            else 1
                        )
                    except (TypeError, ValueError):
                        pc = 1
                    cat = classify(pn)
                    if cat == "EMPTY":
                        continue
                    key = row_key(pn, pc, ic)
                    if key in seen[cat]:
                        continue
                    stem = stem_before_slash(pn)
                    parsed = extract_total(pn)
                    if should_skip(stem) or parsed is None:
                        bucket = "skip"
                        expected = None
                        status = "系统未解析/跳过"
                    else:
                        expected = parsed * pc
                        if expected == ic:
                            bucket = "match"
                            status = "一致"
                        else:
                            bucket = "mismatch"
                            status = "不一致"
                    if len(buckets[cat][bucket]) >= per_bucket:
                        continue
                    seen[cat].add(key)
                    buckets[cat][bucket].append(
                        {
                            "pack_name": pn,
                            "pack_count": pc,
                            "instrument_count": ic,
                            "parsed_per_pack": parsed,
                            "expected_total": expected,
                            "status": status,
                            "bucket": bucket,
                            "hospital": hospital,
                            "file": fn,
                        }
                    )
            wb.close()
    return buckets


def build_rows(buckets: dict[str, dict[str, list[dict]]]) -> list[list]:
    rows: list[list] = []
    bucket_label = {"match": "匹配示例", "mismatch": "争议示例（系统≠账单）", "skip": "跳过/无解析示例"}
    for cat in TYPE_ORDER:
        meta = TAXONOMY_META.get(cat, {"id": cat, "name": cat, "desc": "", "question": ""})
        cat_buckets = buckets.get(cat, {})
        seen_keys: set[tuple] = set()

        def append_example(ex: dict, bucket_name: str) -> None:
            key = row_key(ex["pack_name"], ex["pack_count"], ex["instrument_count"])
            if key in seen_keys:
                return
            seen_keys.add(key)
            rows.append(
                [
                    meta["id"],
                    meta["name"],
                    meta["desc"],
                    meta["question"],
                    bucket_label.get(bucket_name, bucket_name),
                    ex["pack_name"],
                    ex["pack_count"],
                    ex["instrument_count"],
                    ex["parsed_per_pack"] if ex.get("parsed_per_pack") is not None else "",
                    ex["expected_total"] if ex.get("expected_total") is not None else "",
                    ex["status"],
                    ex["hospital"],
                    ex["file"],
                    "",
                    "",
                    "",
                ]
            )

        for bucket_name in ("mismatch", "match", "skip"):
            for ex in cat_buckets.get(bucket_name, []):
                append_example(ex, bucket_name)
        for ex in MANUAL_SUPPLEMENT.get(cat, []):
            append_example(ex, ex.get("bucket", "match"))
        if not seen_keys:
            rows.append(
                [
                    meta["id"],
                    meta["name"],
                    meta["desc"],
                    meta["question"],
                    "（暂无抽样，请补充）",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
    return rows


def write_workbook(out_path: str, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    # 说明 sheet
    ws0 = wb.active
    ws0.title = "填写说明"
    instructions = [
        ["包名计数 taxonomy 客户复核表", ""],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["用途", "请铂康/医院客户逐条确认：包名所示件数 × 包数 是否应等于「器械数」列。"],
        ["", ""],
        ["列说明", ""],
        ["类型编号 T01–T99", "包名计数模式分类，详见「复核明细」每行。"],
        ["示例类别", "争议示例=系统解析与账单器械数不一致；匹配示例=当前一致；跳过=系统未做件数核对。"],
        ["系统解析·单包件数", "当前系统自动从包名提取的单包器械件数。"],
        ["系统期望器械数合计", "系统解析单包件数 × 包数。"],
        ["客户确认列（黄色）", "若计数不对，请在「正确计数规则」中写清规则，如：盆+碗求和、盒1另计1件、按架计1件等。"],
        ["", ""],
        ["类型一览", ""],
    ]
    for cat in TYPE_ORDER:
        m = TAXONOMY_META[cat]
        instructions.append([m["id"], f"{m['name']} — {m['question']}"])
    for r_idx, row in enumerate(instructions, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws0.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 80

    ws = wb.create_sheet("复核明细")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    customer_fill = PatternFill("solid", fgColor="FFF2CC")
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap
            if c_idx >= 14:
                cell.fill = customer_fill

    widths = [8, 22, 36, 32, 14, 42, 6, 12, 14, 16, 10, 24, 28, 18, 16, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export pack name taxonomy review Excel")
    parser.add_argument(
        "--root",
        default=os.path.join(os.path.dirname(os.path.dirname(__file__)), "测试用例"),
    )
    parser.add_argument(
        "--out",
        default=os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "测试用例",
            "包名计数taxonomy客户复核明细-20260819.xlsx",
        ),
    )
    parser.add_argument("--per-bucket", type=int, default=3, help="每类型每类示例最多条数")
    args = parser.parse_args()
    buckets = collect_examples(args.root, per_bucket=args.per_bucket)
    rows = build_rows(buckets)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    write_workbook(args.out, rows)
    total = len(rows)
    by_type = sum(1 for cat in TYPE_ORDER if any(buckets.get(cat, {}).get(b) for b in ("match", "mismatch", "skip")))
    print(f"Wrote {args.out}")
    print(f"rows={total} types_with_samples={by_type}/{len(TYPE_ORDER)}")


if __name__ == "__main__":
    main()
