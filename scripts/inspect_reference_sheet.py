#!/usr/bin/env python3
"""开发期参考表结构探测：表头、sheet 名、合计行。"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def inspect_file(path: Path, sheet_index: int = 0) -> dict:
    wb = load_workbook(path, data_only=True)
    idx = min(sheet_index, len(wb.sheetnames) - 1)
    sheet_name = wb.sheetnames[idx]
    ws = wb[sheet_name]
    headers: list[str] = []
    for r in range(1, min(8, (ws.max_row or 0) + 1)):
        row_vals = []
        for c in range(1, min(12, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                row_vals.append(str(v).strip())
        if len(row_vals) >= 2:
            headers = row_vals
            break
    numeric_total = 0.0
    for r in range(1, (ws.max_row or 0) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                numeric_total += float(v)
    return {
        "path": str(path),
        "sheetNames": wb.sheetnames,
        "activeSheet": sheet_name,
        "headers": headers,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "numericCellSum": round(numeric_total, 2),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect reference export sheet structure")
    parser.add_argument("path", type=Path, help="xlsx/xls reference file")
    parser.add_argument("--sheet", type=int, default=0)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if not args.path.exists():
        print(f"File not found: {args.path}", file=sys.stderr)
        return 1
    result = inspect_file(args.path, args.sheet)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"path: {result['path']}")
        print(f"sheets: {result['sheetNames']}")
        print(f"headers: {result['headers']}")
        print(f"rows: {result['maxRow']} cols: {result['maxColumn']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
