#!/usr/bin/env python3
"""附一 7 月 export parity：import 原始账单 → export-v2 → 与人工核对版零容差比对。"""

from __future__ import annotations

import argparse
import json
import sys
import tempfile
import time
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from batch_june_system_test import import_bill  # noqa: E402
from batch_s8_export_compare import export_bill  # noqa: E402
from lib.api_client import configure_client, get_client  # noqa: E402

HOSPITAL = "黑龙江中医药大学附属第一医院"
DEFAULT_ORIGINAL = Path(
    "/Users/yangxinghui/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_7qwn4vnuj7xo22_508c/temp/drag/附一7月6-30号账单-睿思导出.xlsx"
)
DEFAULT_MANUAL = Path(
    "/Users/yangxinghui/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_7qwn4vnuj7xo22_508c/temp/drag/账单-附属第一医院-人工核对版(1).xlsx"
)


def parse_bill(fp: Path) -> dict[tuple, dict]:
    wb = load_workbook(fp, read_only=True, data_only=True)
    out: dict[tuple, dict] = {}
    for sn in wb.sheetnames:
        if sn == "外来器械":
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        hdr_i = None
        for i, r in enumerate(rows[:20]):
            vals = [str(c).strip() if c is not None else "" for c in r]
            if any(v == "包名" for v in vals):
                hdr_i = i
                break
        if hdr_i is None:
            continue
        hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]

        def col(name: str) -> int | None:
            for j, v in enumerate(hdr):
                if v == name:
                    return j
            return None

        cols = {k: col(k) for k in ("发货日期", "发货单号", "包名", "单价", "总价")}
        for r in rows[hdr_i + 1 :]:
            if not r:
                continue
            pack = r[cols["包名"]] if cols["包名"] is not None else None
            if not pack:
                continue
            ps = str(pack).strip()
            if ps in ("", "合计", "总计", "包数") or "调整" in ps or ps == sn:
                continue
            try:
                unit = float(r[cols["单价"]])
                total = float(r[cols["总价"]])
            except (TypeError, ValueError):
                continue
            key = (
                sn,
                str(r[cols["发货日期"]]),
                str(r[cols["发货单号"]]),
                ps,
            )
            out[key] = {"unit": round(unit, 2), "total": round(total, 2)}
    wb.close()
    return out


def compare(manual: dict, exported: dict) -> tuple[int, list]:
    mismatches: list[dict] = []
    for key, cor in manual.items():
        exp = exported.get(key)
        if not exp:
            continue
        if cor["unit"] != exp["unit"] or cor["total"] != exp["total"]:
            mismatches.append(
                {
                    "key": key,
                    "manual_unit": cor["unit"],
                    "export_unit": exp["unit"],
                    "manual_total": cor["total"],
                    "export_total": exp["total"],
                }
            )
    return len(mismatches), mismatches


def main() -> int:
    p = argparse.ArgumentParser(description="附一 7 月 export 零容差验收")
    p.add_argument("--api", default="http://39.102.213.51:8853")
    p.add_argument("--original", type=Path, default=DEFAULT_ORIGINAL)
    p.add_argument("--manual", type=Path, default=DEFAULT_MANUAL)
    p.add_argument("--out-dir", type=Path, default=ROOT / "测试用例" / ".cli_smoke")
    p.add_argument("--json", action="store_true")
    args = p.parse_args()

    if not args.original.is_file():
        print(f"原始账单不存在: {args.original}", file=sys.stderr)
        return 2
    if not args.manual.is_file():
        print(f"人工核对版不存在: {args.manual}", file=sys.stderr)
        return 2

    configure_client(api_base=args.api.rstrip("/"), mode="direct")
    client = get_client()
    token = client.login(force=True)

    print(f"==> import {args.original.name}")
    t0 = time.time()
    job = import_bill(token, HOSPITAL, args.original)
    job_id = int(
        job.get("jobId")
        or job.get("job_id")
        or job.get("id")
        or (job.get("job") or {}).get("id")
        or 0
    )
    if job_id <= 0:
        raise RuntimeError(f"import 未返回 jobId: {job}")
    print(f"    Job #{job_id} ({time.time() - t0:.1f}s)")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    export_path = args.out_dir / f"zyy-d1-july-export-{job_id}.xlsx"
    print(f"==> export-v2 → {export_path.name}")
    export_bill(token, job_id, export_path)

    manual = parse_bill(args.manual)
    exported = parse_bill(export_path)
    count, samples = compare(manual, exported)
    ok = count == 0

    report = {
        "command": "fuyi-july-export-verify",
        "ok": ok,
        "job_id": job_id,
        "export_path": str(export_path),
        "manual_rows": len(manual),
        "export_rows": len(exported),
        "mismatches": count,
        "sample_mismatches": samples[:15],
    }

    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"\n附一 7 月 export parity: {'PASS' if ok else 'FAIL'}")
        print(f"  Job: {job_id}")
        print(f"  人工核对行: {len(manual)}  导出行: {len(exported)}")
        print(f"  单价/总价不一致: {count} 行（零容差）")
        if samples:
            print("  样例:")
            for m in samples[:8]:
                k = m["key"]
                print(
                    f"    {k[0]}/{k[3][:28]}  "
                    f"人工 {m['manual_unit']}/{m['manual_total']}  "
                    f"导出 {m['export_unit']}/{m['export_total']}"
                )

    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
