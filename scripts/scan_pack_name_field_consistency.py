#!/usr/bin/env python3
"""
包名计数 taxonomy 扫描 + 字段核对覆盖度报告。

与 backend PackNameSpecParser.extractTotalPieceCountFromPackName 逻辑对齐，
用于全库 xlsx 回归：按 T01–T13 分类统计 match/mismatch 率。

Taxonomy（带计数包名类型）
-------------------------
T01  连字符 -N件/-N + 斜杠码         粉刺针-3/Z7526、止血钳-2剪-1/Z1530
T02  连字符 -N + 斜杠码（无「件」）  镊子-1/z7534（并入 T01 统计）
T03  连字符 + 括号内容器           -9（筐1）/w7050、-4（带盒5件）
T04  紧凑复合（无连字符多段）       盆1碗1、镊子1止血钳2
T05  针架复合                       车针架1针4（字段层 skip）
T06  N件盒M / -N盒M（连字符）       机扩针-6盒1、宫腔镜-2件盒1
T07  N件盒M（无连字符/空格）         种植9件盒1、种植盒-10件 盒1
T08  ZSD 器械包编码                 正确 skip
T09  stem 末尾名+数字               排针20、止血钳3（手机序列号 skip）
T10  仅斜杠码无件数                 洗手服/w12050（skip）
T11  尺寸规格                       支抗钉-3 75*20（件数走 T01）
T12  双袋                           持镜钳-2件/Z双2040
T13  -N袋                           器械包-2袋（按 N 件计）
T99  其他/混合                      宫腔镜包26件、开腹包-50件

用法:
  python3 scripts/scan_pack_name_field_consistency.py
  python3 scripts/scan_pack_name_field_consistency.py --root 测试用例 --json-out 测试用例/包名字段核对扫描报告.json
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Any

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PIECE_COUNT = re.compile(r"[-－](\d+)件?")
STANDALONE_PIECE = re.compile(r"(\d+)件")
COMPACT = re.compile(r"([\u4e00-\u9fffA-Za-z]+)(\d+)")
TRAILING = re.compile(r"([\u4e00-\u9fffA-Za-z]+)(\d+)$")
NEEDLE_RACK = re.compile(r"针架\d+针\d+")
HYPHEN_BAG = re.compile(r"[-－]\d+袋")
COMPACT_MULTI = re.compile(r"[\u4e00-\u9fffA-Za-z]\d+[\u4e00-\u9fffA-Za-z]\d+")
GLUED_HYPHEN_ORDER = re.compile(r"(?i)(-\d+)([ZzWw]\d+)$")
GLUED_ORDER_AFTER_PAREN = re.compile(r"(?i)(?<=[）)])([ZzWw]\d+)$")
GLUED_ORDER_AFTER_STEM = re.compile(r"(?i)(?<=[\u4e00-\u9fff\d])([ZzWw]\d+)$")
SPACED_PIECE_THEN_BOX = re.compile(r"\d+件\s+(?:盒|筐|盘)\d+")
PACKAGING_BOX_PAREN = re.compile(r"^(?:带盒|[盒盘]\d*)$")
CONTAINER_AFTER_PIECE = re.compile(r"件\s*(?:盒|筐|盘)(\d+)")
CONTAINER_AFTER_NON_HAN = re.compile(r"(?<![\u4e00-\u9fff])(?:盒|筐|盘)(\d+)")
PAREN_GROUP = re.compile(r"[（(]([^）)]*)[）)]")
PRODUCT_MODEL = re.compile(r"\d+[A-Za-z][A-Za-z0-9]*\d{3,}|[A-Za-z][A-Za-z0-9]*\d{4,}")

PN_COLS = ("包名", "包名称", "pack_name", "PackName")
PC_COLS = ("包数", "pack_count", "PackCount")
IC_COLS = ("器械数", "器械数量", "instrument_count", "InstrumentCount")


def normalize_pack_name(pn: str) -> str:
    s = str(pn).strip()
    s = GLUED_HYPHEN_ORDER.sub(r"\1/\2", s)
    s = GLUED_ORDER_AFTER_PAREN.sub("", s)
    s = GLUED_ORDER_AFTER_STEM.sub("", s)
    return s.strip()


def normalize_stem(stem: str) -> str:
    return stem.replace("，", " ").replace(",", " ").strip()


def stem_before_slash(pn: str) -> str:
    s = normalize_pack_name(pn)
    i = s.find("/")
    return normalize_stem(s[:i].strip() if i > 0 else s.strip())


def is_product_model(stem: str) -> bool:
    if "手机" in stem:
        return True
    if stem.upper().startswith("Z00"):
        return True
    return PRODUCT_MODEL.search(stem) is not None


def should_skip(stem: str) -> bool:
    return bool(NEEDLE_RACK.search(stem) or is_product_model(stem))


def is_packaging_box_paren_only(inner: str) -> bool:
    t = inner.strip()
    if t == "带盒" or (t.startswith("带盒") and not PAREN_BOX_PIECE.search(t)):
        return True
    return PACKAGING_BOX_PAREN.match(t) is not None


def sum_container_tokens(text: str) -> int:
    total = 0
    for m in CONTAINER_AFTER_PIECE.finditer(text):
        total += int(m.group(1))
    stripped = CONTAINER_AFTER_PIECE.sub("件", text)
    for m in CONTAINER_AFTER_NON_HAN.finditer(stripped):
        total += int(m.group(1))
    return total


PAREN_BOX_PIECE = re.compile(r"带盒([\d两二三四五六七八九十]+)件")


def parse_chinese_or_arabic(token: str) -> int:
    if token.isdigit():
        return int(token)
    return {"两": 2, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}.get(token, 0)


def extract_paren_box_total(stem: str) -> int | None:
    max_total = None
    for m in PAREN_GROUP.finditer(stem):
        bm = PAREN_BOX_PIECE.search(m.group(1))
        if bm:
            total = parse_chinese_or_arabic(bm.group(1))
            if max_total is None or total > max_total:
                max_total = total
    return max_total


def sum_parenthesis_container_counts(stem: str) -> int:
    total = 0
    hyphen_piece = bool(PIECE_COUNT.search(stem))
    for m in PAREN_GROUP.finditer(stem):
        inner = m.group(1)
        if PAREN_BOX_PIECE.search(inner):
            continue
        if hyphen_piece and is_packaging_box_paren_only(inner):
            continue
        total += sum_container_tokens(inner)
    return total


def is_planting_or_needle_box_pack(stem: str) -> bool:
    keywords = ("种植", "机扩", "抛光", "环切", "洁牙", "扩针", "ITI", "登腾")
    return any(k in stem for k in keywords)


def is_surgical_pack_with_piece_box_count(stem: str) -> bool:
    dash = stem.find("-")
    if dash <= 0:
        dash = stem.find("－")
    if dash <= 0:
        return False
    return stem[:dash].strip().endswith("包")


def hyphen_count_before_box(text: str) -> int:
    m = re.search(r"[-－](\d+)(?:盒|筐|盘)", text)
    return int(m.group(1)) if m else 0


def sum_explicit_containers(stem: str, base_count: int, compact_base: bool) -> int:
    if compact_base:
        return sum_parenthesis_container_counts(stem)
    if re.search(r"[-－]\d+件", stem):
        if SPACED_PIECE_THEN_BOX.search(stem):
            if is_planting_or_needle_box_pack(stem) or is_surgical_pack_with_piece_box_count(stem):
                without_paren = PAREN_GROUP.sub("", stem)
                return sum_parenthesis_container_counts(stem) + sum_container_tokens(without_paren)
            return sum_parenthesis_container_counts(stem)
        without_paren = PAREN_GROUP.sub("", stem)
        if re.search(r"[-－]\d+件(?:盒|筐|盘)", stem) and is_planting_or_needle_box_pack(stem):
            return sum_parenthesis_container_counts(stem) + sum_container_tokens(without_paren)
        if re.search(r"[-－]\d+件盒", stem) and is_surgical_pack_with_piece_box_count(stem):
            return sum_parenthesis_container_counts(stem) + sum_container_tokens(without_paren)
        return sum_parenthesis_container_counts(stem)
    if re.search(r"[-－]\d+件\s*(?:盒|筐|盘)", stem):
        return sum_parenthesis_container_counts(stem)
    total = sum_parenthesis_container_counts(stem)
    without = PAREN_GROUP.sub("", stem)
    if SPACED_PIECE_THEN_BOX.search(without) and not is_planting_or_needle_box_pack(stem):
        return total
    for m in CONTAINER_AFTER_PIECE.finditer(without):
        total += int(m.group(1))
    stripped = CONTAINER_AFTER_PIECE.sub("件", without)
    if re.search(r"[-－](\d+)(?:盒|筐|盘)", stripped):
        hc = hyphen_count_before_box(stripped)
        if is_planting_or_needle_box_pack(stem) or (2 <= hc <= 12):
            for cm in CONTAINER_AFTER_NON_HAN.finditer(stripped):
                total += int(cm.group(1))
        return total
    for cm in CONTAINER_AFTER_NON_HAN.finditer(stripped):
        total += int(cm.group(1))
    return total


def extract_base(stem: str) -> tuple[int | None, bool]:
    count_stem = PAREN_GROUP.sub("", stem)
    hs = 0
    found = False
    for m in PIECE_COUNT.finditer(count_stem):
        hs += int(m.group(1))
        found = True
    if found:
        return hs, False
    m = STANDALONE_PIECE.search(stem)
    if m:
        return int(m.group(1)), False
    if is_product_model(stem):
        return None, False
    cs = 0
    segs = 0
    for m in COMPACT.finditer(stem):
        cs += int(m.group(2))
        segs += 1
    if segs >= 2:
        return cs, True
    m = TRAILING.search(stem)
    if m:
        name_part = m.group(1)
        trailing_count = int(m.group(2))
        if name_part.endswith("少") or trailing_count > 99:
            return None, False
        return trailing_count, False
    return None, False


def extract_total(pn: str) -> int | None:
    if not pn:
        return None
    stem = stem_before_slash(str(pn).strip())
    if not stem or should_skip(stem):
        return None
    base, compact = extract_base(stem)
    if base is None:
        return None
    paren_total = extract_paren_box_total(stem)
    if paren_total is not None and paren_total > base:
        base = paren_total
    return base + sum_explicit_containers(stem, base, compact)


def classify(pn: str) -> str:
    if not pn or not str(pn).strip():
        return "EMPTY"
    s = str(pn).strip()
    stem = stem_before_slash(s)
    if re.search(r"ZSD\d+", s, re.I):
        return "T08_ZSD"
    if NEEDLE_RACK.search(stem):
        return "T05_针架"
    if HYPHEN_BAG.search(stem):
        return "T13_N袋"
    if re.search(r"[-－]\d+件?盒\d+", stem):
        return "T06_件盒_连字符"
    if re.search(r"\d+件\s*盒\d+", stem) or re.search(r"\d+件盒\d+", stem):
        return "T07_件盒_无连字符"
    if re.search(r"[-－]\d+件?/[ZzWw]\d+", s):
        return "T01_连字符N件"
    if re.search(r"[-－]\d+/[ZzWw]\d+", s):
        return "T02_连字符N"
    if re.search(r"[-－]\d+[（(]", stem):
        return "T03_连字符括号"
    if re.search(r"[\u4e00-\u9fffA-Za-z]+\d+[\u4e00-\u9fffA-Za-z]+\d+", stem) and not re.search(
        r"[-－]", stem
    ):
        return "T04_紧凑复合"
    if re.search(r"[\u4e00-\u9fffA-Za-z]+\d+$", stem) and not re.search(r"[-－]", stem):
        return "T09_末尾数字"
    if re.search(r"/[ZzWw]\d+", s):
        return "T10_仅码"
    if re.search(r"\d+(?:\.\d+)?\s*[×x*]\s*\d+", s):
        return "T11_尺寸规格"
    if "双" in s and re.search(r"/[ZzWw]", s):
        return "T12_双袋"
    return "T99_其他"


def find_header_row(ws) -> tuple[list[str], int] | None:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(c in PN_COLS for c in cells):
            return cells, i
    return None


def scan_xlsx(root: str) -> dict[str, Any]:
    total_c: Counter[str] = Counter()
    match_c: Counter[str] = Counter()
    mismatch_c: Counter[str] = Counter()
    skip_c: Counter[str] = Counter()
    examples: dict[str, list[dict[str, Any]]] = defaultdict(list)
    files_scanned = 0
    rows_scanned = 0

    for dirpath, _, filenames in os.walk(root):
        if "处理后表格" not in dirpath and "原始表格" not in dirpath:
            continue
        for fn in filenames:
            if not fn.endswith(".xlsx") or fn.startswith("~"):
                continue
            path = os.path.join(dirpath, fn)
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            files_scanned += 1
            for ws in wb.worksheets:
                header_info = find_header_row(ws)
                if not header_info:
                    continue
                header, hdr_row = header_info
                idx = {name: header.index(name) for name in header if name}
                pn_i = next((idx[c] for c in PN_COLS if c in idx), None)
                pc_i = next((idx[c] for c in PC_COLS if c in idx), None)
                ic_i = next((idx[c] for c in IC_COLS if c in idx), None)
                if pn_i is None or ic_i is None:
                    continue
                for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                    if not row or all(v is None or str(v).strip() == "" for v in row):
                        continue
                    pn = row[pn_i] if pn_i < len(row) else None
                    if pn is None or str(pn).strip() == "":
                        continue
                    pn = str(pn).strip()
                    try:
                        ic = int(float(row[ic_i])) if ic_i < len(row) and row[ic_i] is not None else None
                    except (TypeError, ValueError):
                        ic = None
                    if ic is None:
                        continue
                    try:
                        pc = (
                            max(1, int(float(row[pc_i])))
                            if pc_i is not None and pc_i < len(row) and row[pc_i] is not None
                            else 1
                        )
                    except (TypeError, ValueError):
                        pc = 1
                    rows_scanned += 1
                    cat = classify(pn)
                    total_c[cat] += 1
                    ext = extract_total(pn)
                    if ext is None:
                        skip_c[cat] += 1
                        continue
                    expected = ext * pc
                    if expected == ic:
                        match_c[cat] += 1
                    else:
                        mismatch_c[cat] += 1
                        if len(examples[cat]) < 6:
                            examples[cat].append(
                                {
                                    "pack_name": pn,
                                    "parsed": ext,
                                    "pack_count": pc,
                                    "expected": expected,
                                    "instrument_count": ic,
                                    "file": os.path.basename(path),
                                }
                            )
            wb.close()

    by_type: dict[str, Any] = {}
    focus = ("T03_连字符括号", "T06_件盒_连字符", "T07_件盒_无连字符")
    for cat in sorted(total_c.keys(), key=lambda c: -total_c[c]):
        validated = match_c[cat] + mismatch_c[cat]
        by_type[cat] = {
            "total_rows": total_c[cat],
            "skip": skip_c[cat],
            "validated": validated,
            "match": match_c[cat],
            "mismatch": mismatch_c[cat],
            "mismatch_rate": round(mismatch_c[cat] / validated, 4) if validated else None,
            "examples": examples.get(cat, []),
        }

    focus_summary = {k: by_type[k] for k in focus if k in by_type}
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "root": root,
        "files_scanned": files_scanned,
        "rows_with_instrument_count": rows_scanned,
        "focus_types_T03_T06_T07": focus_summary,
        "by_type": by_type,
    }


def print_report(report: dict[str, Any]) -> None:
    print("=== 包名字段核对扫描报告 ===")
    print(f"root: {report['root']}")
    print(f"files: {report['files_scanned']}  rows: {report['rows_with_instrument_count']}")
    print("\n--- T03 / T06 / T07 重点类型 ---")
    for cat, stats in report.get("focus_types_T03_T06_T07", {}).items():
        rate = stats.get("mismatch_rate")
        rate_s = f"{100 * rate:.1f}%" if rate is not None else "n/a"
        print(
            f"{cat}: validated={stats['validated']} match={stats['match']} "
            f"mismatch={stats['mismatch']} ({rate_s})"
        )
    print("\n--- 全部分类 ---")
    for cat, stats in report["by_type"].items():
        if stats["validated"] == 0:
            print(f"{cat}: total={stats['total_rows']} (all skip)")
            continue
        rate = stats["mismatch_rate"]
        print(
            f"{cat}: validated={stats['validated']} match={stats['match']} "
            f"mismatch={stats['mismatch']} ({100 * rate:.1f}%)"
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Scan pack name field consistency coverage")
    parser.add_argument(
        "--root",
        default=os.path.join(os.path.dirname(os.path.dirname(__file__)), "测试用例"),
        help="Root directory containing hospital xlsx folders",
    )
    parser.add_argument("--json-out", help="Write JSON report to path")
    args = parser.parse_args()
    report = scan_xlsx(args.root)
    print_report(report)
    if args.json_out:
        os.makedirs(os.path.dirname(args.json_out) or ".", exist_ok=True)
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"\nWrote {args.json_out}")


if __name__ == "__main__":
    main()
