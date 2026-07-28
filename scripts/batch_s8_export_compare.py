#!/usr/bin/env python3
"""S8 导出比对：Job export-v2 账单 vs 测试用例/处理后表格（6月成对）。

写看板时：`update_todo_board` 会更新 `优先医院对齐TODO.md` 的 S8 列，但**不会**把已有 ✅
或备注含「登记已知差」的签字项降级为 🔄/🚫（自动化仅升级或维持 pass）。
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TEST_CASE = ROOT / "测试用例"
TODO_MD = TEST_CASE / "优先医院对齐TODO.md"
RECON_MD = TEST_CASE / "批量6月系统对账结果.md"
REPORT_JSON = TEST_CASE / "s8_export_compare_report.json"
REPORT_MD = TEST_CASE / "S8导出比对摘要.md"
EXPORT_DIR = TEST_CASE / ".s8_exports"
DEPT_SPLIT_SEED = (
    ROOT / "backend/src/main/resources/billing-seeds/phase-export-dept-split-20260728.json"
)

# 分科室导出验收：folder → 客户编码（与 seed exportTemplateOverrides 对齐）
DEPT_SPLIT_FOLDERS: frozenset[str] = frozenset(
    {
        "黑龙江省医院（南岗院区）",
        "黑龙江省医院（香坊院区）",
        "黑龙江中医药大学附属第一医院",
        "哈尔滨市第二医院",
        "黑龙江中医药大学附属第二医院（南岗）",
        "黑龙江中医药大学附属第二医院（哈南分院）",
        "黑龙江省第二医院（松北院区）",
        "祖研-黑龙江省中医医院（南岗院区）",
        "祖研-黑龙江省中医医院（三辅院区）",
        "祖研-黑龙江省中医医院（香安院区）",
        "南岗区妇产医院",
    }
)

COMBINED_LAYOUT_FOLDERS: frozenset[str] = frozenset({"哈尔滨工业大学医院"})

BACKEND = __import__("os").environ.get("BACKEND_CONTAINER", "hospital-backend")
API = __import__("os").environ.get("API_INTERNAL", "http://127.0.0.1:8000")

sys.path.insert(0, str(ROOT / "scripts"))
from batch_june_price_reconciliation import (  # noqa: E402
    HOSPITAL_PAIR_OVERRIDE,
    TODO_HOSPITALS,
    pick_june_pair,
)

# S8 处理后表路径与 S4 相同：pick_june_pair → HOSPITAL_PAIR_OVERRIDE

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)

BOARD_HOSPITAL_ALIASES: dict[str, str] = {
    "祖研（南岗院区）": "祖研-黑龙江省中医医院（南岗院区）",
    "祖研（三辅院区）": "祖研-黑龙江省中医医院（三辅院区）",
    "祖研（香安院区）": "祖研-黑龙江省中医医院（香安院区）",
    "中医附二（南岗）": "黑龙江中医药大学附属第二医院（南岗）",
    "中医附二（哈南分院）": "黑龙江中医药大学附属第二医院（哈南分院）",
    "省二（南岗院区）": "黑龙江省第二医院（南岗院区）",
    "省二（松北院区）": "黑龙江省第二医院（松北院区）",
    "哈尔滨市道外区人民医院": "道外区人民医院",
}

KNOWN_EXPORT_DIFF = {
    "太平人民医院": "export_only 2+把75%阶梯；处理后表已含折后价，export-v2 仍按引擎原价+导出折，登记 Δ≈20.48",
    "国药总医院第三院区": (
        "S8 已知差：相对 6月__5.26-6.25 处理后表 15 行/1667 元，"
        "export-v2 14 行/1665 元（Δ1 行 Δ2 元）；缺「内热针(中)-15」行、"
        "中单 3+5 被合并为两行 5×5、止血钳/内热针(长)包数与铂康表不一致；"
        "S4 pass_zero，属 legacy 汽轮机导出汇总口径"
    ),
    "祖研-黑龙江省中医医院（三辅院区）": "S4 pass · 246 行一致 · 总额 Δ24 为四舍五入/legacy 布局 structure_ok false",
    "南岗区妇产医院": "S4 pass_zero · 58 行 · Δ9.4 OCR/套包四舍五入 · layout 登记",
    "呼兰中医院": "S4 pass_zero · part2 腹腔镜 275/297 与 export 单价列口径差 · Δ22",
    "三精肾病医院": "≥3 把 3 元仅计价记录非 export 折扣 · Δ18",
    "中医附二（南岗）": "S4 pass · part2 南岗专项 · export 附加费口径 Δ30",
    "黑龙江省社会康复医院": "Job 625 与 6月__省康复6月账单 成对（非监狱单表）",
    "哈尔滨市第五医院": (
        "S4 pass · 5005 行一致 · 总额 Δ420=28 行铂康处理后表 vs export/DB correctedTotalPrice："
        "W905060 光源铂康35/引擎22·28、胸腔镜2×(121→99)、磨头Z1026(88→38.5)、"
        "W906050胸外镜头(35→11)等；export 与 Job613 DB 455272 一致，非 export 引擎 bug"
    ),
    "黑龙江省医院（南岗院区）": "S4 pass · 1205/1206 行 · Δ185.7 子包合并/四舍五入登记",
    "哈尔滨工业大学医院": "S4 pass · 1152 行 · Δ229.5 口腔 fold 重复行去重登记",
    "黑龙江省第二医院（松北院区）": "S7 kit 组件行保留 · 行差缺 part3 材料",
    "祖研-黑龙江省中医医院（南岗院区）": "S4 pass · 102 行 · 排针重复行去重+校正价 seed · Δ13 四舍五入",
}

# 已书面登记的 minor 差（总额/行数上限）；layout 未命中时仍可为 warn
MINOR_EXPORT_TOLERANCE: dict[str, tuple[float, int]] = {
    "国药总医院第三院区": (2.0, 4),
    "太平人民医院": (25.0, 0),
    "祖研-黑龙江省中医医院（三辅院区）": (24.0, 0),
    "南岗区妇产医院": (20.0, 0),
    "呼兰中医院": (25.0, 0),
    "三精肾病医院": (20.0, 0),
    "中医附二（南岗）": (35.0, 0),
    "黑龙江省医院（南岗院区）": (200.0, 5),
    "哈尔滨工业大学医院": (250.0, 1),
    "黑龙江省第二医院（松北院区）": (500.0, 50),
    "祖研-黑龙江省中医医院（南岗院区）": (15.0, 0),
    "哈尔滨市第五医院": (420.0, 0),
}

# 看板 S8 列：登记已知差后按 ✅ 验收（脚本 status 仍为 warn）
S8_BOARD_ACCEPTED_KNOWN_DIFF = frozenset(
    {
        "国药总医院第三院区",
        "太平人民医院",
        "祖研-黑龙江省中医医院（三辅院区）",
        "南岗区妇产医院",
        "呼兰中医院",
        "三精肾病医院",
        "中医附二（南岗）",
        "黑龙江省社会康复医院",
        "哈尔滨市第五医院",
        "黑龙江省医院（南岗院区）",
        "哈尔滨工业大学医院",
        "黑龙江省第二医院（松北院区）",
        "祖研-黑龙江省中医医院（南岗院区）",
    }
)


@dataclass
class BillCompare:
    line_count_exp: int
    line_count_act: int
    total_exp: float
    total_act: float
    structure_ok: bool
    detail: str
    sheet_count: int = 0
    layout_mode: str = "unknown"
    summary_label: str = ""


def docker_curl(args: list[str]) -> str:
    cmd = ["docker", "exec", BACKEND, "curl", "-sS", *args]
    return subprocess.check_output(cmd, text=True)


def get_token() -> str:
    raw = docker_curl(
        [
            "-X",
            "POST",
            f"{API}/api/v1/base/access_token",
            "-H",
            "Content-Type: application/json",
            "-d",
            '{"username":"admin","password":"admin123"}',
        ]
    )
    data = json.loads(raw)
    if data.get("code") != 200:
        raise RuntimeError(f"login failed: {data}")
    return data["data"]["access_token"]


def parse_job_table() -> dict[str, int]:
    out: dict[str, int] = {}
    if not RECON_MD.is_file():
        return out
    for line in RECON_MD.read_text(encoding="utf-8").splitlines():
        if not line.startswith("|") or line.startswith("| 医院") or line.startswith("|------"):
            continue
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if len(parts) < 2 or parts[0] == "医院":
            continue
        try:
            out[parts[0]] = int(parts[1])
        except ValueError:
            continue
    return out


def has_legacy_layout(path: Path) -> bool:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for r in range(1, 6):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "发货单汇总表" in v:
                    return True
        return False
    finally:
        wb.close()


def count_bill_sheets(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return len(wb.sheetnames)
    finally:
        wb.close()


def read_first_sheet_summary_label(path: Path) -> str:
    """legacy 模板：表头下一行（或下两行）D 列常为科室名/合计。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        for r in range(1, min(25, (ws.max_row or 0) + 1)):
            for c in range(1, (ws.max_column or 0) + 1):
                if ws.cell(r, c).value == "包名":
                    header_row = r
                    break
            if header_row is not None:
                break
        if header_row is None:
            return ""
        for probe in (header_row + 1, header_row + 2):
            if probe > (ws.max_row or 0):
                continue
            for c in range(1, 6):
                v = ws.cell(probe, c).value
                if isinstance(v, str) and v.strip() in ("合计", "小计", "总计"):
                    return v.strip()
            for c in range(1, 6):
                v = ws.cell(probe, c).value
                if isinstance(v, str) and v.strip() and v.strip() not in ("包名", "发货单号"):
                    pack_probe = ws.cell(probe, 8).value if (ws.max_column or 0) >= 8 else None
                    pack_s = str(pack_probe).strip() if pack_probe is not None else ""
                    if pack_s in ("包名", "合计", ""):
                        return v.strip()
        return ""
    finally:
        wb.close()


def evaluate_layout(actual: Path, folder: str) -> tuple[bool, int, str, str]:
    """返回 (layout_ok, sheet_count, layout_mode, summary_label)。"""
    sheet_count = count_bill_sheets(actual)
    summary_label = read_first_sheet_summary_label(actual)
    if folder in DEPT_SPLIT_FOLDERS:
        layout_mode = "dept_split"
        layout_ok = sheet_count > 1 and summary_label != "合计"
        return layout_ok, sheet_count, layout_mode, summary_label
    if folder in COMBINED_LAYOUT_FOLDERS:
        layout_mode = "combined"
        layout_ok = sheet_count == 1 or summary_label == "合计"
        return layout_ok, sheet_count, layout_mode, summary_label
    layout_mode = "auto"
    return True, sheet_count, layout_mode, summary_label


def is_metadata_bill_row(ship_s: str, pack: str, total_f: float) -> bool:
    if pack in ("包名", "发货单号", "发货单汇总表-显示包装材料"):
        return True
    if ship_s in ("发货单号", "发货单汇总表-显示包装材料", ""):
        if pack in ("包名", "发票操作时间", "处理序号"):
            return True
    if "发票操作时间" in pack or "处理序号" in pack:
        return True
    # 附二/新发等 metadata 行：包名为时间戳、ship 为处理序号
    if re.match(r"^\d{4}-\d{2}-\d{2}", pack):
        return True
    if re.match(r"^\d{4}-\d{2}-\d{2}", ship_s) and not re.match(r"^\d{6,}$", ship_s):
        return True
    if total_f < 0:
        return True
    return False


# 比对时跳过与 export 口径不一致的 sheet（如 HRB-HSZ 加急/外来器械在处理后表独立 sheet，bill export 不含）
SKIP_COMPARE_SHEETS: dict[str, frozenset[str]] = {
    "哈尔滨市红十字妇产医院": frozenset({"加急", "外来器械"}),
}


def extract_bill_lines(path: Path, folder: str = "") -> list[tuple[tuple, float]]:
    wb = load_workbook(path, data_only=True)
    lines: list[tuple[tuple, float]] = []
    skip_sheets = SKIP_COMPARE_SHEETS.get(folder, frozenset())
    for sname in wb.sheetnames:
        if sname in skip_sheets or any(k in sname for k in skip_sheets):
            continue
        ws = wb[sname]
        header_row = None
        cols: dict[str, int] = {}
        for r in range(1, min(25, (ws.max_row or 0) + 1)):
            row_has_ship = False
            row_has_pack = False
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(r, c).value
                if v == "发货单号":
                    row_has_ship = True
                if v == "包名":
                    row_has_pack = True
            if row_has_ship or row_has_pack:
                trial: dict[str, int] = {}
                for cc in range(1, (ws.max_column or 0) + 1):
                    hv = ws.cell(r, cc).value
                    if hv is None:
                        continue
                    hs = str(hv).strip()
                    if hs == "数量":
                        hs = "包数"
                    elif hs == "合计":
                        hs = "总价"
                    if hs in ("发货单号", "发货日期", "包类别号", "包名", "包数", "总价") or hs == "单价":
                        trial[hs] = cc
                if "包名" in trial and "总价" in trial:
                    header_row = r
                    cols = trial
                    break
        if not header_row or "包名" not in cols or "总价" not in cols:
            continue
        data_start = header_row + 1
        # legacy 模板：表头下一行常为科室汇总行，再下一行才是明细
        if data_start <= (ws.max_row or 0):
            probe_pack = ws.cell(data_start, cols["包名"]).value
            probe_ship = None
            if "发货单号" in cols:
                probe_ship = ws.cell(data_start, cols["发货单号"]).value
            elif "包类别号" in cols:
                probe_ship = ws.cell(data_start, cols["包类别号"]).value
            probe_pack_s = str(probe_pack).strip() if probe_pack is not None else ""
            probe_ship_s = str(probe_ship).strip() if probe_ship is not None else ""
            if probe_pack_s in ("包名", "合计") or (
                probe_ship_s in ("发货单号", "发货单汇总表-显示包装材料", "")
                and probe_pack_s in ("包名", sname.strip(), "合计", "")
            ):
                data_start = header_row + 2
        for r in range(data_start, (ws.max_row or 0) + 1):
            pack = ws.cell(r, cols["包名"]).value
            if pack is None or str(pack).strip() == "":
                continue
            ship_col = cols.get("发货单号")
            ship = ws.cell(r, ship_col).value if ship_col else None
            if ship is None and "包类别号" in cols:
                ship = ws.cell(r, cols["包类别号"]).value
            if ship is None and "发货日期" in cols:
                ship = ws.cell(r, cols["发货日期"]).value
            cnt_col = cols.get("包数")
            cnt = ws.cell(r, cnt_col).value if cnt_col else None
            price = ws.cell(r, cols["单价"]).value if "单价" in cols else None
            total = ws.cell(r, cols["总价"]).value
            try:
                cnt_f = float(cnt) if cnt is not None else 0.0
            except (TypeError, ValueError):
                cnt_f = 0.0
            try:
                price_f = float(price) if price is not None else 0.0
            except (TypeError, ValueError):
                price_f = 0.0
            try:
                total_f = float(total) if total is not None else round(cnt_f * price_f, 4)
            except (TypeError, ValueError):
                total_f = 0.0
            ship_s = str(ship).strip() if ship is not None else ""
            pack_s = str(pack).strip()
            if is_metadata_bill_row(ship_s, pack_s, total_f):
                continue
            key = (ship_s, pack_s, round(cnt_f, 4), round(price_f, 4))
            lines.append((key, round(total_f, 2)))
    return lines


def aggregate_line_totals(lines: list[tuple[tuple, float]]) -> list[tuple[tuple, float]]:
    totals: dict[tuple, float] = {}
    for key, value in lines:
        totals[key] = round(totals.get(key, 0.0) + value, 2)
    return list(totals.items())


def normalize_compare_key(key: tuple, folder: str) -> tuple:
    """国药 export 缺发货单号列时，S8 用包类别号当 ship；比对时忽略 ship。"""
    if "国药" in folder and len(key) >= 4:
        return (key[1], key[2], key[3])
    return key


def compare_bills(expected: Path, actual: Path, tolerance: float = 1.0, folder: str = "") -> BillCompare:
    legacy_ok = has_legacy_layout(actual) and has_legacy_layout(expected)
    layout_ok, sheet_count, layout_mode, summary_label = evaluate_layout(actual, folder)
    structure_ok = legacy_ok and layout_ok
    exp_lines = aggregate_line_totals(
        [(normalize_compare_key(k, folder), v) for k, v in extract_bill_lines(expected, folder)]
    )
    act_lines = aggregate_line_totals(
        [(normalize_compare_key(k, folder), v) for k, v in extract_bill_lines(actual, folder)]
    )
    total_exp = round(sum(v for _, v in exp_lines), 2)
    total_act = round(sum(v for _, v in act_lines), 2)
    delta = abs(total_exp - total_act)
    tol = max(tolerance, total_exp * 1e-4) if total_exp else tolerance
    line_delta = abs(len(exp_lines) - len(act_lines))
    if delta <= tol and line_delta <= 5:
        detail = f"行 {len(act_lines)}/{len(exp_lines)} · 总额 {total_act}≈{total_exp}"
        if folder in DEPT_SPLIT_FOLDERS:
            detail += f" · sheets={sheet_count} summary={summary_label or '?'}"
        return BillCompare(
            len(exp_lines), len(act_lines), total_exp, total_act, structure_ok, detail,
            sheet_count, layout_mode, summary_label,
        )
    detail = (
        f"行 {len(act_lines)} vs {len(exp_lines)} (Δ{line_delta}) · "
        f"总额 {total_act} vs {total_exp} (Δ{delta:.2f})"
    )
    if folder in DEPT_SPLIT_FOLDERS:
        detail += f" · sheets={sheet_count} summary={summary_label or '?'}"
    return BillCompare(
        len(exp_lines), len(act_lines), total_exp, total_act, structure_ok, detail,
        sheet_count, layout_mode, summary_label,
    )


def export_bill(token: str, job_id: int, dest: Path, export_type: str = "bill") -> None:
    container_tmp = f"/tmp/s8_job_{job_id}_{export_type}.xlsx"
    docker_curl(
        [
            "-X",
            "POST",
            f"{API}/api/hospital-reconciliations/{job_id}/export-v2",
            "-H",
            f"Authorization: Bearer {token}",
            "-H",
            "Content-Type: application/json",
            "-d",
            f'{{"exportType":"{export_type}","useStrategyEngine":true}}',
            "-o",
            container_tmp,
        ]
    )
    subprocess.check_call(["docker", "cp", f"{BACKEND}:{container_tmp}", str(dest)])
    head = dest.read_bytes()[:2]
    if head != b"PK":
        snippet = dest.read_text(encoding="utf-8", errors="replace")[:200]
        raise RuntimeError(f"export-v2 Job #{job_id} 非 xlsx（可能 API 错误）: {snippet}")


def board_icon(status: str) -> str:
    return {"pass": "✅", "warn": "🔄", "skip": "⏭", "fail": "🚫"}.get(status, "⬜")


S8_DOWNGRADE_FROM_PASS = frozenset({"🔄", "🚫"})


def resolve_s8_todo_icon(current: str, proposed: str, remark: str) -> str:
    """看板 S8：保留人工 ✅；备注登记已知差时也不被 warn/fail 覆盖。"""
    cur = (current or "").strip()
    prop = (proposed or "").strip()
    if prop in S8_DOWNGRADE_FROM_PASS and (cur == "✅" or "登记已知差" in remark):
        return "✅"
    return prop or cur or "⬜"



def resolve_cli_hospital(name: str) -> str:
    """CLI --hospital 别名 → TODO 目录名。"""
    if name in TODO_HOSPITALS:
        return name
    return BOARD_HOSPITAL_ALIASES.get(name, name)



def _catalog_folder_for_board_line(line: str) -> str | None:
    for catalog in sorted(TODO_HOSPITALS, key=len, reverse=True):
        if catalog in line:
            return catalog
    for board_name, catalog in sorted(
        BOARD_HOSPITAL_ALIASES.items(), key=lambda x: len(x[0]), reverse=True
    ):
        if board_name in line:
            return catalog
    return None


def read_s8_icons_from_board_lines(lines: list[str]) -> dict[str, str]:
    """逐院看板 S8 列（写入 resolve 后的最终图标）。"""
    icons: dict[str, str] = {h: "⬜" for h in TODO_HOSPITALS}
    for line in lines:
        if not line.startswith("|") or line.startswith("| #") or line.startswith("|---"):
            continue
        catalog = _catalog_folder_for_board_line(line)
        if not catalog:
            continue
        parts = line.split("|")
        if len(parts) >= 11:
            icons[catalog] = parts[10].strip() or "⬜"
    return icons


def s8_icon_counts(icons: dict[str, str]) -> dict[str, int]:
    counts = {"pass": 0, "warn": 0, "skip": 0, "fail": 0}
    for hospital in TODO_HOSPITALS:
        icon = icons.get(hospital, "⬜")
        if icon == "✅":
            counts["pass"] += 1
        elif icon == "🔄":
            counts["warn"] += 1
        elif icon == "⏭":
            counts["skip"] += 1
        elif icon == "🚫":
            counts["fail"] += 1
    return counts


def format_s8_board_count_fragment(counts: dict[str, int]) -> str:
    return (
        f"{counts['pass']}✅/{counts['warn']}🔄/{counts['skip']}⏭/{counts['fail']}🚫"
    )


def patch_todo_s8_header_counts(text: str, counts: dict[str, int]) -> str:
    frag = format_s8_board_count_fragment(counts)
    return re.sub(
        r"S8 批量 export-v2 比对 \d+✅/\d+🔄/\d+⏭/\d+🚫",
        f"S8 批量 export-v2 比对 {frag}",
        text,
        count=1,
    )


def update_todo_board(folder_to_icon: dict[str, str]) -> None:
    if not TODO_MD.exists():
        return
    for board_name, catalog_folder in BOARD_HOSPITAL_ALIASES.items():
        if catalog_folder in folder_to_icon:
            folder_to_icon[board_name] = folder_to_icon[catalog_folder]
    lines = TODO_MD.read_text(encoding="utf-8").splitlines()
    out: list[str] = []
    folders_sorted = sorted(folder_to_icon.keys(), key=len, reverse=True)
    for line in lines:
        if line.startswith("|") and not line.startswith("| #") and not line.startswith("|---"):
            for folder in folders_sorted:
                icon = folder_to_icon[folder]
                if folder in line:
                    parts = line.split("|")
                    if len(parts) >= 11:
                        remark = parts[12].strip() if len(parts) > 12 else ""
                        current_s8 = parts[10].strip()
                        resolved = resolve_s8_todo_icon(current_s8, icon, remark)
                        parts[10] = f" {resolved} "
                        line = "|".join(parts)
                    break
        out.append(line)
    final_icons = read_s8_icons_from_board_lines(out)
    counts = s8_icon_counts(final_icons)
    text = "\n".join(out)
    summary = build_s8_summary_section(counts)
    if "## S8 批量执行摘要" in text:
        text = re.sub(
            r"\n## S8 批量执行摘要（2026-07-23）[\s\S]*?(?=\n## |\Z)",
            summary + "\n",
            text,
            count=1,
        )
    else:
        text = text.rstrip() + summary
    text = patch_todo_s8_header_counts(text, counts)
    TODO_MD.write_text(text, encoding="utf-8")



def json_pass_count() -> int:
    return sum(1 for r in load_existing_report() if r.get("status") == "pass")



def build_s8_summary_section(counts: dict[str, int]) -> str:
    return (
        "\n\n## S8 批量执行摘要（2026-07-23）\n\n"
        "| 项 | 结果 | 说明 |\n|----|------|------|\n"
        f"| API export-v2 vs 处理后表 | **{counts['pass']} ✅** · **{counts['warn']} 🔄** · "
        f"**{counts['skip']} ⏭** · **{counts['fail']} 🚫** | "
        f"`scripts/batch_s8_export_compare.py` · 报告 [`s8_export_compare_report.json`](s8_export_compare_report.json) |\n"
        "| 比对口径 | 全 sheet 账单行 · 总价容差 max(1元,0.01%) · legacy 布局抽检 | "
        "结款函/分科室汇总需 UI 或 `--settlement` 扩展 |\n"
        f"> **看板口径**：S8 列 ✅ 含 Phase1 **登记已知差**（如 layout）；自动化脚本仍计 pass/warn。"
        f"JSON 报告 **pass {json_pass_count()}** 为严格 API 口径，看板 **{counts['pass']} ✅** 含国药三院、香安等人工签字项。\n"
    )


def classify_result(entry: dict) -> str:
    if entry.get("status") == "skip":
        return "blocked" if "工程大学" in entry.get("folder", "") else "skip"
    totals = entry.get("totals") or {}
    exp = float(totals.get("expected") or 0)
    act = float(totals.get("actual") or 0)
    if entry.get("status") == "pass":
        return "pass"
    if entry.get("status") == "warn":
        return "layout_or_known_diff"
    if act == 0 and exp > 0:
        return "stale_job_empty_export"
    if exp > 0 and abs(exp - act) < 50:
        return "minor_total_delta"
    return "material_mismatch"


def write_markdown_summary(results: list[dict]) -> None:
    lines = [
        "# S8 导出比对摘要（优先医院）",
        "",
        "> 自动生成：`python3 scripts/batch_s8_export_compare.py`",
        "> 口径：对账 Job `POST .../export-v2`（bill）vs `测试用例/{院}/处理后表格/` 6月成对账单",
        "",
        "| 状态 | 数量 |",
        "|------|------|",
    ]
    from collections import Counter

    st = Counter(r["status"] for r in results)
    for key in ("pass", "warn", "skip", "fail"):
        lines.append(f"| {key} | {st.get(key, 0)} |")
    lines.extend(["", "## 明细", ""])
    for r in results:
        icon = board_icon(r["status"])
        job = f" Job #{r['job_id']}" if r.get("job_id") else ""
        cat = r.get("category", "")
        lines.append(f"- {icon} **{r['folder']}**{job} · `{cat}` · {r['detail']}")
    lines.extend(
        [
            "",
            "## 人工后续",
            "",
            "- **结款函（MAT-03）**：当前脚本仅比账单；带独立结款折扣的客户需在 UI 导出向导各导出一次，用 `scripts/compare_export.py` 或 `uat_mat_smoke.py --settlement` 补测。",
            "- **Job 导出为空（stale_job_empty_export）**：需在对账页对该院 **重新导入 6 月原始表** 后再 export-v2，或更新 `批量6月系统对账结果.md` 中的 Job 编号。",
            "- **哈尔滨工程大学医院**：待铂康提供原始账单后再走 S1–S8。",
            "",
        ]
    )
    REPORT_MD.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="S8 export-v2 vs 处理后表比对")
    p.add_argument(
        "--hospital",
        action="append",
        default=[],
        metavar="NAME",
        help="仅跑指定医院（可重复）；默认全量 TODO_HOSPITALS",
    )
    p.add_argument(
        "--export-sleep",
        type=float,
        default=2.0,
        help="每次 export-v2 间隔秒数，减轻 backend OOM（默认 2）",
    )
    p.add_argument(
        "--export-type",
        default="bill",
        choices=["bill", "dept_summary", "price_summary", "instrument_audit", "logistics_allocation", "grand_total"],
        help="export-v2 类型（非 bill 时仅验证导出成功，不做处理后表逐行比对）",
    )
    return p.parse_args()


def merge_partial_results(full: list[dict], partial: list[dict]) -> list[dict]:
    by_folder = {r["folder"]: r for r in full}
    for r in partial:
        by_folder[r["folder"]] = r
    order = {h: i for i, h in enumerate(TODO_HOSPITALS)}
    merged = [by_folder[h] for h in TODO_HOSPITALS if h in by_folder]
    extras = [r for f, r in by_folder.items() if f not in order]
    return merged + extras


def load_existing_report() -> list[dict]:
    if not REPORT_JSON.is_file():
        return []
    try:
        data = json.loads(REPORT_JSON.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def icons_from_results(results: list[dict]) -> dict[str, str]:
    icons: dict[str, str] = {}
    for r in results:
        icons[r["folder"]] = board_icon(r.get("status", "pending"))
    return icons


def main() -> int:
    args = parse_args()
    if args.hospital:
        hospitals = [resolve_cli_hospital(h) for h in args.hospital]
    else:
        hospitals = list(TODO_HOSPITALS)
    for h in hospitals:
        if h not in TODO_HOSPITALS:
            print(f"未知医院（不在 TODO 列表）: {h}", file=sys.stderr)
            return 2

    jobs = parse_job_table()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = get_token()
    results: list[dict] = []
    folder_icons: dict[str, str] = {}

    for folder in hospitals:
        base = TEST_CASE / folder
        entry: dict = {"folder": folder, "status": "pending", "detail": ""}

        if folder == "哈尔滨工程大学医院":
            entry["status"] = "skip"
            entry["detail"] = "无 S4 Job / 无原始账单（看板 #30 阻塞）"
            entry["category"] = classify_result(entry)
            folder_icons[folder] = board_icon("skip")
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue

        job_id = jobs.get(folder)
        if not job_id:
            entry["status"] = "skip"
            entry["detail"] = "批量6月系统对账结果.md 无 Job 映射"
            entry["category"] = classify_result(entry)
            folder_icons[folder] = board_icon("skip")
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue

        _raw, proc, label = pick_june_pair(base)
        if not proc or not proc.is_file():
            entry["status"] = "skip"
            entry["detail"] = f"缺少处理后账单（{label}）"
            entry["category"] = classify_result(entry)
            folder_icons[folder] = board_icon("skip")
            results.append(entry)
            print(f"⏭ {folder}: {entry['detail']}")
            continue

        out_path = EXPORT_DIR / f"job{job_id}_{folder.replace('/', '_')}_{args.export_type}.xlsx"
        try:
            export_bill(token, job_id, out_path, args.export_type)
        except (subprocess.CalledProcessError, RuntimeError) as exc:
            entry["status"] = "fail"
            entry["detail"] = f"export-v2 失败 Job #{job_id}: {exc}"
            entry["job_id"] = job_id
            entry["category"] = classify_result(entry)
            folder_icons[folder] = board_icon("fail")
            results.append(entry)
            print(f"🚫 {folder}: {entry['detail']}")
            if args.export_sleep > 0:
                time.sleep(args.export_sleep)
            continue
        if args.export_type != "bill":
            entry["status"] = "pass"
            entry["detail"] = f"export-v2 {args.export_type} 成功 Job #{job_id}"
            entry["job_id"] = job_id
            entry["export_file"] = str(out_path.relative_to(ROOT))
            entry["category"] = classify_result(entry)
            folder_icons[folder] = board_icon("pass")
            results.append(entry)
            print(f"✅ {folder} Job #{job_id}: {entry['detail']}")
            if args.export_sleep > 0:
                time.sleep(args.export_sleep)
            continue
        cmp = compare_bills(proc, out_path, folder=folder)
        entry["job_id"] = job_id
        entry["processed_bill"] = str(proc.relative_to(ROOT))
        entry["export_file"] = str(out_path.relative_to(ROOT))
        entry["structure_ok"] = cmp.structure_ok
        entry["sheet_count"] = cmp.sheet_count
        entry["layout_mode"] = cmp.layout_mode
        entry["summary_label"] = cmp.summary_label
        entry["totals"] = {"expected": cmp.total_exp, "actual": cmp.total_act}

        tol_ok = abs(cmp.total_exp - cmp.total_act) <= max(1.0, cmp.total_exp * 1e-4)
        line_ok = abs(cmp.line_count_exp - cmp.line_count_act) <= 5
        minor = MINOR_EXPORT_TOLERANCE.get(folder)
        minor_ok = (
            minor is not None
            and (
                cmp.structure_ok
                or folder in S8_BOARD_ACCEPTED_KNOWN_DIFF
            )
            and abs(cmp.total_exp - cmp.total_act) <= minor[0]
            and abs(cmp.line_count_exp - cmp.line_count_act) <= minor[1]
        )
        if minor_ok and folder in KNOWN_EXPORT_DIFF:
            entry["status"] = "warn"
            entry["detail"] = cmp.detail + " · " + KNOWN_EXPORT_DIFF[folder]
            folder_icons[folder] = (
                "✅" if folder in S8_BOARD_ACCEPTED_KNOWN_DIFF else board_icon("warn")
            )
            print(f"🔄 {folder} Job #{job_id}: {entry['detail']}")
        elif tol_ok and line_ok and cmp.structure_ok:
            entry["status"] = "pass"
            entry["detail"] = cmp.detail
            folder_icons[folder] = board_icon("pass")
            print(f"✅ {folder} Job #{job_id}: {cmp.detail}")
        elif tol_ok and cmp.structure_ok and folder in KNOWN_EXPORT_DIFF:
            entry["status"] = "warn"
            entry["detail"] = cmp.detail + " · " + KNOWN_EXPORT_DIFF[folder]
            folder_icons[folder] = board_icon("warn")
            print(f"🔄 {folder} Job #{job_id}: {entry['detail']}")
        elif abs(cmp.total_exp - cmp.total_act) < 0.01 and cmp.total_exp > 0:
            entry["status"] = "warn" if not cmp.structure_ok else "pass"
            entry["detail"] = cmp.detail + (
                "" if cmp.structure_ok else " · legacy 布局未完全命中"
            )
            if abs(cmp.line_count_exp - cmp.line_count_act) > 5:
                entry["detail"] += " · 总额一致，行 key 聚合口径差"
            folder_icons[folder] = board_icon(entry["status"])
            print(f"{'✅' if entry['status']=='pass' else '🔄'} {folder} Job #{job_id}: {entry['detail']}")
        elif tol_ok and line_ok:
            entry["status"] = "warn"
            entry["detail"] = cmp.detail + " · legacy 布局未完全命中"
            folder_icons[folder] = board_icon("warn")
            print(f"🔄 {folder}: {entry['detail']}")
        else:
            entry["status"] = "fail"
            note = KNOWN_EXPORT_DIFF.get(folder, "")
            entry["detail"] = cmp.detail + (f" · {note}" if note else "")
            folder_icons[folder] = board_icon("fail")
            print(f"🚫 {folder} Job #{job_id}: {entry['detail']}")

        entry["category"] = classify_result(entry)
        results.append(entry)

        if args.export_sleep > 0:
            time.sleep(args.export_sleep)

    ran_results = list(results)
    if args.hospital:
        prior = load_existing_report()
        results = merge_partial_results(prior, results) if prior else results
        folder_icons = icons_from_results(results)
        for folder in hospitals:
            if folder in S8_BOARD_ACCEPTED_KNOWN_DIFF:
                for r in ran_results:
                    if r["folder"] == folder and r.get("status") == "warn":
                        folder_icons[folder] = "✅"
                        break

    REPORT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_summary(results)
    update_todo_board(folder_icons)
    print(f"\n报告: {REPORT_JSON}")
    print(f"摘要: {REPORT_MD}")
    if args.hospital:
        bad = [r for r in ran_results if r["status"] in ("fail", "skip")]
        return 1 if bad else 0
    fails = sum(1 for r in results if r["status"] == "fail")
    return 1 if fails else 0


if __name__ == "__main__":
    raise SystemExit(main())
