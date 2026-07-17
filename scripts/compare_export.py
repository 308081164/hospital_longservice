#!/usr/bin/env python3
"""
Compare two hospital export Excel files (MAT-02/MAT-03 diff helper).

Usage:
  python scripts/compare_export.py expected.xlsx actual.xlsx [--sheet 0] [--tolerance 0.01]
  python scripts/compare_export.py --batch manifest.json
  python scripts/compare_export.py expected.xlsx actual.xlsx --settlement-total-col 8

Exit code 0 when within tolerance, 1 when mismatches found.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def cell_value(cell) -> str | float | None:
    if cell is None or cell.value is None:
        return None
    v = cell.value
    if isinstance(v, (int, float)):
        return float(v)
    return str(v).strip()


def compare_sheets(path_a: Path, path_b: Path, sheet_index: int, tolerance: float) -> list[str]:
    wb_a = load_workbook(path_a, data_only=True)
    wb_b = load_workbook(path_b, data_only=True)
    if sheet_index >= len(wb_a.sheetnames) or sheet_index >= len(wb_b.sheetnames):
        return [f"Sheet index {sheet_index} out of range"]
    name_a = wb_a.sheetnames[sheet_index]
    name_b = wb_b.sheetnames[sheet_index]
    ws_a = wb_a[name_a]
    ws_b = wb_b[name_b]
    diffs: list[str] = []
    max_row = max(ws_a.max_row or 0, ws_b.max_row or 0)
    max_col = max(ws_a.max_column or 0, ws_b.max_column or 0)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            va = cell_value(ws_a.cell(row=r, column=c))
            vb = cell_value(ws_b.cell(row=r, column=c))
            if va == vb:
                continue
            if isinstance(va, float) and isinstance(vb, float) and abs(va - vb) <= tolerance:
                continue
            diffs.append(f"R{r}C{c}: {va!r} != {vb!r} (sheets {name_a}/{name_b})")
    return diffs


def settlement_total(path: Path, sheet_index: int, total_col: int) -> float | None:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    total = 0.0
    found = False
    for r in range(2, (ws.max_row or 0) + 1):
        val = cell_value(ws.cell(row=r, column=total_col))
        if isinstance(val, float):
            total += val
            found = True
    return round(total, 2) if found else None


def run_batch(manifest_path: Path, default_tolerance: float) -> int:
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    pairs = data if isinstance(data, list) else data.get("pairs", [])
    failed = 0
    for item in pairs:
        expected = Path(item["expected"])
        actual = Path(item["actual"])
        tolerance = float(item.get("tolerance", default_tolerance))
        sheet = int(item.get("sheet", 0))
        label = item.get("label", expected.name)
        diffs = compare_sheets(expected, actual, sheet, tolerance)
        if diffs:
            failed += 1
            print(f"FAIL {label}: {len(diffs)} diff(s)")
            for line in diffs[:5]:
                print(" ", line)
        else:
            print(f"OK   {label}")
    return 1 if failed else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare two export xlsx files")
    parser.add_argument("expected", type=Path, nargs="?")
    parser.add_argument("actual", type=Path, nargs="?")
    parser.add_argument("--batch", type=Path, help="JSON manifest of expected/actual pairs")
    parser.add_argument("--sheet", type=int, default=0, help="Sheet index (0-based)")
    parser.add_argument("--tolerance", type=float, default=0.01, help="Numeric tolerance (CNY)")
    parser.add_argument("--settlement-total-col", type=int, help="Validate settlement column sum")
    args = parser.parse_args()

    if args.batch:
        if not args.batch.exists():
            print("Manifest not found", file=sys.stderr)
            return 2
        return run_batch(args.batch, args.tolerance)

    if not args.expected or not args.actual:
        parser.print_help()
        return 2

    if not args.expected.exists() or not args.actual.exists():
        print("File not found", file=sys.stderr)
        return 2

    diffs = compare_sheets(args.expected, args.actual, args.sheet, args.tolerance)
    if args.settlement_total_col:
        exp_total = settlement_total(args.expected, args.sheet, args.settlement_total_col)
        act_total = settlement_total(args.actual, args.sheet, args.settlement_total_col)
        if exp_total is not None and act_total is not None:
            delta = abs(exp_total - act_total)
            print(f"Settlement column total: expected={exp_total}, actual={act_total}, delta={delta:.4f}")
            if delta > args.tolerance:
                diffs.append(f"Settlement total delta {delta} > {args.tolerance}")

    if not diffs:
        print("OK — no differences within tolerance")
        return 0
    print(f"Found {len(diffs)} difference(s):")
    for line in diffs[:50]:
        print(" ", line)
    if len(diffs) > 50:
        print(f"  ... and {len(diffs) - 50} more")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
