#!/usr/bin/env python3
"""Mark reconciliation rows urgent from processed bill 加急 sheet or sheet_name."""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)

from batch_june_price_reconciliation import pick_june_pair  # noqa: E402
from batch_s8_export_compare import load_job_map  # noqa: E402

MYSQL = ["docker", "exec", "hospital-mysql", "sh", "-c"]

DEFAULT_BILL = {
    615: ROOT / "测试用例/新发红十字医院/处理后表格/6月__新发红十字医院6月账单.xlsx",
    635: ROOT / "测试用例/哈尔滨仁胜医院/处理后表格/6月__仁胜6月账单.xlsx",
}


def mysql_exec(sql: str) -> str:
    cmd = MYSQL + [
        'mysql -uhospital -p"$MYSQL_PASSWORD" --default-character-set=utf8mb4 hospital -N -e '
        + repr(sql)
    ]
    return subprocess.check_output(cmd, text=True, cwd=ROOT).strip()


def urgent_keys_from_bill(xlsx: Path) -> set[tuple[str, str, int]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    keys: set[tuple[str, str, int]] = set()
    for sn in wb.sheetnames:
        if "加急" not in sn:
            continue
        ws = wb[sn]
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}
        order_i = next((idx[k] for k in idx if "发货单" in k or k == "orderNo"), None)
        pack_i = next((idx[k] for k in idx if "包名" in k or k == "packName"), None)
        count_i = next((idx[k] for k in idx if "包数" in k or k == "packCount"), None)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            order = str(row[order_i]).strip() if order_i is not None and row[order_i] is not None else ""
            pack = str(row[pack_i]).strip() if pack_i is not None and row[pack_i] is not None else ""
            count = int(float(row[count_i])) if count_i is not None and row[count_i] is not None else 1
            if order or pack:
                keys.add((order, pack, count))
    wb.close()
    return keys


def enrich_from_sheet_name(job_id: int) -> int:
    before = int(
        mysql_exec(
            f"SELECT COUNT(*) FROM hospital_reconciliation_row WHERE job_id={job_id} "
            "AND sheet_name LIKE '%加急%' AND (is_urgent IS NULL OR is_urgent=0)"
        )
        or "0"
    )
    if before:
        mysql_exec(
            f"UPDATE hospital_reconciliation_row SET is_urgent=1 WHERE job_id={job_id} "
            "AND sheet_name LIKE '%加急%' AND (is_urgent IS NULL OR is_urgent=0)"
        )
    return before


def enrich_from_bill(job_id: int, xlsx: Path, dry_run: bool) -> int:
    keys = urgent_keys_from_bill(xlsx)
    if not keys:
        return 0
    updated = 0
    for order, pack, count in keys:
        order_sql = order.replace("'", "''")
        pack_sql = pack.replace("'", "''")
        sql = (
            f"UPDATE hospital_reconciliation_row SET is_urgent=1 "
            f"WHERE job_id={job_id} AND pack_name='{pack_sql}' AND pack_count={count} "
            f"AND (is_urgent IS NULL OR is_urgent=0)"
        )
        if order:
            sql += f" AND order_no='{order_sql}'"
        if dry_run:
            cnt = mysql_exec(
                sql.replace("UPDATE hospital_reconciliation_row SET is_urgent=1", "SELECT COUNT(*) FROM hospital_reconciliation_row")
                .split(" AND (is_urgent")[0]
            )
            updated += int(cnt or 0)
        else:
            mysql_exec(sql)
            updated += 1
    return updated


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich is_urgent flags")
    parser.add_argument("--job-id", type=int, action="append", default=[])
    parser.add_argument("--job-map", type=Path, default=ROOT / "测试用例/job_baseline_stable.json")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    job_ids = list(args.job_id) or [615, 635]
    total = 0
    for jid in job_ids:
        total += enrich_from_sheet_name(jid)
        bill = DEFAULT_BILL.get(jid)
        if bill and bill.is_file():
            total += enrich_from_bill(jid, bill, args.dry_run)
            print(f"Job #{jid}: bill 加急 keys from {bill.name}")
    print(f"Done · ~{total} updates")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
