#!/usr/bin/env python3
"""S8 额外表格导出比对：export-v2 summary types vs 测试用例/处理后表格参考表。

用法:
  ./scripts/run-python-host.sh scripts/batch_s8_summary_export_compare.py --export-type price_summary
  ./scripts/run-python-host.sh scripts/batch_s8_summary_export_compare.py --export-type all
"""

from __future__ import annotations

import argparse
import fnmatch
import json
import sys
import time
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TEST_CASE = ROOT / "测试用例"
MANIFEST = ROOT / "scripts" / "summary_reference_manifest.json"
REPORT_JSON = TEST_CASE / "s8_summary_compare_report.json"
REPORT_MD = TEST_CASE / "S8汇总导出比对摘要.md"
EXPORT_DIR = TEST_CASE / ".s8_summary_exports"

BACKEND = __import__("os").environ.get("BACKEND_CONTAINER", "hospital-backend")
API = __import__("os").environ.get("API_INTERNAL", "http://127.0.0.1:8000")

sys.path.insert(0, str(ROOT / "scripts"))
from batch_s8_export_compare import docker_curl, get_token, parse_job_table  # noqa: E402

SUMMARY_TYPES = (
    "price_summary",
    "instrument_audit",
    "dept_summary",
    "logistics_allocation",
    "grand_summary",
)

KNOWN_SUMMARY_DIFF: dict[str, str] = {
    "哈尔滨市第五医院|dept_summary": "L3 allocation 列布局与铂康分科室汇总有差，登记 warn",
    "哈尔滨市第五医院|grand_summary": "merge 二门诊口径首版骨架",
    "黑龙江中医药大学附属第一医院|dept_summary": "附一首版各科室费用汇总骨架",
    "黑龙江中医药大学附属第一医院|logistics_allocation": "附一物流明细 vs 4列汇总首版",
    "黑龙江省医院（南岗院区)|price_summary": "3月参考 fallback",
    "黑龙江省医院（香坊院区)|price_summary": "3月参考 fallback",
    "黑龙江中医药大学附属第二医院（南岗）|price_summary": "包数据 xls 格式差",
}

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(2)


@dataclass
class CompareResult:
    hospital: str
    export_type: str
    status: str
    job_id: int | None
    reference: str | None
    total_exp: float | None
    total_act: float | None
    delta: float | None
    detail: str


def load_manifest() -> list[dict]:
    data = json.loads(MANIFEST.read_text(encoding="utf-8"))
    return data.get("entries", data if isinstance(data, list) else [])


def find_reference(folder: Path, glob_pattern: str, month: str) -> Path | None:
    proc = folder / "处理后表格"
    if not proc.is_dir():
        return None
    candidates = []
    for f in proc.iterdir():
        if f.suffix.lower() not in (".xlsx", ".xls"):
            continue
        if fnmatch.fnmatch(f.name, glob_pattern):
            candidates.append(f)
    if not candidates:
        return None
    month_files = [f for f in candidates if month in f.name]
    pool = month_files or candidates
    return sorted(pool, key=lambda p: p.name)[-1]


def sheet_numeric_total(path: Path) -> float:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total = 0.0
    for r in range(1, (ws.max_row or 0) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                total += float(v)
    return round(total, 2)


def export_summary(job_id: int, export_type: str, token: str, sleep_s: float) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORT_DIR / f"job{job_id}_{export_type}.xlsx"
    payload = json.dumps({"exportType": export_type, "useStrategyEngine": True})
    cmd = [
        "curl", "-sfS",
        "-H", f"Authorization: Bearer {token}",
        "-H", "Content-Type: application/json",
        "-d", payload,
        f"{API}/api/hospital-reconciliations/{job_id}/export-v2",
        "-o", str(out),
    ]
    docker_curl(cmd)
    time.sleep(sleep_s)
    return out


def resolve_job(hospital: str, jobs: dict[str, int]) -> int | None:
    if hospital in jobs:
        return jobs[hospital]
    for key, jid in jobs.items():
        if hospital in key or key in hospital:
            return jid
    return None


def compare_entry(entry: dict, jobs: dict[str, int], token: str, sleep_s: float) -> CompareResult:
    hospital = entry["hospital"]
    export_type = entry["exportType"]
    folder = TEST_CASE / hospital
    ref = find_reference(folder, entry["referenceGlob"], entry.get("referenceMonth", "6月"))
    job_id = resolve_job(hospital, jobs)
    key = f"{hospital}|{export_type}"

    if job_id is None:
        return CompareResult(hospital, export_type, "skip", None, str(ref) if ref else None,
                             None, None, None, "无 Job")
    if ref is None:
        return CompareResult(hospital, export_type, "skip", job_id, None,
                             None, None, None, "无参考表")

    try:
        actual = export_summary(job_id, export_type, token, sleep_s)
        total_exp = sheet_numeric_total(ref)
        total_act = sheet_numeric_total(actual)
        delta = abs(total_exp - total_act)
        tol = float(entry.get("tolerance", 0.01))
        if delta <= tol:
            status = "pass"
        elif key in KNOWN_SUMMARY_DIFF or entry.get("referenceMonthFallback"):
            status = "warn"
        else:
            status = "fail"
        detail = KNOWN_SUMMARY_DIFF.get(key, "")
        if entry.get("referenceMonthFallback"):
            detail = (detail + " · reference_month_fallback").strip(" ·")
        return CompareResult(hospital, export_type, status, job_id, str(ref),
                             total_exp, total_act, round(delta, 2), detail)
    except Exception as e:
        return CompareResult(hospital, export_type, "fail", job_id, str(ref),
                             None, None, None, str(e))


def write_reports(results: list[CompareResult], export_type_filter: str) -> None:
    payload = {
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "exportTypeFilter": export_type_filter,
        "results": [
            {
                "hospital": r.hospital,
                "exportType": r.export_type,
                "status": r.status,
                "jobId": r.job_id,
                "reference": r.reference,
                "totalExpected": r.total_exp,
                "totalActual": r.total_act,
                "delta": r.delta,
                "detail": r.detail,
            }
            for r in results
        ],
    }
    REPORT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = [
        "# S8 汇总导出比对摘要",
        "",
        f"- 过滤类型: `{export_type_filter}`",
        f"- 生成时间: {payload['generatedAt']}",
        "",
        "| 医院 | 类型 | 状态 | Job | Δ | 说明 |",
        "|------|------|------|-----|---|------|",
    ]
    for r in results:
        lines.append(
            f"| {r.hospital} | {r.export_type} | {r.status} | {r.job_id or '—'} | "
            f"{r.delta if r.delta is not None else '—'} | {r.detail or '—'} |"
        )
    REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--export-type",
        default="all",
        choices=["all", *SUMMARY_TYPES],
    )
    parser.add_argument("--export-sleep", type=float, default=1.5)
    args = parser.parse_args()

    entries = load_manifest()
    if args.export_type != "all":
        entries = [e for e in entries if e["exportType"] == args.export_type]

    jobs = parse_job_table()
    token = get_token()
    results = [compare_entry(e, jobs, token, args.export_sleep) for e in entries]
    write_reports(results, args.export_type)

    failed = sum(1 for r in results if r.status == "fail")
    print(json.dumps({"failed": failed, "total": len(results)}, ensure_ascii=False))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
