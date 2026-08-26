#!/usr/bin/env python3
"""生成22家特殊计价医院测试文件（每家正向+负向，共44份）"""
import math
import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "系统内置测试用例")

# ── 包装材料价格表 ──
PACK_PRICE = {10: 16.5, 15: 19.5, 20: 22.5, 25: 24.5, 30: 29.5}
ONE_PIECE_PRICE = {10: 22, 15: 25, 20: 28, 25: 30, 30: 35}
LOWTEMP_TIER = {
    1: 22, 2: 44, 3: 66, 4: 88, 5: 88, 6: 110, 7: 132, 8: 154,
    9: 165, 10: 165, 11: 187, 12: 209, 13: 231, 14: 253, 15: 253,
    16: 275, 17: 297, 18: 300, 19: 300, 20: 300, 21: 322, 22: 344,
    23: 366, 24: 388, 25: 388, 26: 410, 27: 432, 28: 454, 29: 465,
    30: 465, 31: 487, 32: 509, 33: 531, 34: 553, 35: 553, 36: 575,
    37: 597, 38: 600, 39: 600, 40: 600,
}

def ceil_div(n, d):
    return math.ceil(n / d)

def fold_price(count, bag_size, with_packaging=True):
    instrument = ceil_div(count, 5) * 5.5
    if with_packaging and count <= 10:
        return round(instrument + PACK_PRICE[bag_size], 2)
    return round(instrument, 2)

def fold_price_unit(count, unit_price, threshold, ratio, with_packaging=True, bag_size=20):
    """通用FOLD规则价格计算"""
    instrument = ceil_div(count, ratio) * unit_price
    if with_packaging and count <= threshold:
        return round(instrument + PACK_PRICE[bag_size], 2)
    return round(instrument, 2)

# ── 通用规则混合模式测试条目 ──
def generic_mix_cases(hospital_name):
    """为每家医院生成2-3条通用规则混合模式测试条目"""
    return [
        {"hospital": hospital_name, "dept": "混合测试", "type": "低温等离子灭菌",
         "packName": "克氏针-5件/Z7530", "material": "纸塑袋 20cm",
         "count": 5, "packCount": 1, "rule": "通用规则混合-克氏针5件",
         "price": fold_price(5, 20)},
        {"hospital": hospital_name, "dept": "混合测试", "type": "低温等离子灭菌",
         "packName": "测试器械包-9件", "material": "纸塑袋 20cm",
         "count": 9, "packCount": 1, "rule": "通用规则混合-低温9件阶梯",
         "price": float(LOWTEMP_TIER[9])},
        {"hospital": hospital_name, "dept": "混合测试", "type": "低温等离子灭菌",
         "packName": "软镜-1件", "material": "纸塑袋 20cm",
         "count": 1, "packCount": 1, "rule": "通用规则混合-软镜固定300",
         "price": 300.0},
    ]

# ════════════════════════════════════════════
# 22家医院特殊计价规则测试用例
# ════════════════════════════════════════════

def build_hospital_cases():
    """构建所有22家医院的测试用例"""
    hospitals = {}

    # 1. 冰城医美
    h = "冰城医美"
    hospitals[h] = []
    for name, cnt, fee, rule in [
        ("环钻包", 5, 3, "环钻包件数*5.5+3"),
        ("环钻包", 8, 3, "环钻包8件"),
        ("整形手术包", 6, 3, "整形手术包件数*5.5+3"),
        ("脂充包", 4, 5, "脂充包件数*5.5+5"),
        ("脂充包", 7, 5, "脂充包7件"),
    ]:
        hospitals[h].append({
            "hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"{name}-{cnt}件", "material": "无纺布",
            "count": cnt, "packCount": 1, "rule": rule,
            "price": round(cnt * 5.5 + fee, 2),
        })

    # 2. 电机厂医院
    h = "电机厂医院"
    hospitals[h] = []
    # 缝合针: 1*5.5+2.5=8
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "缝合针-3件/Z7530", "material": "纸塑袋 15cm",
        "count": 3, "packCount": 1, "rule": "电机厂缝合针=8", "price": 8.0})
    # 双<3件: 5.5*件数+标准包材
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "双爪钳-2件/Z3020", "material": "纸塑袋 20cm",
        "count": 2, "packCount": 1, "rule": "电机厂双<3件含标准包材",
        "price": round(2 * 5.5 + PACK_PRICE[20], 2)})
    # 双>=3件: 5.5*件数+内层纸塑袋标准包材
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "双爪钳-4件/Z3020", "material": "纸塑袋 20cm",
        "count": 4, "packCount": 1, "rule": "电机厂双>=3件含内层袋包材",
        "price": round(4 * 5.5 + PACK_PRICE[20], 2)})
    # 指针(≤10): ceil(count/5)*5.5+包材
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "指针-7件/Z7530", "material": "纸塑袋 20cm",
        "count": 7, "packCount": 1, "rule": "电机厂指针7件含包材",
        "price": fold_price(7, 20)})
    # 指针(>10): ceil(count/5)*5.5
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "指针-12件/Z7530", "material": "纸塑袋 20cm",
        "count": 12, "packCount": 1, "rule": "电机厂指针12件免包材",
        "price": fold_price(12, 20, False)})
    # 棉球W60=25, W90=30, W120=35
    for w, p in [("W60", 25), ("W90", 30), ("W120", 35)]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"棉球敷料包-{w}", "material": "无纺布",
            "count": 1, "packCount": 1, "rule": f"电机厂棉球{w}={p}", "price": float(p)})
    # 纱布W70=25, W150=35
    for w, p in [("W70", 25), ("W150", 35)]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"纱布敷料包-{w}", "material": "无纺布",
            "count": 1, "packCount": 1, "rule": f"电机厂纱布{w}={p}", "price": float(p)})

    # 3. 方南南医院
    h = "方南南医院"
    hospitals[h] = []
    for name in ["P钻", "根管锉", "光滑针", "机扩针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
                "packName": f"{name}-{cnt}件/Z7530", "material": "纸塑袋 20cm",
                "count": cnt, "packCount": 1,
                "rule": f"方南南{name}{cnt}件{'含包材' if wp else '免包材'}",
                "price": fold_price(cnt, 20, wp)})

    # 4. 东北农业大学
    h = "东北农业大学"
    hospitals[h] = []
    for name in ["根管针", "机锉", "牙探针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
                "packName": f"{name}-{cnt}件/Z7530", "material": "纸塑袋 20cm",
                "count": cnt, "packCount": 1,
                "rule": f"东北农大{name}{cnt}件{'含包材' if wp else '免包材'}",
                "price": fold_price(cnt, 20, wp)})
    # 根管锉(≤11): ceil(count/5)*5.6+包材; >11: ceil(count/5)*5.6
    for cnt in [6, 12]:
        wp = cnt <= 11
        inst = ceil_div(cnt, 5) * 5.6
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"根管锉-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"东北农大根管锉{cnt}件5.6单价{'含包材' if wp else '免包材'}",
            "price": p})

    # 5. 哈尔滨工程大学
    h = "哈尔滨工程大学"
    hospitals[h] = []
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "孔巾敷料包-25cm", "material": "纸塑袋 25cm",
        "count": 1, "packCount": 1, "rule": "工程大孔巾≥20cm=4", "price": 4.0})
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "孔巾敷料包-15cm", "material": "纸塑袋 15cm",
        "count": 1, "packCount": 1, "rule": "工程大孔巾<20cm=2.5", "price": 2.5})

    # 6. 松电慢病
    h = "松电慢病"
    hospitals[h] = []
    for cnt in [5, 12]:
        wp = cnt <= 10
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"机扩针-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"松电慢病机扩针{cnt}件{'含包材' if wp else '免包材'}",
            "price": fold_price(cnt, 20, wp)})

    # 7. 航天风华
    h = "航天风华"
    hospitals[h] = []
    for cnt in [5, 12]:
        wp = cnt <= 10
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"镍钛锉-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"航天风华镍钛锉{cnt}件{'含包材' if wp else '免包材'}",
            "price": fold_price(cnt, 20, wp)})

    # 8. 市五院（二门诊）
    h = "市五院（二门诊）"
    hospitals[h] = []
    for name, w, p in [("驱血带", "W50", 25), ("驱血带", "W60", 25), ("驱血带", "W90", 30), ("驱血带", "W120", 35), ("驱血带", "W150", 35)]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"{name}-{w}", "material": "无纺布",
            "count": 1, "packCount": 1, "rule": f"市五院二门诊{name}{w}={p}", "price": float(p)})

    # 9. 九州医院
    h = "九州医院"
    hospitals[h] = []
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "方盘-3件", "material": "纸塑袋 20cm",
        "count": 3, "packCount": 1, "rule": "九州方盘固定5.5", "price": 5.5})

    # 10. 博尚医院
    h = "博尚医院"
    hospitals[h] = []
    for name, price in [
        ("旋切器1戳卡1胶帽4/Z2045", 66),
        ("旋切器1胶帽4/Z2045", 44),
        ("旋切器1线1胶帽4/Z2045", 66),
    ]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": name, "material": "纸塑袋 20cm",
            "count": 6, "packCount": 1, "rule": f"博尚{name}固定{price}",
            "price": float(price)})

    # 11. 黑龙江省海员总医院（松北）
    h = "黑龙江省海员总医院（松北）"
    hospitals[h] = []
    # 胶帽≤5: 按1件, 低温标准包材价格
    for bag in [20, 30]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": f"胶帽-3件/Z2045", "material": f"纸塑袋 {bag}cm",
            "count": 3, "packCount": 1, "rule": f"海员胶帽≤5按1件{bag}cm={ONE_PIECE_PRICE[bag]}",
            "price": float(ONE_PIECE_PRICE[bag])})
    # 胶帽>5: ceil(count/5)*22
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "胶帽-7件/Z2045", "material": "纸塑袋 20cm",
        "count": 7, "packCount": 1, "rule": "海员胶帽>5: ceil(7/5)*22=44",
        "price": float(ceil_div(7, 5) * 22)})
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "胶帽-12件/Z2045", "material": "纸塑袋 20cm",
        "count": 12, "packCount": 1, "rule": "海员胶帽>5: ceil(12/5)*22=66",
        "price": float(ceil_div(12, 5) * 22)})

    # 12. 黑龙江省妇幼保健院（人口）
    h = "黑龙江省妇幼保健院（人口）"
    hospitals[h] = []
    # 密封胶圈等≤5: 按一件, 低温标准包材
    for bag in [20, 30]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": "密封胶圈-4件/Z2045", "material": f"纸塑袋 {bag}cm",
            "count": 4, "packCount": 1, "rule": f"妇幼密封胶圈≤5按1件{bag}cm={ONE_PIECE_PRICE[bag]}",
            "price": float(ONE_PIECE_PRICE[bag])})
    # >5: ceil(count/5)*22
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "密封胶圈-8件/Z2045", "material": "纸塑袋 20cm",
        "count": 8, "packCount": 1, "rule": "妇幼密封胶圈>5: ceil(8/5)*22=44",
        "price": float(ceil_div(8, 5) * 22)})
    # 根管锉, 针类组件, 机扩锉, 机扩针, 荡洗针, 加长锉, 洗髓针, 彩色锉, 手扩锉
    for name in ["根管锉", "机扩锉", "机扩针", "荡洗针", "加长锉", "洗髓针", "彩色锉", "手扩锉", "针类组件"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
                "packName": f"{name}-{cnt}件/Z7530", "material": "纸塑袋 20cm",
                "count": cnt, "packCount": 1,
                "rule": f"妇幼{name}{cnt}件{'含包材' if wp else '免包材'}",
                "price": fold_price(cnt, 20, wp)})
    # 针包: 1*5.5+包材
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
        "packName": "针包-3件/Z7530", "material": "纸塑袋 20cm",
        "count": 3, "packCount": 1, "rule": "妇幼针包按1件=5.5+22.5=28",
        "price": round(5.5 + PACK_PRICE[20], 2)})
    # 针多少盒1(≤6): [ceil(针/5)+1]*5.5+包材
    for cnt in [5, 8]:
        wp = cnt <= 6
        inst = (ceil_div(cnt, 5) + 1) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"针多少盒1-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"妇幼针多少盒1 {cnt}件{'含包材' if wp else '免包材'}",
            "price": p})
    # 新腹腔镜镜头: 标准收费+8元 (1件/z2060, 低温)
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "新腹腔镜镜头-1件/z2060", "material": "纸塑袋 20cm",
        "count": 1, "packCount": 1, "rule": "妇幼腹腔镜镜头标准+8=36",
        "price": round(ONE_PIECE_PRICE[20] + 8, 2)})

    # 13. 祖研南岗
    h = "祖研-黑龙江省中医医院（南岗院区）"
    hospitals[h] = []
    # 排针(≤20): ceil(count/10)*5.5+包材; >20: ceil(count/10)*5.5
    for cnt in [10, 15, 25, 30]:
        wp = cnt <= 20
        inst = ceil_div(cnt, 10) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"排针-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"祖研南岗排针{cnt}件/10{'含包材' if wp else '免包材'}",
            "price": p})

    # 14. 黑龙江省社会康复医院
    h = "黑龙江省社会康复医院"
    hospitals[h] = []
    fixed_items = [
        ("抛光车针盒6件盒1/Z1026", 16.5),
        ("环切器械盒-14件盒1/Z2032", 16.5),
        ("种植盒(黄绿)8件盒1/Z2032", 16.5),
        ("种植盒(金属)12件盒1/Z2032", 16.5),
        ("种植盒(蓝白)12件盒1/Z2032", 16.5),
        ("种植盒-14件盒1/W6050", 16.5),
        ("梁光强器械盒(粉)10件盒1/Z2030", 16.5),
        ("梁光强器械盒-21件盒1/W6050", 22),
        ("大车针盒-1/Z1526", 22),
        ("内提工具盒10件盒1/W6050", 44),
        ("百诺工具盒-19件盒1/W6050", 44),
        ("ITI种植盒-32件盒1/W12050", 44),
        ("登腾种植盒-34件盒1/W6050", 44),
        ("登腾种植盒-36件盒1/W6050", 44),
    ]
    for name, price in fixed_items:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": name, "material": "纸塑袋 20cm",
            "count": 1, "packCount": 1, "rule": f"康复医院{name}固定{price}",
            "price": float(price)})

    # 15. 哈尔滨市道里区妇幼保健院
    h = "哈尔滨市道里区妇幼保健院"
    hospitals[h] = []
    for name in ["棉花针", "洗髓针"]:
        for cnt in [5, 12]:
            wp = cnt <= 10
            hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
                "packName": f"{name}-{cnt}件/Z7530", "material": "纸塑袋 20cm",
                "count": cnt, "packCount": 1,
                "rule": f"道里妇幼{name}{cnt}件{'含包材' if wp else '免包材'}",
                "price": fold_price(cnt, 20, wp)})

    # 16. 春语医疗美容医院
    h = "春语医疗美容医院"
    hospitals[h] = []
    # 塑料管/管子≤10: 按一件, 标准低温包材
    for bag in [20, 30]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": "塑料管-5件", "material": f"纸塑袋 {bag}cm",
            "count": 5, "packCount": 1, "rule": f"春语塑料管≤10按1件{bag}cm={ONE_PIECE_PRICE[bag]}",
            "price": float(ONE_PIECE_PRICE[bag])})
    # >10: ceil(count/10)*22
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "管子-15件", "material": "纸塑袋 20cm",
        "count": 15, "packCount": 1, "rule": "春语管子>10: ceil(15/10)*22=44",
        "price": float(ceil_div(15, 10) * 22)})

    # 17. 黑龙江总工会医院
    h = "黑龙江总工会医院"
    hospitals[h] = []
    for name in ["12°镜头-1/Z2060", "30°镜头镜鞘-2/Z2060", "检查镜-2/Z2060",
                 "宫腔检查镜30度-2件/Z2060", "沈大12度镜头-1(新)/Z2060"]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": name, "material": "纸塑袋 20cm",
            "count": 1, "packCount": 1, "rule": f"总工会{name}标准+8={ONE_PIECE_PRICE[20]+8}",
            "price": round(ONE_PIECE_PRICE[20] + 8, 2)})

    # 18. 哈尔滨基准生物有限公司
    h = "哈尔滨基准生物有限公司"
    hospitals[h] = []
    for temp_type in ["低温等离子灭菌", "ETO"]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": temp_type,
            "packName": "氩氦刀-1件", "material": "纸塑袋 20cm",
            "count": 1, "packCount": 1, "rule": f"基准氩氦刀{temp_type}固定150",
            "price": 150.0})

    # 19. 索菲医疗美容门诊
    h = "索菲医疗美容门诊"
    hospitals[h] = []
    # 面吸针≥3: 件数*5.5 (双层袋不额外收费)
    for cnt in [3, 5, 8]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": f"面吸针-{cnt}件/Z2045", "material": "双层袋",
            "count": cnt, "packCount": 1, "rule": f"索菲面吸针{cnt}件={cnt*5.5}",
            "price": round(cnt * 5.5, 2)})

    # 20. 省监狱管理局医院
    h = "省监狱管理局医院"
    hospitals[h] = []
    # 密封件/密封胶圈≤5: 按一件, 低温标准包材
    for bag in [20, 30]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
            "packName": "密封件-3件/Z2045", "material": f"纸塑袋 {bag}cm",
            "count": 3, "packCount": 1, "rule": f"监狱密封件≤5按1件{bag}cm={ONE_PIECE_PRICE[bag]}",
            "price": float(ONE_PIECE_PRICE[bag])})
    # >5: ceil(count/5)*22
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "密封胶圈-8件/Z2045", "material": "纸塑袋 20cm",
        "count": 8, "packCount": 1, "rule": "监狱密封胶圈>5: ceil(8/5)*22=44",
        "price": float(ceil_div(8, 5) * 22)})

    # 21. 呼兰中医院
    h = "呼兰中医院"
    hospitals[h] = []
    # 铂康-阑尾包/外科包: 件数*5.5+13
    for name, cnt in [("铂康-阑尾包", 5), ("铂康-外科包", 8), ("外科包", 6)]:
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"{name}-{cnt}件", "material": "无纺布",
            "count": cnt, "packCount": 1, "rule": f"呼兰{name}={cnt}*5.5+13",
            "price": round(cnt * 5.5 + 13, 2)})
    # 胶帽≤5: 按一件, 低温标准包材
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "胶帽-3件/Z2045", "material": "纸塑袋 20cm",
        "count": 3, "packCount": 1, "rule": f"呼兰胶帽≤5按1件={ONE_PIECE_PRICE[20]}",
        "price": float(ONE_PIECE_PRICE[20])})
    # 胶帽>5: ceil(count/5)*22
    hospitals[h].append({"hospital": h, "dept": "测试科", "type": "低温等离子灭菌",
        "packName": "胶帽-8件/Z2045", "material": "纸塑袋 20cm",
        "count": 8, "packCount": 1, "rule": "呼兰胶帽>5: ceil(8/5)*22=44",
        "price": float(ceil_div(8, 5) * 22)})

    # 22. 平房区人民医院
    h = "平房区人民医院"
    hospitals[h] = []
    # 针盒1针多少(≤6): [ceil(针/5)+1]*5.5+包材; >6: [ceil(针/5)+1]*5.5
    for cnt in [5, 8]:
        wp = cnt <= 6
        inst = (ceil_div(cnt, 5) + 1) * 5.5
        p = round(inst + PACK_PRICE[20], 2) if wp else round(inst, 2)
        hospitals[h].append({"hospital": h, "dept": "测试科", "type": "高温灭菌",
            "packName": f"针盒1针多少-{cnt}件/Z7530", "material": "纸塑袋 20cm",
            "count": cnt, "packCount": 1,
            "rule": f"平房针多少{cnt}件{'含包材' if wp else '免包材'}",
            "price": p})

    return hospitals

# ── 生成Excel文件 ──
HEADERS = ["序号", "医院名称", "科室", "灭菌类型", "包名称", "包装材料",
           "器械数", "包数", "单价", "总价", "规则说明"]

def make_excel(cases, filepath, hospital_name, is_negative=False):
    error_indices = set()
    if is_negative:
        random.seed(hash(hospital_name) % 2**32)
        total = len(cases)
        error_count = max(2, total // 5)
        error_indices = set(random.sample(range(total), min(error_count, total)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    title = f"{hospital_name}{'负向' if is_negative else '正向'}测试"
    if is_negative:
        title += "（红色为故意标注的错误结果）"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    red_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, case in enumerate(cases):
        row = i + 3
        price = case["price"]
        is_error = i in error_indices
        if is_error:
            price = round(price + 5.0, 2)

        vals = [
            i + 1, case["hospital"], case["dept"], case["type"],
            case["packName"], case["material"], case["count"],
            case["packCount"], price, round(price * case["packCount"], 2),
            case["rule"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if col != 11 else "left")
            if is_error and col in (9, 10):
                c.font = red_font
                c.fill = red_fill

    widths = [6, 18, 10, 16, 32, 16, 8, 8, 10, 10, 32]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filepath)
    tag = "负向" if is_negative else "正向"
    err_count = len(error_indices)
    print(f"  {hospital_name} {tag}: {len(cases)}条用例, {err_count}条错误")

def main():
    hospitals = build_hospital_cases()
    print(f"共{len(hospitals)}家医院\n")

    for h_name, cases in hospitals.items():
        # 添加通用规则混合模式测试条目
        all_cases = cases + generic_mix_cases(h_name)

        safe_name = h_name.replace("/", "_").replace("（", "(").replace("）", ")")
        positive_path = os.path.join(OUTPUT_DIR, f"{safe_name}-正向测试.xlsx")
        negative_path = os.path.join(OUTPUT_DIR, f"{safe_name}-负向测试.xlsx")

        make_excel(all_cases, positive_path, h_name, is_negative=False)
        make_excel(all_cases, negative_path, h_name, is_negative=True)

    print(f"\n全部{len(hospitals)*2}份测试文件生成完成！")
    print(f"输出目录: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
