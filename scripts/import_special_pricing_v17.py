#!/usr/bin/env python3
"""Export 特殊收费(17).xlsx to structured JSON (merged-cell aware).

「包名称带"xxx"」→ match_mode=contains（关键词 xxx）
纯包名（如 环钻包）→ match_mode=exact（指定名称）
合并单元格的值向下/向右继承到同组每一行。
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs/source/特殊收费(17).xlsx"
DEFAULT_OUT = ROOT / "测试用例/.special_pricing_v17_extract.json"

HOSPITAL_TO_CODE: dict[str, str] = {
    "东北农业大学": "NEAU-YY",
    "九州医院": "JIUZHOU-FK",
    "冰城医美": "BINGCHENG-YM",
    "博尚医院": "BOSHANG-YY",
    "哈尔滨基准生物有限公司": "JZSW-BIO",
    "哈尔滨市道里区妇幼保健院": "DL-FUCHAN",
    "市五院（二门诊）": "HRB-WY-EM",
    "方南南医院": "FNN-YY",
    "春语医疗美容医院": "CHUNYU-YL",
    "松电慢病": "HRB-SD-MB",
    "电机厂医院": "GUOYAO-2",
    "省监狱管理局医院": "HLJ-JYGLJ-YY",
    "祖研-黑龙江省中医医院（南岗院区）": "ZUYAN-NG",
    "祖研-黑龙江省中医医院（三辅院区）": "ZUYAN-SF",
    "索菲医疗美容门诊": "SUOFEI-YL",
    "航天风华": "HRB-HTFH",
    "黑龙江总工会医院": "HL-ZGH",
    "黑龙江省妇幼保健院（人口）": "HLJ-FY-RK",
    "黑龙江省海员总医院（松北）": "HAIYUAN-SB",
    "黑龙江省社会康复医院": "SHKF-YY",
    "呼兰中医院": "HULAN-TCM",
    "呼兰区第一人民医院": "HULAN-RM",
    "哈尔滨市第五医院": "HRB-WY",
    "平房区人民医院": "PFQ-RM",
    "新发红十字医院": "XINFA-HSZ",
    "黑龙江省远东心脑血管医院": "YUANDONG-XN",
}


def build_merged_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int, int], object]:
    merged: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = val
    return merged


def cell(ws, merged: dict, row: int, col: int):
    return merged.get((row, col), ws.cell(row, col).value)


def norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_pack_name(text: str) -> tuple[str, str | None]:
    """Return (match_mode, keyword)."""
    text = norm(text)
    if not text:
        return "unknown", None
    m = re.search(r'包名称带[“"](.+?)[”"]', text)
    if m:
        return "contains", m.group(1)
    if "通用" in text:
        return "generic", None
    return "exact", text


def parse_hospital_sheet(ws) -> list[dict]:
    merged = build_merged_map(ws)
    state = {
        "hospital_seq": "",
        "hospital": "",
        "item": "",
        "pack_name": "",
        "pack_type": "",
        "note": "",
    }
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if norm(cell(ws, merged, r, 1)):
            state["hospital_seq"] = norm(cell(ws, merged, r, 1))
        if norm(cell(ws, merged, r, 2)):
            state["hospital"] = norm(cell(ws, merged, r, 2))
        if norm(cell(ws, merged, r, 3)):
            state["item"] = norm(cell(ws, merged, r, 3))
        if norm(cell(ws, merged, r, 4)):
            state["pack_name"] = norm(cell(ws, merged, r, 4))
        if norm(cell(ws, merged, r, 5)):
            state["pack_type"] = norm(cell(ws, merged, r, 5))
        if norm(cell(ws, merged, r, 8)):
            state["note"] = norm(cell(ws, merged, r, 8))

        inst = norm(cell(ws, merged, r, 6))
        rule = norm(cell(ws, merged, r, 7))
        if not rule and not inst:
            continue

        mode, keyword = parse_pack_name(state["pack_name"])
        rows.append(
            {
                "row": r,
                "hospital_seq": state["hospital_seq"],
                "hospital": state["hospital"],
                "customer_code": HOSPITAL_TO_CODE.get(state["hospital"]),
                "item": state["item"],
                "pack_name": state["pack_name"],
                "match_mode": mode,
                "keyword": keyword,
                "pack_type": state["pack_type"],
                "instrument_count": inst or None,
                "rule": rule,
                "note": state["note"] or None,
            }
        )
    return rows


def parse_generic_sheet(ws) -> list[dict]:
    merged = build_merged_map(ws)
    state = {"seq": "", "pack_name": "", "pack_type": "", "note": ""}
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if norm(cell(ws, merged, r, 1)):
            state["seq"] = norm(cell(ws, merged, r, 1))
        if norm(cell(ws, merged, r, 2)):
            state["pack_name"] = norm(cell(ws, merged, r, 2))
        if norm(cell(ws, merged, r, 3)):
            state["pack_type"] = norm(cell(ws, merged, r, 3))
        if norm(cell(ws, merged, r, 6)):
            state["note"] = norm(cell(ws, merged, r, 6))

        inst = norm(cell(ws, merged, r, 4))
        rule = norm(cell(ws, merged, r, 5))
        if not rule and not inst:
            continue
        mode, keyword = parse_pack_name(state["pack_name"])
        rows.append(
            {
                "row": r,
                "seq": state["seq"],
                "pack_name": state["pack_name"],
                "match_mode": mode,
                "keyword": keyword,
                "pack_type": state["pack_type"],
                "instrument_count": inst or None,
                "rule": rule,
                "note": state["note"] or None,
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise FileNotFoundError(args.xlsx)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    hospital = parse_hospital_sheet(wb["各医院特殊收费"])
    generic = parse_generic_sheet(wb["通用特殊收费"])

    payload = {
        "source": str(args.xlsx),
        "hospital_rows": hospital,
        "generic_rows": generic,
        "hospital_to_code": HOSPITAL_TO_CODE,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    hospitals = sorted({r["hospital"] for r in hospital if r["hospital"]})
    print(f"Wrote {args.out} ({len(hospital)} hospital rows, {len(generic)} generic rows, {len(hospitals)} hospitals)")


if __name__ == "__main__":
    main()
